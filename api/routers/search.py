"""
Router de Búsqueda live en OpenAlex.
Permite buscar publicaciones en la API de OpenAlex sin ingesta.
"""

import io
import logging
from datetime import datetime as _dt
from typing import Optional, List

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from fastapi import APIRouter, File, Query, HTTPException, UploadFile
from starlette.responses import StreamingResponse

from config import openalex_config, institution
from extractors.base import normalize_doi

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/search", tags=["Búsqueda en Vivo"])


# ─────────────────────────────────────────────────────────────────────────────
# Definición de columnas del Excel de enriquecimiento
# (key_interna, label_encabezado, ancho_columna)
# ─────────────────────────────────────────────────────────────────────────────

_FOUND_COLS: list[tuple[str, str, int]] = [
    # Datos aportados por el usuario
    ("titulo",                    "Título (input)",          50),
    ("año",                       "Año (input)",              8),
    ("doi",                       "DOI (input)",             36),
    # Método de localización + afiliación
    ("oa_metodo",                 "Método búsqueda",          16),
    ("oa_autor_institucional",    "Autor de la U",           14),
    ("oa_autores_institucionales","Autores de la U",         46),
    # Revista e ISSN  ← prioridad
    ("oa_revista",                "Revista",                 40),
    ("oa_issn",                   "ISSN-L",                  16),
    ("oa_issn_todos",             "Todos los ISSN",          28),
    ("oa_editorial",              "Editorial",               28),
    # Datos de la publicación
    ("oa_titulo",                 "Título en OpenAlex",      50),
    ("oa_año",                    "Año (OA)",                8),
    ("oa_doi",                    "DOI (OA)",                36),
    ("oa_tipo",                   "Tipo",                    18),
    ("oa_idioma",                 "Idioma",                  10),
    # Acceso abierto
    ("oa_open_access",            "Acceso abierto",          16),
    ("oa_status_oa",              "Estado OA",               16),
    # Métricas
    ("oa_citas",                  "Citas",                   10),
    # Enlace
    ("oa_url",                    "URL",                     40),
    # Autores
    ("oa_autores",                "Autores",                 52),
    # ID OpenAlex (al final, referencia técnica)
    ("oa_work_id",                "Work ID (OA)",            32),
]

_NOT_FOUND_COLS: list[tuple[str, str, int]] = [
    ("titulo", "Título",  52),
    ("año",    "Año",       8),
    ("doi",    "DOI",      40),
]

# Colores de encabezado por grupo de columna
_COL_HEADER_COLOR: dict[str, str] = {
    "titulo":                     "1F4E79",
    "año":                        "1F4E79",
    "doi":                        "1F4E79",
    "oa_metodo":                  "2E75B6",
    "oa_autor_institucional":     "7030A0",
    "oa_autores_institucionales": "7030A0",
    "oa_revista":                 "375623",
    "oa_issn":                    "375623",
    "oa_issn_todos":              "375623",
    "oa_editorial":               "375623",
}
_DEFAULT_HDR_COLOR = "16537E"


def _build_enrich_excel(
    found: list[dict],
    verify_inst: list[dict],
    found_no_inst: list[dict],
    verify: list[dict],
    not_found: list[dict],
    total: int,
) -> io.BytesIO:
    """Genera un Excel de 5 hojas: Encontrados (U), Revisar inst. U, Sin afiliación U, Verificación manual, No encontrados."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl no disponible")

    wb = openpyxl.Workbook()
    now_str = _dt.now().strftime("%d/%m/%Y %H:%M")

    def _fill(color: str) -> PatternFill:
        return PatternFill(fill_type="solid", fgColor=color)

    def _font(bold=False, color="000000", size=10) -> Font:
        return Font(bold=bold, color=color, size=size)

    def _center(wrap=False) -> Alignment:
        return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

    def _left(wrap=False) -> Alignment:
        return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

    _side  = Side(style="thin", color="BFBFBF")
    _border = Border(left=_side, right=_side, top=_side, bottom=_side)

    # ── HOJA 1: Encontrados ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = f"Encontrados ({len(found)})"
    ncols1 = len(_FOUND_COLS)

    # Fila 1 — título del reporte
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols1)
    t = ws1.cell(row=1, column=1)
    t.value     = f"Enriquecimiento OpenAlex  ·  {len(found)} encontrados de {total}  ·  {now_str}"
    t.font      = _font(bold=True, color="FFFFFF", size=12)
    t.fill      = _fill("1F4E79")
    t.alignment = _center()
    ws1.row_dimensions[1].height = 24

    # Fila 2 — leyenda de grupos de columnas
    ws1.row_dimensions[2].height = 16
    _group_ranges = [
        (1,  3,  "DATOS ORIGINALES",   "1F4E79"),
        (4,  4,  "BÚSQL./CONFIANZA",   "2E75B6"),
        (5,  6,  "AFILIACIÓN U",       "7030A0"),
        (7,  10, "REVISTA / ISSN",     "375623"),
        (11, 17, "PUBLICACIÓN",        "16537E"),
        (18, 19, "ACCESO ABIERTO",     "7030A0"),
        (20, 20, "MÉTRICAS",           "833C00"),
        (21, 21, "ENLACE",             "595959"),
        (22, 22, "AUTORES",            "404040"),
        (23, 23, "ID TÉCNICO",         "808080"),
    ]
    for c1, c2, label, color in _group_ranges:
        if c1 == c2:
            ws1.cell(row=2, column=c1).value = label
        else:
            ws1.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
            ws1.cell(row=2, column=c1).value = label
        for ci in range(c1, c2 + 1):
            cell = ws1.cell(row=2, column=ci)
            cell.fill      = _fill(color)
            cell.font      = _font(bold=True, color="FFFFFF", size=9)
            cell.alignment = _center()
            cell.border    = _border

    # Fila 3 — encabezados de columnas
    for ci, (key, label, _) in enumerate(_FOUND_COLS, start=1):
        cell = ws1.cell(row=3, column=ci, value=label)
        cell.fill      = _fill(_COL_HEADER_COLOR.get(key, _DEFAULT_HDR_COLOR))
        cell.font      = _font(bold=True, color="FFFFFF", size=10)
        cell.alignment = _center(wrap=True)
        cell.border    = _border
    ws1.row_dimensions[3].height = 36

    # Filas de datos
    _ISSN_FILL     = _fill("E2EFDA")   # verde claro para ISSN
    _DATA_FILL     = _fill("EBF3FB")   # azul muy claro para el resto
    _OA_YES_FILL   = _fill("C6EFCE")   # verde acceso abierto
    _OA_NO_FILL    = _fill("FFCCCC")   # rojo sin acceso abierto
    _INST_YES_FILL = _fill("C6EFCE")   # verde institucional
    _INST_NO_FILL  = _fill("FFCCCC")   # rojo sin afiliación
    _ISSN_BOLD     = {"oa_issn", "oa_issn_todos", "oa_revista"}
    _WRAP_KEYS     = {"titulo", "oa_titulo", "oa_autores", "oa_issn_todos", "oa_autores_institucionales"}

    for ri, row in enumerate(found, start=4):
        for ci, (key, _, _) in enumerate(_FOUND_COLS, start=1):
            val = row.get(key, "")
            if isinstance(val, bool):
                val = "Sí" if val else "No"
            cell = ws1.cell(row=ri, column=ci, value=(val if val is not None else ""))
            cell.border    = _border
            cell.alignment = _left(wrap=(key in _WRAP_KEYS))

            if key in ("oa_issn", "oa_issn_todos", "oa_revista", "oa_editorial"):
                cell.fill = _ISSN_FILL
                cell.font = _font(bold=(key in _ISSN_BOLD), size=10)
            elif key == "oa_open_access":
                cell.fill = _OA_YES_FILL if val == "Sí" else _OA_NO_FILL
                cell.font = _font(bold=True, size=10)
            elif key == "oa_autor_institucional":
                cell.fill = _INST_YES_FILL if val == "Sí" else _INST_NO_FILL
                cell.font = _font(bold=True, size=10)
            else:
                cell.fill = _DATA_FILL
        ws1.row_dimensions[ri].height = 15

    # Anchos de columna
    for ci, (_, _, width) in enumerate(_FOUND_COLS, start=1):
        ws1.column_dimensions[get_column_letter(ci)].width = width

    ws1.freeze_panes = "A4"
    ws1.auto_filter.ref = f"A3:{get_column_letter(ncols1)}3"

    # ── HOJA 2: Revisión de institucionalidad ────────────────────────────────
    ws_vi = wb.create_sheet(title=f"Revisar inst. U ({len(verify_inst)})")
    ws_vi.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols1)
    t_vi = ws_vi.cell(row=1, column=1)
    t_vi.value     = (
        f"Revisar inst. U  ·  {len(verify_inst)} registros  ·  "
        f"Algunos autores con afiliación U (≤50%) — revisar manualmente  ·  {now_str}"
    )
    t_vi.font      = _font(bold=True, color="FFFFFF", size=12)
    t_vi.fill      = _fill("7030A0")   # púrpura
    t_vi.alignment = _center()
    ws_vi.row_dimensions[1].height = 24

    for c1, c2, label, color in _group_ranges:
        if c1 == c2:
            ws_vi.cell(row=2, column=c1).value = label
        else:
            ws_vi.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
            ws_vi.cell(row=2, column=c1).value = label
        for ci in range(c1, c2 + 1):
            cell = ws_vi.cell(row=2, column=ci)
            cell.fill      = _fill(color)
            cell.font      = _font(bold=True, color="FFFFFF", size=9)
            cell.alignment = _center()
            cell.border    = _border
    ws_vi.row_dimensions[2].height = 16

    for ci, (key, label, _) in enumerate(_FOUND_COLS, start=1):
        cell = ws_vi.cell(row=3, column=ci, value=label)
        cell.fill      = _fill(_COL_HEADER_COLOR.get(key, _DEFAULT_HDR_COLOR))
        cell.font      = _font(bold=True, color="FFFFFF", size=10)
        cell.alignment = _center(wrap=True)
        cell.border    = _border
    ws_vi.row_dimensions[3].height = 36

    _VI_FILL = _fill("EAD1DC")   # lavanda rosado — revisión de institucionalidad
    for ri, row in enumerate(verify_inst, start=4):
        for ci, (key, _, _) in enumerate(_FOUND_COLS, start=1):
            val = row.get(key, "")
            if isinstance(val, bool):
                val = "Sí" if val else "No"
            elif key == "oa_autor_institucional" and val == "verificar":
                val = "Revisar"
            cell = ws_vi.cell(row=ri, column=ci, value=(val if val is not None else ""))
            cell.border    = _border
            cell.alignment = _left(wrap=(key in _WRAP_KEYS))
            if key in ("oa_issn", "oa_issn_todos", "oa_revista", "oa_editorial"):
                cell.fill = _ISSN_FILL
                cell.font = _font(bold=(key in _ISSN_BOLD), size=10)
            elif key == "oa_open_access":
                cell.fill = _OA_YES_FILL if val == "Sí" else _OA_NO_FILL
                cell.font = _font(bold=True, size=10)
            elif key == "oa_autor_institucional":
                cell.fill = _fill("FFEB9C")   # amarillo — en revisión
                cell.font = _font(bold=True, size=10)
            else:
                cell.fill = _VI_FILL
        ws_vi.row_dimensions[ri].height = 15

    for ci, (_, _, width) in enumerate(_FOUND_COLS, start=1):
        ws_vi.column_dimensions[get_column_letter(ci)].width = width
    ws_vi.freeze_panes = "A4"
    ws_vi.auto_filter.ref = f"A3:{get_column_letter(ncols1)}3"

    # ── HOJA 3: Sin afiliación institucional ─────────────────────────────────
    ws_ni = wb.create_sheet(title=f"Sin afiliación U ({len(found_no_inst)})")
    ws_ni.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols1)
    t_ni = ws_ni.cell(row=1, column=1)
    t_ni.value     = (
        f"Sin autor de la U  ·  {len(found_no_inst)} registros  ·  "
        f"Encontrados en OpenAlex pero sin afiliación institucional  ·  {now_str}"
    )
    t_ni.font      = _font(bold=True, color="FFFFFF", size=12)
    t_ni.fill      = _fill("843C0C")   # rojo-ladrillo
    t_ni.alignment = _center()
    ws_ni.row_dimensions[1].height = 24

    for c1, c2, label, color in _group_ranges:
        if c1 == c2:
            ws_ni.cell(row=2, column=c1).value = label
        else:
            ws_ni.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
            ws_ni.cell(row=2, column=c1).value = label
        for ci in range(c1, c2 + 1):
            cell = ws_ni.cell(row=2, column=ci)
            cell.fill      = _fill(color)
            cell.font      = _font(bold=True, color="FFFFFF", size=9)
            cell.alignment = _center()
            cell.border    = _border
    ws_ni.row_dimensions[2].height = 16

    for ci, (key, label, _) in enumerate(_FOUND_COLS, start=1):
        cell = ws_ni.cell(row=3, column=ci, value=label)
        cell.fill      = _fill(_COL_HEADER_COLOR.get(key, _DEFAULT_HDR_COLOR))
        cell.font      = _font(bold=True, color="FFFFFF", size=10)
        cell.alignment = _center(wrap=True)
        cell.border    = _border
    ws_ni.row_dimensions[3].height = 36

    _NI_FILL = _fill("FCE4D6")   # salmón claro
    for ri, row in enumerate(found_no_inst, start=4):
        for ci, (key, _, _) in enumerate(_FOUND_COLS, start=1):
            val = row.get(key, "")
            if isinstance(val, bool):
                val = "Sí" if val else "No"
            cell = ws_ni.cell(row=ri, column=ci, value=(val if val is not None else ""))
            cell.border    = _border
            cell.alignment = _left(wrap=(key in _WRAP_KEYS))
            if key in ("oa_issn", "oa_issn_todos", "oa_revista", "oa_editorial"):
                cell.fill = _ISSN_FILL
                cell.font = _font(bold=(key in _ISSN_BOLD), size=10)
            elif key == "oa_open_access":
                cell.fill = _OA_YES_FILL if val == "Sí" else _OA_NO_FILL
                cell.font = _font(bold=True, size=10)
            elif key == "oa_autor_institucional":
                cell.fill = _INST_NO_FILL
                cell.font = _font(bold=True, size=10)
            else:
                cell.fill = _NI_FILL
        ws_ni.row_dimensions[ri].height = 15

    for ci, (_, _, width) in enumerate(_FOUND_COLS, start=1):
        ws_ni.column_dimensions[get_column_letter(ci)].width = width
    ws_ni.freeze_panes = "A4"
    ws_ni.auto_filter.ref = f"A3:{get_column_letter(ncols1)}3"

    # ── HOJA 4: Verificación manual ──────────────────────────────────────────
    ws_v = wb.create_sheet(title=f"Verificación manual ({len(verify)})")
    ws_v.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols1)
    t_v = ws_v.cell(row=1, column=1)
    t_v.value     = (
        f"Verificación manual  ·  {len(verify)} registros  ·  "
        f"Coincidencia parcial (año u otros datos difieren)  ·  {now_str}"
    )
    t_v.font      = _font(bold=True, color="FFFFFF", size=12)
    t_v.fill      = _fill("7F3F00")   # naranja oscuro
    t_v.alignment = _center()
    ws_v.row_dimensions[1].height = 24

    for c1, c2, label, color in _group_ranges:
        if c1 == c2:
            ws_v.cell(row=2, column=c1).value = label
        else:
            ws_v.merge_cells(start_row=2, start_column=c1, end_row=2, end_column=c2)
            ws_v.cell(row=2, column=c1).value = label
        for ci in range(c1, c2 + 1):
            cell = ws_v.cell(row=2, column=ci)
            cell.fill      = _fill(color)
            cell.font      = _font(bold=True, color="FFFFFF", size=9)
            cell.alignment = _center()
            cell.border    = _border
    ws_v.row_dimensions[2].height = 16

    for ci, (key, label, _) in enumerate(_FOUND_COLS, start=1):
        cell = ws_v.cell(row=3, column=ci, value=label)
        cell.fill      = _fill(_COL_HEADER_COLOR.get(key, _DEFAULT_HDR_COLOR))
        cell.font      = _font(bold=True, color="FFFFFF", size=10)
        cell.alignment = _center(wrap=True)
        cell.border    = _border
    ws_v.row_dimensions[3].height = 36

    _V_FILL = _fill("FFF2CC")   # amarillo claro
    for ri, row in enumerate(verify, start=4):
        for ci, (key, _, _) in enumerate(_FOUND_COLS, start=1):
            val = row.get(key, "")
            if isinstance(val, bool):
                val = "Sí" if val else "No"
            elif key == "oa_autor_institucional" and val == "verificar":
                val = "Revisar"
            cell = ws_v.cell(row=ri, column=ci, value=(val if val is not None else ""))
            cell.border    = _border
            cell.alignment = _left(wrap=(key in _WRAP_KEYS))
            if key in ("oa_issn", "oa_issn_todos", "oa_revista", "oa_editorial"):
                cell.fill = _ISSN_FILL
                cell.font = _font(bold=(key in _ISSN_BOLD), size=10)
            elif key == "oa_open_access":
                cell.fill = _OA_YES_FILL if val == "Sí" else _OA_NO_FILL
                cell.font = _font(bold=True, size=10)
            elif key == "oa_autor_institucional":
                if val == "Sí":
                    cell.fill = _INST_YES_FILL
                elif val == "Revisar":
                    cell.fill = _fill("FFEB9C")   # amarillo — en revisión
                else:
                    cell.fill = _INST_NO_FILL
                cell.font = _font(bold=True, size=10)
            else:
                cell.fill = _V_FILL
        ws_v.row_dimensions[ri].height = 15

    for ci, (_, _, width) in enumerate(_FOUND_COLS, start=1):
        ws_v.column_dimensions[get_column_letter(ci)].width = width
    ws_v.freeze_panes = "A4"
    ws_v.auto_filter.ref = f"A3:{get_column_letter(ncols1)}3"

    # ── HOJA 5: No encontrados ───────────────────────────────────────────────
    ws2 = wb.create_sheet(title=f"No encontrados ({len(not_found)})")
    ncols2 = len(_NOT_FOUND_COLS) + 1  # +1 para la columna Nota

    # Fila 1 — título
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols2)
    t2 = ws2.cell(row=1, column=1)
    t2.value     = f"No encontrados en OpenAlex  ·  {len(not_found)} de {total}  ·  {now_str}"
    t2.font      = _font(bold=True, color="FFFFFF", size=12)
    t2.fill      = _fill("595959")
    t2.alignment = _center()
    ws2.row_dimensions[1].height = 24

    # Fila 2 — encabezados
    nf_cols_ext = _NOT_FOUND_COLS + [("_nota", "Motivo / Nota", 50)]
    for ci, (_, label, _) in enumerate(nf_cols_ext, start=1):
        cell = ws2.cell(row=2, column=ci, value=label)
        cell.fill      = _fill("404040")
        cell.font      = _font(bold=True, color="FFFFFF", size=10)
        cell.alignment = _center(wrap=True)
        cell.border    = _border
    ws2.row_dimensions[2].height = 28

    # Filas de datos
    _NF_FILL = _fill("FFF2CC")   # amarillo claro
    for ri, row in enumerate(not_found, start=3):
        has_doi    = bool(str(row.get("doi",    "")).strip())
        has_title  = bool(str(row.get("titulo", "")).strip())
        has_issn   = bool(str(row.get("issn",   "")).strip())
        has_rev    = bool(str(row.get("revista","")).strip())
        if has_doi:
            nota = "DOI consultado en OpenAlex — no encontrado"
        elif has_issn:
            nota = "Búsqueda por ISSN + título (6 etapas) — sin coincidencia suficiente"
        elif has_rev:
            nota = "Búsqueda por nombre de revista + título — sin coincidencia suficiente"
        elif has_title:
            nota = "Búsqueda por título (6 etapas) — similitud insuficiente o sin resultados"
        else:
            nota = "Sin título ni DOI — no se pudo buscar"

        for ci, (key, _, _) in enumerate(_NOT_FOUND_COLS, start=1):
            val = row.get(key, "")
            cell = ws2.cell(row=ri, column=ci, value=(val if val is not None else ""))
            cell.fill      = _NF_FILL
            cell.alignment = _left(wrap=(key == "titulo"))
            cell.border    = _border

        nota_cell = ws2.cell(row=ri, column=len(_NOT_FOUND_COLS) + 1, value=nota)
        nota_cell.fill      = _NF_FILL
        nota_cell.alignment = _left(wrap=True)
        nota_cell.border    = _border
        ws2.row_dimensions[ri].height = 15

    for ci, (_, _, width) in enumerate(nf_cols_ext, start=1):
        ws2.column_dimensions[get_column_letter(ci)].width = width

    ws2.freeze_panes = "A3"

    # ── Guardar en buffer ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _get_session() -> requests.Session:
    """Sesión HTTP con retry. No reintenta en 429 (rate-limit debe respetarse)."""
    s = requests.Session()
    retries = Retry(total=3, backoff_factor=0.5, status_forcelist=[500, 502, 503])
    s.mount("https://", HTTPAdapter(max_retries=retries))
    return s


# ── GET /search/openalex ─────────────────────────────────────

@router.get("/openalex", summary="Buscar en OpenAlex", tags=["OpenAlex"])
def search_openalex(
    query: Optional[str] = Query(None, description="Texto libre de búsqueda"),
    doi: Optional[str] = Query(None),
    title: Optional[str] = Query(None),
    author: Optional[str] = Query(None),
    year_from: Optional[int] = Query(None),
    year_to: Optional[int] = Query(None),
    max_results: int = Query(25, ge=1, le=200),
):
    """
    Búsqueda en vivo en la API de OpenAlex.
    Retorna publicaciones con metadatos y autores institucionales.
    """
    params = {
        "per_page": min(max_results, 200),
        "mailto": institution.contact_email,
    }
    if openalex_config.api_key:
        params["api_key"] = openalex_config.api_key

    # Construir filtros
    filters = []
    if doi:
        ndoi = normalize_doi(doi)
        if ndoi:
            filters.append(f"doi:https://doi.org/{ndoi}")
    if year_from and year_to:
        filters.append(f"from_publication_date:{year_from}-01-01")
        filters.append(f"to_publication_date:{year_to}-12-31")
    elif year_from:
        filters.append(f"from_publication_date:{year_from}-01-01")
    elif year_to:
        filters.append(f"to_publication_date:{year_to}-12-31")

    if filters:
        params["filter"] = ",".join(filters)

    # Búsqueda por texto
    search_parts = []
    if query:
        search_parts.append(query)
    if title:
        search_parts.append(title)
    if author:
        search_parts.append(author)
    if search_parts:
        params["search"] = " ".join(search_parts)

    if not params.get("search") and not params.get("filter"):
        raise HTTPException(400, "Debe proporcionar al menos un criterio de búsqueda")

    try:
        session = _get_session()
        resp = session.get(openalex_config.base_url, params=params, timeout=openalex_config.timeout)
    except requests.RequestException as e:
        logger.error(f"Error de red consultando OpenAlex: {e}")
        raise HTTPException(502, f"Error al consultar OpenAlex: {e}")

    if resp.status_code == 429:
        retry_after = resp.headers.get("Retry-After", "desconocido")
        try:
            body = resp.json()
            msg = body.get("message") or "Límite de créditos diarios agotado."
        except Exception:
            msg = "Límite d e solicitudes superado en OpenAlex."
        logger.warning(f"OpenAlex rate-limit. Retry-After: {retry_after}s. {msg}")
        raise HTTPException(
            status_code=429,
            detail={
                "error": "rate_limit",
                "message": msg,
                "retry_after_seconds": retry_after,
                "hint": "Configure OA_KEY en el .env para mayor cuota, o espere al reset de medianoche UTC.",
            },
        )

    try:
        resp.raise_for_status()
        data = resp.json()
    except requests.HTTPError as e:
        logger.error(f"HTTP error desde OpenAlex: {e}")
        raise HTTPException(502, f"Error al consultar OpenAlex: {e}")
    except Exception as e:
        logger.error(f"Error parseando respuesta OpenAlex: {e}")
        raise HTTPException(502, "Respuesta inesperada de OpenAlex")

    results = data.get("results", [])
    ror_id = institution.ror_id

    output = []
    for work in results:
        # Identificar autores institucionales
        inst_authors = []
        all_authors = []

        for authorship in work.get("authorships", []):
            author_info = authorship.get("author", {})
            author_name = author_info.get("display_name", "Desconocido")
            author_orcid = author_info.get("orcid")
            openalex_id = author_info.get("id", "")

            is_inst = False
            for inst in authorship.get("institutions", []):
                if inst.get("ror") == ror_id:
                    is_inst = True
                    break

            entry = {
                "name": author_name,
                "orcid": author_orcid,
                "openalex_id": openalex_id,
                "is_institutional": is_inst,
            }
            all_authors.append(entry)
            if is_inst:
                inst_authors.append(entry)

        # Fuente / landing page
        primary_loc = work.get("primary_location") or {}
        source_info = primary_loc.get("source") or {}
        landing_url = primary_loc.get("landing_page_url")

        oa_info = work.get("open_access", {})

        output.append({
            "openalex_id": work.get("id", ""),
            "doi": work.get("doi"),
            "title": work.get("title", ""),
            "publication_year": work.get("publication_year"),
            "publication_type": work.get("type"),
            "cited_by_count": work.get("cited_by_count", 0),
            "is_open_access": oa_info.get("is_oa", False),
            "oa_status": oa_info.get("oa_status"),
            "source_journal": source_info.get("display_name"),
            "issn": source_info.get("issn_l"),
            "landing_page_url": landing_url or work.get("doi"),
            "all_authors": all_authors,
            "institutional_authors": inst_authors,
            "institutional_authors_count": len(inst_authors),
        })

    return {
        "count": data.get("meta", {}).get("count", len(output)),
        "results": output,
    }


# ── POST /search/enrich-excel ─────────────────────────────────────────────────

@router.post(
    "/enrich-excel",
    summary="Enriquecer Excel con OpenAlex",
    tags=["OpenAlex"],
    response_description="Excel (.xlsx) con hojas 'Encontrados' y 'No encontrados'",
)
async def enrich_excel_openalex(
    file: UploadFile = File(
        ...,
        description="Archivo Excel (.xlsx) con columnas: Título, Año, doi",
    ),
    fuzzy_threshold: int = Query(
        80, ge=50, le=100,
        description="Umbral mínimo de similitud (%) para búsqueda por título cuando no hay DOI",
    ),
):
    """
    Recibe un Excel con columnas **Título**, **Año** y **doi**, busca cada
    publicación en OpenAlex y devuelve un Excel con dos hojas:

    - **Encontrados** — Datos enriquecidos: revista, **ISSN-L**, **todos los ISSN**,
      tipo, editorial, acceso abierto, citas, autores, URL, etc.
    - **No encontrados** — Publicaciones que OpenAlex no pudo resolver, con
      una nota explicando el motivo.

    **Estrategia de búsqueda (6 etapas en cascada):**
    1. **DOI** — consulta exacta por lotes (hasta 25 DOIs por request).
    2. **Título + año** — fallback para DOIs no resueltos; 4 reintentos
       (título original, normalizado, sin año, truncado).
    3. **Título + año** — igual para registros sin DOI.
    4. **ISSN + título** — filtra por `primary_location.source.issn` **y**
       `locations.source.issn`; admite múltiples ISSNs separados por `;`.
    5. **Nombre de revista + título** — cuando hay columna `revista` pero no ISSN.
    6. **Solo título (umbral alto 88%)** — último recurso para los restantes.
    """
    fname = (file.filename or "").lower()
    if not fname.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "El archivo debe ser .xlsx o .xls")

    data = await file.read()
    if not data:
        raise HTTPException(400, "El archivo está vacío")

    try:
        from extractors.openalex import OpenAlexEnricher, OpenAlexRateLimitError
    except Exception as exc:
        logger.error(f"[enrich-excel] Error cargando módulo OpenAlex: {exc}")
        raise HTTPException(500, f"Error inicializando OpenAlex: {exc}")

    try:
        enricher = OpenAlexEnricher()
        enricher.MIN_SCORE = float(fuzzy_threshold)
        all_rows = enricher.enrich_from_excel_bytes(data)
    except OpenAlexRateLimitError as exc:
        raise HTTPException(
            status_code=429,
            detail=str(exc),
            headers={"Retry-After": str(exc.retry_after)},
        )
    except ValueError as exc:
        raise HTTPException(422, str(exc))
    except Exception as exc:
        logger.error(f"[enrich-excel] Error procesando archivo: {exc}")
        raise HTTPException(500, f"Error al procesar el archivo: {exc}")

    found          = [r for r in all_rows if r.get("oa_encontrado") and r.get("oa_confianza") != "verificar" and r.get("oa_autor_institucional") is True]
    verify_inst    = [r for r in all_rows if r.get("oa_encontrado") and r.get("oa_confianza") != "verificar" and r.get("oa_autor_institucional") == "verificar"]
    found_no_inst  = [r for r in all_rows if r.get("oa_encontrado") and r.get("oa_confianza") != "verificar" and r.get("oa_autor_institucional") is False]
    verify         = [r for r in all_rows if r.get("oa_encontrado") and r.get("oa_confianza") == "verificar"]
    not_found      = [r for r in all_rows if not r.get("oa_encontrado")]

    logger.info(
        f"[enrich-excel] Total: {len(all_rows)} | "
        f"Con autor U: {len(found)} | Revisar inst.: {len(verify_inst)} | Sin autor U: {len(found_no_inst)} | "
        f"Verificar: {len(verify)} | No encontrados: {len(not_found)}"
    )

    try:
        buf = _build_enrich_excel(found, verify_inst, found_no_inst, verify, not_found, len(all_rows))
    except Exception as exc:
        logger.error(f"[enrich-excel] Error generando Excel: {exc}")
        raise HTTPException(500, f"Error generando el Excel de respuesta: {exc}")

    filename = f"openalex_{_dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )