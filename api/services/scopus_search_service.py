"""
Servicio de búsqueda masiva de publicaciones en Scopus.
"""
import logging
import asyncio
from typing import List, Dict, Tuple, Optional
import io
import time

import openpyxl
from starlette.concurrency import run_in_threadpool

from extractors.scopus import ScopusExtractor, ScopusAPIError
from api.schemas.scopus import ScopusPublicationSearchResult

logger = logging.getLogger("scopus_search")


class ScopusSearchService:
    """Servicio para búsqueda masiva de productos en Scopus."""

    def __init__(self, max_workers: int = 5):
        self.max_workers = max_workers
        self.extractor = ScopusExtractor()
        self.delay_between_requests = 0.1  # Delay por slot concurrente (no secuencial)

    async def search_publications_batch(
        self,
        file_bytes: bytes,
        progress_callback=None,
    ) -> Tuple[List[ScopusPublicationSearchResult], List[ScopusPublicationSearchResult]]:
        """
        Busca un lote de publicaciones en Scopus desde un archivo Excel.
        Procesa hasta max_workers publicaciones en paralelo usando asyncio.gather
        con un Semaphore para respetar los límites de la API de Scopus.

        Args:
            file_bytes: Contenido del archivo Excel
            progress_callback: Función para reportar progreso

        Returns:
            Tupla (encontrados, no_encontrados)
        """
        # 1. Leer Excel
        logger.info("Leyendo archivo Excel...")
        publications = await run_in_threadpool(self._read_publications_from_excel, file_bytes)
        logger.info(f"Se leyeron {len(publications)} publicaciones")

        # 2. Detectar y agrupar duplicados
        unique_pubs, duplicates_map = self._detect_duplicates(publications)
        logger.info(
            f"Publicaciones únicas: {len(unique_pubs)}, "
            f"filas duplicadas: {len(publications) - len(unique_pubs)}, "
            f"workers concurrentes: {self.max_workers}"
        )

        found_results: List[ScopusPublicationSearchResult] = []
        not_found_results: List[ScopusPublicationSearchResult] = []
        total = len(unique_pubs)
        completed = 0

        # Semáforo para limitar requests simultáneos a Scopus
        semaphore = asyncio.Semaphore(self.max_workers)
        # Lock para escritura segura en las listas de resultados
        results_lock = asyncio.Lock()

        async def process_one(pub: Dict) -> None:
            nonlocal completed
            async with semaphore:
                # Pequeño delay dentro del semáforo para no saturar la API
                await asyncio.sleep(self.delay_between_requests)
                try:
                    result = await self._search_single_publication(pub)
                except Exception as e:
                    logger.error(f"Error procesando fila {pub.get('row_num')}: {e}", exc_info=True)
                    result = ScopusPublicationSearchResult(
                        row_num=pub.get("row_num", 0),
                        title=pub.get("title", ""),
                        year=pub.get("year"),
                        doi=pub.get("doi"),
                        issn=pub.get("issn"),
                        magazine=pub.get("magazine"),
                        found_in_scopus=False,
                        search_method="error",
                        search_query=str(e),
                    )

            # Actualizar listas de resultados (fuera del semáforo para no bloquerlo)
            async with results_lock:
                row_nums = duplicates_map.get(pub["_unique_key"], [pub.get("row_num", 0)])
                for row_num in row_nums:
                    result_copy = result.model_copy(deep=True)
                    result_copy.row_num = row_num
                    if result_copy.found_in_scopus:
                        found_results.append(result_copy)
                    else:
                        not_found_results.append(result_copy)

                completed += 1
                if progress_callback and completed % 10 == 0:
                    await progress_callback(completed, total, len(found_results), len(not_found_results))
                if completed % 50 == 0:
                    pct = completed / total * 100
                    logger.info(f"Progreso: {completed}/{total} ({pct:.0f}%) — encontrados: {len(found_results)}")

        # 3. Procesar todas las publicaciones únicas en paralelo
        tasks = [process_one(pub) for pub in unique_pubs]
        await asyncio.gather(*tasks)

        logger.info(
            f"Búsqueda completada: {len(found_results)} encontrados, "
            f"{len(not_found_results)} no encontrados (incluyendo duplicados)"
        )

        return found_results, not_found_results
    
    def _detect_duplicates(self, publications: List[Dict]) -> Tuple[List[Dict], Dict[str, List[int]]]:
        """
        Detecta publicaciones duplicadas y agrupa filas por contenido único.
        
        Args:
            publications: Lista de publicaciones desde Excel
        
        Returns:
            Tupla (publicaciones_unicas, mapa_duplicados)
            donde mapa_duplicados[unique_key] = [row_nums que comparten ese contenido]
        """
        unique_map: Dict[str, Dict] = {}
        duplicates_map: Dict[str, List[int]] = {}
        
        for pub in publications:
            # Crear clave única normalizando título, DOI e ISSN
            key = self._create_unique_key(pub)
            pub["_unique_key"] = key
            
            if key not in unique_map:
                # Primera vez viendo este contenido
                unique_map[key] = pub
                duplicates_map[key] = [pub["row_num"]]
            else:
                # Duplicado detectado
                duplicates_map[key].append(pub["row_num"])
                logger.info(
                    f"Duplicado detectado: Fila {pub['row_num']} "
                    f"coincide con fila {duplicates_map[key][0]}\n"
                    f"  Título: {pub.get('title', '')[:60]}"
                )
        
        unique_pubs = list(unique_map.values())
        return unique_pubs, duplicates_map
    
    def _create_unique_key(self, pub: Dict) -> str:
        """Crea una clave única para detectar duplicados."""
        # Normalizar: DOI exacto (más confiable) → Título + Año + ISSN
        
        doi = pub.get("doi", "").strip().lower()
        if doi:
            return f"doi:{doi}"
        
        # Si no DOI, usar título + año + issn
        title = pub.get("title", "").strip().lower()
        year = pub.get("year", "")
        issn = pub.get("issn", "").strip().lower()
        
        # Normalizar espacios en blanco múltiples
        title = " ".join(title.split())
        
        if title and year:
            return f"title-year-issn:{title}|{year}|{issn}"
        elif title:
            return f"title:{title}"
        elif issn:
            return f"issn:{issn}"
        else:
            # Fallback: usar revista si está disponible
            magazine = pub.get("magazine", "").strip().lower()
            return f"magazine:{magazine}"
    
    async def _search_single_publication(self, pub_data: Dict) -> ScopusPublicationSearchResult:
        """
        Busca una sola publicación en Scopus con estrategia cascada:
        1. DOI exacto (más preciso)
        2. Título (búsqueda por nombre)
        3. ISSN + Año (búsqueda por revista)
        """
        row_num = pub_data.get("row_num", 0)
        title = pub_data.get("title", "").strip()
        year = pub_data.get("year")
        doi = pub_data.get("doi", "").strip()
        issn = pub_data.get("issn", "").strip()
        magazine = pub_data.get("magazine", "").strip()
        
        # ESTRATEGIA 1: Buscar por DOI (más preciso)
        if doi:
            logger.info(f"[Fila {row_num}] ✓ Intentando DOI: {doi}")
            result = await run_in_threadpool(self.extractor.search_by_doi, doi)
            if result:
                result_dict = self._standardrecord_to_dict(result)
                logger.info(f"[Fila {row_num}] ✓✓ ENCONTRADO en Scopus por DOI")
                return self._build_result_found(
                    row_num, pub_data, result_dict,
                    search_method="doi",
                    search_query=doi
                )
            logger.warning(f"[Fila {row_num}] ✗ No encontrado por DOI, intentando título...")
        
        # ESTRATEGIA 2: Buscar por título
        if title:
            logger.info(f"[Fila {row_num}] ✓ Intentando TÍTULO: {title[:60]}...")
            query = f'TITLE-ABS-KEY("{title}")'
            result = await run_in_threadpool(self._search_by_advanced_query, query, title)
            if result:
                logger.info(f"[Fila {row_num}] ✓✓ ENCONTRADO en Scopus por TÍTULO")
                return self._build_result_found(
                    row_num, pub_data, result,
                    search_method="title",
                    search_query=query
                )
            logger.warning(f"[Fila {row_num}] ✗ No encontrado por título, intentando ISSN...")
        
        # ESTRATEGIA 3: Buscar por ISSN + año
        if issn:
            logger.info(f"[Fila {row_num}] ✓ Intentando ISSN: {issn}")
            query = f'ISSN("{issn}")'
            if year:
                query += f' AND PUBYEAR = {year}'
                logger.info(f"[Fila {row_num}]   + Filtro año: {year}")
            result = await run_in_threadpool(self._search_by_advanced_query, query, issn)
            if result:
                logger.info(f"[Fila {row_num}] ✓✓ ENCONTRADO en Scopus por ISSN")
                return self._build_result_found(
                    row_num, pub_data, result,
                    search_method="issn",
                    search_query=query
                )
            logger.warning(f"[Fila {row_num}] ✗ No encontrado por ISSN")
        
        # NO ENCONTRADO en ninguna estrategia
        logger.error(f"[Fila {row_num}] ✗✗ NO ENCONTRADO - Todas las estrategias fallaron")
        return ScopusPublicationSearchResult(
            row_num=row_num,
            title=title,
            year=year,
            doi=doi,
            issn=issn,
            magazine=magazine,
            found_in_scopus=False,
            search_method="na",
            search_query="",
        )
    
    def _standardrecord_to_dict(self, record) -> Dict:
        """Convierte un StandardRecord a diccionario."""
        return {
            'source_id': record.source_id,
            'title': record.title,
            'source_title': record.source_journal,
            'doi': record.doi,
            'pub_year': record.publication_year,
            'issn': record.issn,
        }
    
    def _search_by_advanced_query(self, query: str, identifier: str) -> Optional[Dict]:
        """Ejecuta una búsqueda avanzada en Scopus y retorna el primer resultado."""
        try:
            params = {
                "query": query,
                "count": 1,
                "field": (
                    "dc:identifier,doi,dc:title,prism:publicationName,"
                    "prism:coverDate,subtypeDescription,citedby-count,"
                    "author,prism:issn,openaccess,openaccessFlag,"
                    "dc:description,authkeywords,prism:volume,prism:issueIdentifier,"
                    "prism:pageRange,afid,affiliation"
                ),
            }
            response = self.extractor.session.get(
                self.extractor.SEARCH_URL,
                params=params,
                timeout=self.extractor.config.timeout,
            )
            response.raise_for_status()
            
            # Intentar parsear como JSON
            try:
                data = response.json()
            except Exception as json_err:
                logger.warning(
                    f"Error parsing JSON en búsqueda '{identifier}': {json_err}\n"
                    f"Response status: {response.status_code}\n"
                    f"Response text: {response.text[:200]}"
                )
                return None
            
            entries = data.get("search-results", {}).get("entry", [])
            
            if entries and not ("error" in str(entries).lower()):
                # Parsear el primer resultado
                result_dict = {}
                entry = entries[0]
                
                # Extraer campos principales
                result_dict['source_id'] = entry.get("dc:identifier", "").replace("SCOPUS_ID:", "").strip()
                result_dict['title'] = entry.get("dc:title", "")
                result_dict['source_title'] = entry.get("prism:publicationName", "")
                result_dict['doi'] = entry.get("prism:doi", "")
                
                # Año de publicación
                cover_date = entry.get("prism:coverDate", "")
                result_dict['pub_year'] = int(cover_date[:4]) if cover_date and len(cover_date) >= 4 else None
                
                # ISSN
                result_dict['issn'] = entry.get("prism:issn", "")
                
                return result_dict
                
        except Exception as e:
            logger.warning(f"Error en búsqueda avanzada '{identifier}': {e}")
        
        return None
    
    def _build_result_found(self, row_num: int, pub_data: Dict, scopus_result, search_method: str, search_query: str) -> ScopusPublicationSearchResult:
        """Construye un resultado de publicación encontrada."""
        # scopus_result es un dict con los campos extraídos
        
        # Determinar coincidencias
        matched_fields = []
        if pub_data.get("title", "").lower().strip() == scopus_result.get('title', '').lower().strip():
            matched_fields.append("title")
        if pub_data.get("year") == scopus_result.get('pub_year'):
            matched_fields.append("year")
        if pub_data.get("doi", "").lower().strip() == (scopus_result.get('doi', '') or "").lower().strip():
            matched_fields.append("doi")
        if pub_data.get("issn", "").lower().strip() == (scopus_result.get('issn', '') or "").lower().strip():
            matched_fields.append("issn")
        
        return ScopusPublicationSearchResult(
            row_num=row_num,
            title=pub_data.get("title", ""),
            year=pub_data.get("year"),
            doi=pub_data.get("doi"),
            issn=pub_data.get("issn"),
            magazine=pub_data.get("magazine"),
            found_in_scopus=True,
            scopus_id=scopus_result.get('source_id'),
            scopus_title=scopus_result.get('title'),
            scopus_journal=scopus_result.get('source_title'),
            scopus_doi=scopus_result.get('doi'),
            scopus_issn=scopus_result.get('issn'),
            search_method=search_method,
            search_query=search_query,
            matched_fields=matched_fields,
        )
    
    def _read_publications_from_excel(self, file_bytes: bytes) -> List[Dict]:
        """Lee publicaciones desde un archivo Excel."""
        publications = []
        
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            
            # Detectar encabezados
            headers = {}
            first_row = None
            
            for row_idx, row in enumerate(ws.iter_rows(max_row=5, values_only=False), 1):
                row_values = [cell.value for cell in row]
                if self._looks_like_header_row(row_values):
                    first_row = row_idx
                    for col_idx, cell in enumerate(row, 1):
                        header_name = str(cell.value or "").lower().strip()
                        headers[col_idx] = header_name
                    break
            
            if not first_row:
                raise ValueError("No se detectó fila de encabezados en el Excel")
            
            logger.info(f"Encabezados detectados en fila {first_row}: {list(headers.values())}")
            
            # Leer datos
            for row_idx, row in enumerate(ws.iter_rows(min_row=first_row + 1, values_only=True), first_row + 1):
                if not any(row):  # Saltar filas vacías
                    continue
                
                pub = self._parse_row(row, headers, row_idx)
                if pub:
                    publications.append(pub)
            
        except Exception as e:
            logger.error(f"Error leyendo Excel: {e}", exc_info=True)
            raise ValueError(f"Error leyendo archivo Excel: {e}")
        
        return publications
    
    def _parse_row(self, row: tuple, headers: Dict[int, str], row_idx: int) -> Optional[Dict]:
        """Parsea una fila del Excel."""
        pub = {
            "row_num": row_idx
        }
        
        for col_idx, header_name in headers.items():
            if col_idx - 1 >= len(row):
                continue
            
            value = row[col_idx - 1]
            
            # Mapear encabezado a campo
            if header_name in ("título", "title", "título del artículo", "titulo"):
                pub["title"] = str(value or "").strip()
            elif header_name in ("año", "year", "year published"):
                try:
                    pub["year"] = int(value) if value else None
                except:
                    pub["year"] = None
            elif header_name in ("doi",):
                pub["doi"] = str(value or "").strip()
            elif header_name in ("issn",):
                pub["issn"] = str(value or "").strip()
            elif header_name in ("revista", "source title", "magazine", "journal", "source_title"):
                pub["magazine"] = str(value or "").strip()
        
        # Validar que al menos tenga título
        if not pub.get("title"):
            logger.warning(f"Fila {row_idx} sin título, se omite")
            return None
        
        return pub
    
    def _looks_like_header_row(self, row_values: list) -> bool:
        """Detecta si una fila contiene encabezados."""
        known_headers = {
            "título", "title", "año", "year", "doi", "issn", "revista", 
            "magazine", "source title", "journal", "source_title",
            "título del artículo", "titulo", "year published", "openalex_id",
            "tipo", "documento", "autores", "authors", "institucion",
        }
        
        non_empty = [str(c or "").lower().strip() for c in row_values if c is not None and str(c).strip()]
        return any(v in known_headers for v in non_empty)
