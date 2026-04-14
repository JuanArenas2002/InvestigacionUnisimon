"""
Servicio para extraer la producción académica de autores desde Scopus por Author ID.
Procesa múltiples autores en paralelo.
"""

import logging
import asyncio
import os
from typing import List, Dict, Optional, Tuple
import io

import openpyxl
from starlette.concurrency import run_in_threadpool

from extractors.scopus import ScopusExtractor, ScopusAPIError

logger = logging.getLogger("scopus_author_production")


# ════════════════════════════════════════════════════════════════════════════════
# OBTENER PERFIL DEL AUTOR
# ════════════════════════════════════════════════════════════════════════════════

async def get_scopus_author_profile(author_id: str) -> dict:
    """
    Obtiene el perfil del autor desde Scopus Author Retrieval API.
    Intenta conexión, pero si falla devuelve dict vacío (no crashea).
    """
    try:
        import requests
        from config import scopus_config

        url = f"{scopus_config.base_url}/author/author_id/{author_id}"
        headers = {
            "Accept": "application/json",
            "X-ELS-APIKey": scopus_config.api_key or "",
        }

        resp = requests.get(url, headers=headers, timeout=5)
        if resp.status_code == 200:
            data = resp.json()
            author_entry = data.get("author-retrieval-response", [])
            if isinstance(author_entry, list):
                author_entry = author_entry[0] if author_entry else {}

            # ═══════════════════════════════════════════════════════════
            # EXTRAER NOMBRE desde preferred-name
            # ═══════════════════════════════════════════════════════════
            author_profile = author_entry.get("author-profile", {})
            preferred_name = author_profile.get("preferred-name", {})

            given_name = preferred_name.get("given-name", "")
            surname = preferred_name.get("surname", "")
            name = f"{given_name} {surname}".strip()

            if not name:
                # Fallback al indexed-name
                name = preferred_name.get("indexed-name", "")

            # ═══════════════════════════════════════════════════════════
            # EXTRAER INSTITUCIÓN ACTUAL
            # ═══════════════════════════════════════════════════════════
            aff_current = author_profile.get("affiliation-current", {})
            institution = ""
            if aff_current and isinstance(aff_current, dict):
                # La afiliación actual es una lista dentro de affiliation-current
                aff_list = aff_current.get("affiliation", [])
                if aff_list and isinstance(aff_list, list) and len(aff_list) > 0:
                    aff = aff_list[0]
                    if isinstance(aff, dict):
                        # Intentar obtener el nombre de la institución
                        ip_doc = aff.get("ip-doc", {})
                        if isinstance(ip_doc, dict):
                            # Preferir afdispname que es el nombre legible
                            institution = ip_doc.get("afdispname", "") or \
                                        ip_doc.get("preferred-name", {}).get("$", "") or \
                                        ip_doc.get("sort-name", "")

            # ═══════════════════════════════════════════════════════════
            # EXTRAER ÁREAS DE INVESTIGACIÓN (subject-areas)
            # ═══════════════════════════════════════════════════════════
            subject_areas_list = []
            subject_areas_data = author_entry.get("subject-areas", {})
            if subject_areas_data and isinstance(subject_areas_data, dict):
                areas = subject_areas_data.get("subject-area", [])
                if areas and isinstance(areas, list):
                    subject_areas_list = [
                        area.get("$", "") or area.get("@abbrev", "")
                        for area in areas if isinstance(area, dict)
                    ]
                    # Deduplicar y limitar a 5
                    subject_areas_list = list(dict.fromkeys(subject_areas_list))[:5]

            subject_areas = "; ".join(filter(None, subject_areas_list)) if subject_areas_list else ""

            logger.info(f"Perfil Scopus: {name} @ {institution} [{subject_areas}]")
            return {
                "name": name,
                "institution_current": institution,
                "subject_areas": subject_areas
            }
    except Exception as e:
        logger.warning(f"No se pudo obtener perfil de {author_id}: {e}")

    return {"name": "", "institution_current": "", "subject_areas": ""}


logger = logging.getLogger("scopus_author_production")


# ════════════════════════════════════════════════════════════════════════════════
# FUNCIONES HELPER
# ════════════════════════════════════════════════════════════════════════════════

def _extract_authors_string(authors_list: List) -> str:
    """
    Convierte una lista de autores (dicts o strings) en un string separado por ";".
    
    Maneja múltiples formatos:
    - Lista de dicts: [{"name": "Juan Pérez", ...}, {"name": "María García", ...}]
    - Lista de strings: ["Juan Pérez", "María García"]
    - Lista de objetos con atributo 'name'
    """
    if not authors_list or not isinstance(authors_list, list):
        return ""
    
    author_names = []
    for author in authors_list:
        if not author:
            continue
        
        # Si es dict
        if isinstance(author, dict):
            name = author.get("name", "")
            if name:
                author_names.append(str(name).strip())
        
        # Si es string
        elif isinstance(author, str):
            name = author.strip()
            if name:
                author_names.append(name)
        
        # Si es objeto con atributo 'name'
        elif hasattr(author, "name"):
            name = getattr(author, "name", "")
            if name:
                author_names.append(str(name).strip())
        
        # Fallback: convertir a string
        else:
            name = str(author).strip()
            if name and name not in ["None", ""]:
                author_names.append(name)
    
    return "; ".join(author_names)


class ScopusAuthorProductionService:
    """Servicio para extraer producción de autores de Scopus."""

    def __init__(self, max_workers: int = 3):
        self.max_workers = max_workers
        self.extractor = ScopusExtractor()
        self.delay_between_requests = 0.2  # Delay entre requests

    async def process_author_ids(
        self,
        file_bytes: bytes,
        progress_callback=None,
    ) -> List[Dict]:
        """
        Procesa un archivo Excel con IDs de autores de Scopus.
        
        El Excel debe tener una columna con "author_id" o similar.
        Busca la producción de cada autor y retorna lista de dict con:
        - author_id
        - author_name (si está disponible o se obtiene de Scopus)
        - publications (lista de publicaciones)
        
        Args:
            file_bytes: Contenido del archivo Excel
            progress_callback: Función para reportar progreso
            
        Returns:
            Lista de dicts con información de cada autor y sus publicaciones
        """
        logger.info("Leyendo archivo Excel...")
        author_ids = await run_in_threadpool(self._read_author_ids_from_excel, file_bytes)
        logger.info(f"Se leyeron {len(author_ids)} autores")

        if not author_ids:
            logger.warning("No se encontraron IDs de autores en el Excel")
            return []

        # Procesamiento paralelo con límite de workers
        semaphore = asyncio.Semaphore(self.max_workers)
        results_lock = asyncio.Lock()
        results: List[Dict] = []
        
        total = len(author_ids)
        completed = 0

        async def process_one(author_data: Dict) -> None:
            nonlocal completed
            async with semaphore:
                await asyncio.sleep(self.delay_between_requests)
                try:
                    author_id = author_data.get("author_id")
                    logger.info(f"Procesando autor {author_id}...")

                    # ══════════════════════════════════════════════════════════
                    # PASO 1: Obtener perfil verdadero del investigador
                    # ══════════════════════════════════════════════════════════
                    profile = await get_scopus_author_profile(author_id)

                    # Nombre verdadero desde el perfil
                    author_name = ""
                    subject_areas = ""
                    institution_current = ""
                    if profile and profile.get("name"):
                        author_name = profile.get("name", "")
                        subject_areas = profile.get("subject_areas", "")
                        institution_current = profile.get("institution_current", "")
                        logger.info(f"  Perfil: {author_name} ({institution_current})")

                    # Fallback al nombre del Excel si el perfil falla
                    if not author_name:
                        author_name = author_data.get("author_name", f"Autor {author_id}")

                    # ══════════════════════════════════════════════════════════
                    # PASO 2: Obtener publicaciones
                    # ══════════════════════════════════════════════════════════
                    logger.info(f"Extrayendo producción para autor {author_id}...")
                    publications = await self._fetch_author_publications(author_id)

                    result = {
                        "author_id": author_id,
                        "author_name": author_name,
                        "affiliation": author_data.get("affiliation", ""),
                        "institution_current": institution_current,  # Del perfil
                        "subject_areas": subject_areas,  # Del perfil
                        "publications_count": len(publications),
                        "publications": publications,
                        "status": "success",
                        "error": None,
                    }

                    async with results_lock:
                        results.append(result)

                except Exception as e:
                    logger.error(f"Error procesando autor {author_data.get('author_id')}: {e}", exc_info=True)

                    result = {
                        "author_id": author_data.get("author_id"),
                        "author_name": author_data.get("author_name", "") or f"Autor {author_data.get('author_id')}",
                        "affiliation": author_data.get("affiliation", ""),
                        "institution_current": "",
                        "subject_areas": "",
                        "publications_count": 0,
                        "publications": [],
                        "status": "error",
                        "error": str(e),
                    }

                    async with results_lock:
                        results.append(result)

                finally:
                    completed += 1
                    if progress_callback:
                        progress_callback(completed, total)

        # Ejecutar extracciones en paralelo
        tasks = [process_one(author_data) for author_data in author_ids]
        await asyncio.gather(*tasks)
        
        logger.info(f"Procesamiento completado. {len(results)} autores procesados")
        return results

    def _read_author_ids_from_excel(self, file_bytes: bytes) -> List[Dict]:
        """
        Lee el Excel y extrae IDs de autores.
        Busca columnas con nombres como: author_id, authorid, scopus_id, scopus_author_id
        Opcionalmente: author_name, affiliation
        """
        file_obj = io.BytesIO(file_bytes)
        wb = openpyxl.load_workbook(file_obj)
        ws = wb.active
        
        # Leer headers
        headers = {}
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=False), start=1):
            values = [cell.value for cell in row]
            # Buscar fila con headers - busca cualquier columna que parezca tener IDs
            has_id_like_header = any(
                v and isinstance(v, str) and any(kw in v.lower() for kw in 
                ["author_id", "authorid", "scopus", "researcher_id", "id"]) 
                for v in values
            )
            if has_id_like_header:
                header_row = row_idx
                for col_idx, cell in enumerate(row, start=1):
                    if cell.value:
                        headers[col_idx] = cell.value.lower().strip()
                logger.info(f"Headers encontrados en fila {header_row}: {headers}")
                break
        
        if not header_row:
            logger.warning("No se encontró fila de headers, usando fila 1")
            header_row = 1
            for col_idx, cell in enumerate(ws[header_row], start=1):
                if cell.value:
                    headers[col_idx] = cell.value.lower().strip()
        
        logger.info(f"Headers procesados: {headers}")
        
        # Encontrar columnas relevantes con búsqueda más flexible
        author_id_col = None
        author_name_col = None
        affiliation_col = None
        
        for col_idx, header in headers.items():
            header_lower = header.lower()
            
            # Columna de ID
            if any(kw in header_lower for kw in ["author_id", "authorid", "scopus_id", "scopus_author_id", "researcher_id", "scopus author"]):
                author_id_col = col_idx
                logger.info(f"Columna Author ID encontrada: col {col_idx} = '{header}'")
            
            # Columna de Nombre
            elif any(kw in header_lower for kw in ["author_name", "nombre", "name", "researcher_name", "fullname", "full_name"]):
                author_name_col = col_idx
                logger.info(f"Columna Author Name encontrada: col {col_idx} = '{header}'")
            
            # Columna de Afiliación
            elif any(kw in header_lower for kw in ["affiliation", "afiliación", "organization", "organization", "institution", "institución"]):
                affiliation_col = col_idx
                logger.info(f"Columna Affiliation encontrada: col {col_idx} = '{header}'")
        
        if not author_id_col:
            raise ValueError("No se encontró columna de author_id. Use: author_id, authorid, scopus_id, etc.")
        
        logger.info(f"Columnas detectadas: author_id={author_id_col}, author_name={author_name_col}, affiliation={affiliation_col}")
        
        # Leer datos - usar dict para deduplicar por author_id
        unique_authors = {}  # key=author_id, value=author_data
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=False), start=header_row + 1):
            try:
                # Validar que la fila existe
                if not row:
                    continue
                
                # Acceder de forma segura a author_id
                author_id = None
                if author_id_col and author_id_col - 1 < len(row):
                    author_id_cell = row[author_id_col - 1]
                    author_id = author_id_cell.value if author_id_cell else None
                
                if not author_id or not str(author_id).strip():
                    continue
                
                author_id_str = str(author_id).strip()
                
                # Si el autor ya existe, preferir el que tenga nombre
                if author_id_str in unique_authors:
                    existing = unique_authors[author_id_str]
                    
                    # Actualizar nombre si es mejor
                    if author_name_col and author_name_col - 1 < len(row):
                        try:
                            name_cell = row[author_name_col - 1]
                            name_value = name_cell.value if name_cell else None
                            if name_value and str(name_value).strip() and not existing["author_name"]:
                                existing["author_name"] = str(name_value).strip()
                        except (IndexError, TypeError):
                            pass
                    
                    # Actualizar afiliación si es mejor
                    if affiliation_col and affiliation_col - 1 < len(row):
                        try:
                            aff_cell = row[affiliation_col - 1]
                            aff_value = aff_cell.value if aff_cell else None
                            if aff_value and str(aff_value).strip() and not existing["affiliation"]:
                                existing["affiliation"] = str(aff_value).strip()
                        except (IndexError, TypeError):
                            pass
                    
                    continue
                
                # Crear entrada nueva para este autor
                author_data = {
                    "author_id": author_id_str,
                    "author_name": "",
                    "affiliation": "",
                }
                
                # Extraer nombre de forma segura
                if author_name_col and author_name_col - 1 < len(row):
                    try:
                        name_cell = row[author_name_col - 1]
                        name_value = name_cell.value if name_cell else None
                        if name_value:
                            author_data["author_name"] = str(name_value).strip()
                    except (IndexError, TypeError):
                        pass
                
                # Extraer afiliación de forma segura
                if affiliation_col and affiliation_col - 1 < len(row):
                    try:
                        aff_cell = row[affiliation_col - 1]
                        aff_value = aff_cell.value if aff_cell else None
                        if aff_value:
                            author_data["affiliation"] = str(aff_value).strip()
                    except (IndexError, TypeError):
                        pass
                
                unique_authors[author_id_str] = author_data
                logger.debug(f"Autor leído: {author_data}")
            
            except Exception as e:
                logger.warning(f"Error procesando fila {row_idx}: {e}", exc_info=False)
                continue
        
        # Convertir a lista
        author_ids = list(unique_authors.values())
        logger.info(f"Se extrajeron {len(author_ids)} autores únicos")
        return author_ids

    async def _fetch_author_publications(self, author_id: str) -> List[Dict]:
        """
        Busca todas las publicaciones de un autor en Scopus.
        """
        try:
            # Query base por autor
            query = f"AU-ID({author_id})"

            # Filtro institucional opcional desde .env
            raw_aff_ids = os.getenv("SCOPUS_AFFILIATION_IDS", "")
            aff_ids = [aid.strip() for aid in raw_aff_ids.split(",") if aid and aid.strip()]
            if aff_ids:
                aff_clause = " OR ".join([f"AF-ID({aid})" for aid in aff_ids])
                query = f"{query} AND ({aff_clause})"

            logger.info(f"Ejecutando query: {query}")
            
            # Extractor retorna los registros de Scopus
            records = await run_in_threadpool(
                self.extractor.extract,
                query=query,
                max_results=None
            )
            
            if not records:
                logger.info(f"No se encontraron registros para autor {author_id}")
                return []
            
            # Transformar registros a dict con campos importantes
            # Usar dict para deduplicar por Scopus ID o DOI
            unique_publications = {}  # key=scopus_id o doi, value=pub
            
            for record in records:
                try:
                    # record es un StandardRecord (dataclass), convertir a dict
                    if hasattr(record, "to_dict"):
                        record_dict = record.to_dict()
                    else:
                        # Fallback: convertir atributos a dict manualmente
                        record_dict = {
                            "source_id": getattr(record, "source_id", ""),
                            "title": getattr(record, "title", ""),
                            "doi": getattr(record, "doi", ""),
                            "publication_year": getattr(record, "publication_year", ""),
                            "publication_date": getattr(record, "publication_date", ""),
                            "publication_type": getattr(record, "publication_type", ""),
                            "source_journal": getattr(record, "source_journal", ""),
                            "authors": getattr(record, "authors", []),
                            "citation_count": getattr(record, "citation_count", 0),
                            "issn": getattr(record, "issn", ""),
                            "is_open_access": getattr(record, "is_open_access", ""),
                            "oa_status": getattr(record, "oa_status", ""),
                            "url": getattr(record, "url", ""),
                            "raw_data": getattr(record, "raw_data", {}),
                        }
                    
                    # Extraer autores como string con mejor manejo
                    authors_list = record_dict.get("authors", [])
                    authors_str = _extract_authors_string(authors_list)
                    if not authors_str and record_dict.get("raw_data"):
                        # Fallback: intentar del raw_data
                        raw_authors = record_dict["raw_data"].get("authors", "")
                        authors_str = raw_authors if isinstance(raw_authors, str) else ""
                    
                    # Extraer datos del raw_data
                    raw_data = record_dict.get("raw_data", {}) or {}
                    if not isinstance(raw_data, dict):
                        raw_data = {}
                    
                    # Intentar obtener datos de múltiples fuentes
                    volume = raw_data.get("volume", "") or getattr(record, "volume", "")
                    issue = raw_data.get("issue", "") or getattr(record, "issue", "")
                    pages = raw_data.get("pages", "") or getattr(record, "pages", "")
                    eissn = raw_data.get("eissn", "") or getattr(record, "eissn", "")
                    description = raw_data.get("description", "") or raw_data.get("abstract", "") or getattr(record, "description", "")
                    keywords = raw_data.get("keywords", "") or getattr(record, "keywords", "")
                    source_type = raw_data.get("source_type", "") or getattr(record, "source_type", "")
                    aggregation_type = raw_data.get("aggregation_type", "") or getattr(record, "aggregation_type", "")
                    subtype_description = raw_data.get("subtype_description", "") or getattr(record, "subtype_description", "")
                    
                    # Construir string de keywords si es lista
                    if isinstance(keywords, list):
                        keywords = "; ".join([str(k) for k in keywords if k])
                    
                    # Construir URL de Scopus si no viene en el registro
                    scopus_id = record_dict.get("source_id", "")
                    url = record_dict.get("url", "") or ""
                    if not url and scopus_id:
                        # Construir URL automáticamente en formato Scopus
                        url = f"https://www.scopus.com/record/display.uri?eid=2-s2.0-{scopus_id}"
                    elif not url and record_dict.get("doi"):
                        # Si no hay URL pero hay DOI, usar enlace DOI
                        url = f"https://doi.org/{record_dict.get('doi')}"
                    
                    pub = {
                        "scopus_id": scopus_id or "",
                        "title": record_dict.get("title", "") or "",
                        "doi": record_dict.get("doi", "") or "",
                        "publication_year": record_dict.get("publication_year", "") or "",
                        "publication_date": record_dict.get("publication_date", "") or "",
                        "publication_type": record_dict.get("publication_type", "") or subtype_description or aggregation_type or "",
                        "source_title": record_dict.get("source_journal", "") or "",
                        "source_type": source_type or "",
                        "authors": authors_str or "",
                        "citation_count": record_dict.get("citation_count", 0) or 0,
                        "volume": str(volume) if volume else "",
                        "issue": str(issue) if issue else "",
                        "pages": str(pages) if pages else "",
                        "issn": record_dict.get("issn", "") or "",
                        "eissn": str(eissn) if eissn else "",
                        "is_open_access": str(record_dict.get("is_open_access", "")) if record_dict.get("is_open_access") is not None else "",
                        "oa_status": record_dict.get("oa_status", "") or "",
                        "keywords": str(keywords) if keywords else "",
                        "description": str(description[:500]) if description else "",  # Limitar a 500 chars
                        "url": url,
                    }
                    
                    # Deduplicar por Scopus ID primero, luego por DOI
                    dedup_key = None
                    if pub.get("scopus_id"):
                        dedup_key = f"scopus:{pub['scopus_id']}"
                    elif pub.get("doi"):
                        dedup_key = f"doi:{pub['doi']}"
                    else:
                        # Si no tiene ID ni DOI, usar el título normalizado
                        dedup_key = f"title:{pub.get('title', '')[:50]}"
                    
                    # Si ya existe esta publicación, saltar
                    if dedup_key not in unique_publications:
                        unique_publications[dedup_key] = pub
                
                except Exception as e:
                    logger.warning(f"Error procesando registro individual: {e}", exc_info=False)
                    continue
            
            # Convertir a lista
            publications = list(unique_publications.values())
            
            # Filtrar publicaciones inválidas (sin título ni scopus_id)
            valid_publications = [
                pub for pub in publications
                if pub.get("title", "").strip() or pub.get("scopus_id", "").strip()
            ]
            
            logger.info(f"Se encontraron {len(valid_publications)} publicaciones válidas para autor {author_id} (originales: {len(publications)})")
            return valid_publications
            
        except ScopusAPIError as e:
            logger.error(f"Error en API de Scopus para autor {author_id}: {e}")
            raise
        except Exception as e:
            logger.error(f"Error inesperado extrayendo producción de {author_id}: {e}")
            raise
