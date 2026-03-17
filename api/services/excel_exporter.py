"""
Servicio para exportar publicaciones a Excel con metadatos completos.
Reutilizable desde FastAPI y scripts standalone.
"""

import io
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Dict, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from extractors.scopus import ScopusExtractor
from db.session import get_session
from db.models import Author, PublicationAuthor, CanonicalPublication

logger = logging.getLogger(__name__)

# ════════════════════════════════════════════════════════════════════════════════
# ESTILOS DE EXCEL
# ════════════════════════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
ALTER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
SUMMARY_FONT = Font(bold=True, color="375623", size=11)
BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

COLUMNS = [
    ("Año", "publication_year", 8),
    ("Título", "title", 50),
    ("Autores", "authors", 40),
    ("Journal", "source_journal", 35),
    ("DOI", "doi", 25),
    ("ISSN", "issn", 15),
    ("Tipo", "publication_type", 15),
    ("Citas", "citation_count", 8),
    ("Open Access", "is_open_access", 12),
    ("Scopus ID", "source_id", 20),
]


def extract_publications_for_export(
    query: str,
    author_id: str,
    affiliation_ids: Optional[List[str]] = None,
    verbose: bool = False,
) -> Tuple[List[dict], str]:
    """
    Extrae todas las publicaciones usando Scopus API.
    
    Returns:
        (lista de publicaciones, nombre del investigador)
    """
    try:
        logger.info(f"Extrayendo publicaciones con query: {query}")
        extractor = ScopusExtractor()
        
        records = extractor.extract(query=query, max_results=None)
        logger.info(f"Scopus retornó {len(records) if records else 0} registros")
        
        if not records:
            logger.warning(f"No se encontraron publicaciones para la consulta: {query}")
            return [], "Desconocido"
        
        # Extraer nombre del investigador del primer registro
        investigator_name = "Desconocido"
        if records and records[0].authors:
            # Buscar el autor que coincida con el author_id (scopus_id)
            for rec in records:
                for author in rec.authors:
                    if author.get('scopus_id') and str(author['scopus_id']).replace('SCOPUS_ID:', '') == str(author_id):
                        investigator_name = author.get('name', 'Desconocido')
                        logger.info(f"Investigador encontrado: {investigator_name}")
                        break
                if investigator_name != "Desconocido":
                    break
        
        # Convertir StandardRecord a dict y ordenar por año descendente
        publications = []
        for rec in records:
            pub_dict = {
                'publication_year': rec.publication_year,
                'title': rec.title,
                'authors': ', '.join([a.get('name', '') for a in rec.authors]) if rec.authors else '',
                'source_journal': rec.source_journal,
                'doi': rec.doi,
                'issn': rec.issn,
                'publication_type': rec.publication_type,
                'citation_count': rec.citation_count or 0,
                'is_open_access': 'Sí' if rec.is_open_access else 'No',
                'source_id': rec.source_id,
                'url': rec.url,
                'oa_status': rec.oa_status,
            }
            publications.append(pub_dict)
        
        publications.sort(key=lambda x: x['publication_year'] or 0, reverse=True)
        
        if verbose:
            logger.info(f"✓ Extraídas {len(publications)} publicaciones para {investigator_name}")
        
        return publications, investigator_name
    
    except Exception as e:
        logger.error(f"Error extrayendo publicaciones: {str(e)}", exc_info=True)
        raise


def generate_publications_excel_bytes(
    publications: List[dict],
    investigator_name: str,
    author_id: str,
    query_used: str = None,
) -> bytes:
    """
    Genera un archivo Excel en bytes con las publicaciones.
    
    Args:
        publications: Lista de dicts con metadatos de publicaciones
        investigator_name: Nombre del investigador
        author_id: ID del investigador en Scopus
    
    Returns:
        Bytes del archivo Excel
    """
    wb = openpyxl.Workbook()
    
    # ────────────────────────────────────────────────────────────────────────────
    # HOJA 1: RESUMEN
    # ────────────────────────────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Resumen"
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 50
    
    # Título
    ws_summary['A1'] = "RESUMEN DE PRODUCCIÓN CIENTÍFICA"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:B1')
    
    # Información del investigador
    row = 3
    ws_summary[f'A{row}'] = "Investigador:"
    ws_summary[f'B{row}'] = investigator_name
    ws_summary[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws_summary[f'A{row}'] = "Scopus ID:"
    ws_summary[f'B{row}'] = author_id
    ws_summary[f'A{row}'].font = Font(bold=True)
    
    row += 1
    ws_summary[f'A{row}'] = "Fecha de extracción:"
    ws_summary[f'B{row}'] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws_summary[f'A{row}'].font = Font(bold=True)
    
    # Estadísticas
    row += 2
    ws_summary[f'A{row}'] = "ESTADÍSTICAS"
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    total_pubs = len(publications)
    ws_summary[f'A{row}'] = "Total de publicaciones:"
    ws_summary[f'B{row}'] = total_pubs
    ws_summary[f'A{row}'].font = Font(bold=True)
    
    row += 1
    years = sorted(set([p['publication_year'] for p in publications if p['publication_year']]))
    ws_summary[f'A{row}'] = "Rango temporal:"
    ws_summary[f'B{row}'] = f"{min(years)} - {max(years)}" if years else "N/A"
    ws_summary[f'A{row}'].font = Font(bold=True)
    
    row += 1
    total_citations = sum([p['citation_count'] for p in publications])
    ws_summary[f'A{row}'] = "Total de citas:"
    ws_summary[f'B{row}'] = total_citations
    ws_summary[f'A{row}'].font = Font(bold=True)
    
    row += 1
    open_access = len([p for p in publications if p['is_open_access'] == 'Sí'])
    ws_summary[f'A{row}'] = "Acceso abierto:"
    ws_summary[f'B{row}'] = f"{open_access} ({100*open_access/total_pubs:.1f}%)"
    ws_summary[f'A{row}'].font = Font(bold=True)
    
    # Consulta utilizada
    row += 2
    ws_summary[f'A{row}'] = "CONSULTA UTILIZADA"
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    ws_summary[f'A{row}'] = "Scopus Query:"
    ws_summary[f'A{row}'].font = Font(bold=True)
    ws_summary.merge_cells(f'A{row}:B{row}')
    
    row += 1
    if query_used:
        ws_summary[f'A{row}'] = query_used
        ws_summary[f'A{row}'].alignment = Alignment(wrap_text=True)
        ws_summary.merge_cells(f'A{row}:B{row}')
    
    # Publicaciones por año
    row += 3
    ws_summary[f'A{row}'] = "PUBLICACIONES POR AÑO"
    ws_summary[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    ws_summary[f'A{row}'] = "Año"
    ws_summary[f'B{row}'] = "Cantidad"
    ws_summary[f'A{row}'].font = SUMMARY_FONT
    ws_summary[f'B{row}'].font = SUMMARY_FONT
    ws_summary[f'A{row}'].fill = SUMMARY_FILL
    ws_summary[f'B{row}'].fill = SUMMARY_FILL
    
    pub_by_year = {}
    for pub in publications:
        year = pub['publication_year']
        if year:
            pub_by_year[year] = pub_by_year.get(year, 0) + 1
    
    for year in sorted(pub_by_year.keys(), reverse=True):
        row += 1
        ws_summary[f'A{row}'] = year
        ws_summary[f'B{row}'] = pub_by_year[year]
    
    # ────────────────────────────────────────────────────────────────────────────
    # HOJA 2: DETALLE
    # ────────────────────────────────────────────────────────────────────────────
    ws_detail = wb.create_sheet("Publicaciones")
    
    # Headers
    for col_idx, (column_name, _, width) in enumerate(COLUMNS, 1):
        col_letter = get_column_letter(col_idx)
        ws_detail.column_dimensions[col_letter].width = width
        
        cell = ws_detail[f'{col_letter}1']
        cell.value = column_name
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER_THIN
    
    # Datos
    for row_idx, pub in enumerate(publications, 2):
        for col_idx, (_, field_name, _) in enumerate(COLUMNS, 1):
            col_letter = get_column_letter(col_idx)
            cell = ws_detail[f'{col_letter}{row_idx}']
            
            value = pub.get(field_name, '')
            cell.value = value
            cell.border = BORDER_THIN
            
            # Alternar colores
            if row_idx % 2 == 0:
                cell.fill = ALTER_FILL
            
            # Alineación
            if field_name in ['publication_year', 'citation_count']:
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left', wrap_text=True)
    
    # Congelar encabezados
    ws_detail.freeze_panes = 'A2'
    
    # Convertir a bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    logger.info(f"Excel generado: {total_pubs} publicaciones, {len(publications)} registros")
    
    return output.getvalue()


def generate_publications_excel_file(
    author_id: str,
    affiliation_ids: Optional[List[str]] = None,
    output_dir: Path = Path("reports/exports"),
) -> Dict:
    """
    Genera un archivo Excel y lo guarda en disco.
    
    Args:
        author_id: ID del autor en Scopus
        affiliation_ids: IDs de afiliación opcionales
        output_dir: Directorio de salida
    
    Returns:
        Dict con información del archivo generado
    """
    try:
        logger.info(f"[EXCEL] Iniciando generación para AU-ID: {author_id}")
        
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"[EXCEL] Directorio creado: {output_dir}")
        
        # Construir consulta Scopus
        if affiliation_ids:
            aff_clause = " OR ".join([f"AF-ID ( {aff_id} )" for aff_id in affiliation_ids])
            query = f"AU-ID ( {author_id} ) AND ({aff_clause})"
        else:
            query = f"AU-ID ( {author_id} )"
        
        logger.info(f"[EXCEL] Query Scopus: {query}")
        
        # Extraer publicaciones
        publications, investigator_name = extract_publications_for_export(
            query=query,
            author_id=author_id,
            affiliation_ids=affiliation_ids,
            verbose=True,
        )
        
        if not publications:
            logger.error(f"[EXCEL] No se encontraron publicaciones")
            raise ValueError("No se encontraron publicaciones para exportar")
        
        logger.info(f"[EXCEL] {len(publications)} publicaciones extraídas de Scopus")
        
        # Generar Excel
        logger.info(f"[EXCEL] Generando bytes de Excel...")
        excel_bytes = generate_publications_excel_bytes(
            publications=publications,
            investigator_name=investigator_name,
            author_id=author_id,
            query_used=query,
        )
        
        logger.info(f"[EXCEL] Excel generado ({len(excel_bytes)} bytes)")
        
        # Generar nombre de archivo
        cleanup_name = investigator_name.lower().replace(' ', '_').replace('.', '')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"produccion_{cleanup_name}_{timestamp}.xlsx"
        
        # Guardar
        file_path = output_dir / filename
        logger.info(f"[EXCEL] Guardando archivo: {file_path}")
        
        with open(file_path, 'wb') as f:
            f.write(excel_bytes)
        
        logger.info(f"✓ [EXCEL] Archivo guardado exitosamente: {file_path}")
        
        return {
            'filename': filename,
            'file_path': str(file_path),
            'investigator_name': investigator_name,
            'total_publications': len(publications),
            'size_mb': len(excel_bytes) / (1024 * 1024),
        }
    
    except Exception as e:
        logger.error(f"✗ [EXCEL] Error fatal: {str(e)}", exc_info=True)
        raise
