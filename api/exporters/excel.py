"""
Exportador de Excel para resultados de cobertura de revistas (Serial Title API).

GENERA archivos .xlsx en memoria usando openpyxl con:
  - Encabezados con estilo
  - Columnas auto-ajustadas
  - Colores condicionales por estado de la revista
  - Hoja de resumen

LEE archivos .xlsx subidos para extraer ISSNs de la primera columna.
"""

import io
import logging
import re
from datetime import datetime
from typing import List

logger_excel = logging.getLogger("excel")

import openpyxl
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter


# ── Paleta de colores ─────────────────────────────────────────────────────────

COLOR_HEADER_BG   = "1F4E79"   # azul oscuro
COLOR_HEADER_FONT = "FFFFFF"   # blanco
COLOR_ACTIVE      = "C6EFCE"   # verde claro
COLOR_DISCONT     = "FFCCCC"   # rojo claro
COLOR_UNKNOWN     = "FFF2CC"   # amarillo claro
COLOR_ERROR       = "E0E0E0"   # gris
COLOR_ALT_ROW     = "EBF3FB"   # azul muy claro (filas alternadas)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ── Columnas del reporte ──────────────────────────────────────────────────────

COLUMNS = [
    ("ISSN",             "issn",            15),
    ("Título",           "title",           45),
    ("Editorial",        "publisher",       30),
    ("Estado",           "status",          14),
    ("¿Descontinuada?",  "is_discontinued", 16),
    ("Desde (año)",      "coverage_from",   13),
    ("Hasta (año)",      "coverage_to",     13),
    ("Áreas temáticas",  "subject_areas",   40),
    ("Source ID Scopus", "source_id",       18),
    ("Error",            "error",           35),
]


# ── Función principal ─────────────────────────────────────────────────────────

def generate_journal_coverage_excel(results: List[dict]) -> bytes:
    """
    Genera un Excel en memoria a partir de los resultados de cobertura.

    Args:
        results: Lista de dicts retornados por SerialTitleExtractor.get_bulk_coverage()

    Returns:
        bytes del archivo .xlsx listo para enviar como respuesta HTTP.
    """
    wb = openpyxl.Workbook()

    # ── Hoja 1: Detalle ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Cobertura de Revistas"

    _write_title_row(ws, len(COLUMNS))
    _write_header_row(ws)
    _write_data_rows(ws, results)
    _auto_adjust_columns(ws)

    # ── Hoja 2: Resumen ───────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Resumen")
    _write_summary_sheet(ws_summary, results)

    # ── Serializar a bytes ────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ── Helpers privados ──────────────────────────────────────────────────────────

def _write_title_row(ws, num_cols: int):
    """Fila 1: título del reporte."""
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = (
        f"Reporte de Cobertura de Revistas en Scopus  —  "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    title_cell.font = Font(bold=True, size=13, color=COLOR_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=num_cols
    )


def _write_header_row(ws):
    """Fila 2: encabezados de columna con estilo."""
    header_fill = PatternFill(
        fill_type="solid", fgColor=COLOR_HEADER_BG
    )
    for col_idx, (header, _, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color=COLOR_HEADER_FONT, size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 30


def _write_data_rows(ws, results: List[dict]):
    """Filas de datos a partir de la fila 3."""
    for row_idx, item in enumerate(results, start=3):
        fill_color = _row_fill_color(item, row_idx)
        row_fill = PatternFill(fill_type="solid", fgColor=fill_color)

        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            value = item.get(key)

            # Serialización especial
            if key == "subject_areas" and isinstance(value, list):
                value = " | ".join(value) if value else None
            elif key == "is_discontinued":
                value = "Sí" if value else "No"

            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )

        ws.row_dimensions[row_idx].height = 20


def _row_fill_color(item: dict, row_idx: int) -> str:
    """Retorna el color de fondo según estado de la revista."""
    if item.get("error"):
        return COLOR_ERROR
    status = (item.get("status") or "").lower()
    if "discontinued" in status or "inactive" in status:
        return COLOR_DISCONT
    if status == "inactiva":
        return "FCE5CD"   # naranja muy claro — inactiva (sin confirmación de Scopus)
    if "active" in status:
        # Alternar tono verde para legibilidad
        return COLOR_ACTIVE if row_idx % 2 == 0 else "D9F0DD"
    return COLOR_ALT_ROW if row_idx % 2 == 0 else "FFFFFF"


def _auto_adjust_columns(ws):
    """Aplica los anchos de columna definidos en COLUMNS."""
    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width


def _write_summary_sheet(ws, results: List[dict]):
    """Hoja de resumen con estadísticas del lote."""
    total     = len(results)
    found     = sum(1 for r in results if not r.get("error") and r.get("title"))
    not_found = sum(1 for r in results if r.get("error") == "Revista no encontrada en Scopus.")
    errors    = sum(1 for r in results if r.get("error") and r.get("error") != "Revista no encontrada en Scopus.")
    active    = sum(1 for r in results if (r.get("status") or "").lower() == "active")
    discont   = sum(1 for r in results if (r.get("status") or "").lower() in ("discontinued", "inactive", "inactiva"))
    unknown   = sum(1 for r in results if (r.get("status") or "").lower() == "unknown" and not r.get("error"))

    header_fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)

    rows = [
        ("Métrica",              "Valor"),
        ("Total ISSNs consultados", total),
        ("Revistas encontradas",    found),
        ("No encontradas (404)",    not_found),
        ("Errores de API",          errors),
        ("",                        ""),
        ("Activas",                 active),
        ("Descontinuadas",          discont),
        ("Estado desconocido",      unknown),
        ("",                        ""),
        ("Fecha de generación", datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    for row_idx, (label, value) in enumerate(rows, start=1):
        cell_a = ws.cell(row=row_idx, column=1, value=label)
        cell_b = ws.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            for cell in (cell_a, cell_b):
                cell.font = Font(bold=True, color=COLOR_HEADER_FONT)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = THIN_BORDER
        elif label:
            cell_a.font = Font(bold=True)
            cell_a.border = THIN_BORDER
            cell_b.border = THIN_BORDER
            cell_b.alignment = Alignment(horizontal="center")
            if row_idx % 2 == 0:
                cell_a.fill = PatternFill(fill_type="solid", fgColor=COLOR_ALT_ROW)
                cell_b.fill = PatternFill(fill_type="solid", fgColor=COLOR_ALT_ROW)


# ── Lector de ISSNs desde archivo Excel subido ───────────────────────────

def read_issns_from_excel(file_bytes: bytes) -> list[str]:
    """
    Lee un archivo .xlsx y extrae todos los ISSNs de la primera columna (A).

    - Omite la primera fila si parece un encabezado (texto, no ISSN).
    - Acepta ISSNs con o sin guion: '2595-3982' o '25953982'.
    - Elimina duplicados preservando el orden de aparición.
    - Ignora celdas vacías.

    Args:
        file_bytes: Contenido binario del .xlsx.

    Returns:
        Lista de ISSNs únicos como strings.

    Raises:
        ValueError: Si el archivo no es válido o no contiene ISSNs reconocibles.
    """
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
    except Exception as e:
        raise ValueError(f"No se pudo leer el archivo Excel: {e}")

    ws = wb.active
    # Patrón: 7-8 caracteres: 4 dígitos, guion opcional, 3 dígitos + dígito/X
    ISSN_RE = re.compile(r"^\d{4}-?\d{3}[\dXx]$", re.IGNORECASE)

    seen: set[str] = set()
    issns: list[str] = []

    for row_idx, row in enumerate(
        ws.iter_rows(min_col=1, max_col=1, values_only=True), start=1
    ):
        raw = row[0]
        if raw is None:
            continue
        value = str(raw).strip()
        if not value:
            continue

        # Omitir encabezado textual en la primera fila
        normalized = value.replace("-", "")
        if row_idx == 1 and not ISSN_RE.match(value) and not re.match(r"^[\dXx]{7,8}$", normalized, re.I):
            continue

        # Validar: con guion o sin guion (7-8 chars alfanum)
        if ISSN_RE.match(value) or re.match(r"^[\dXx]{7,8}$", normalized, re.I):
            if value not in seen:
                seen.add(value)
                issns.append(value)

    wb.close()

    if not issns:
        raise ValueError(
            "No se encontraron ISSNs válidos en la columna A del archivo. "
            "Asegúrese de que la columna A contenga ISSNs "
            "(formatos aceptados: 2595-3982 ó 25953982)."
        )

    return issns


# ── Lector de Excel de exportación Scopus ────────────────────────────────────

# Nombres de las columnas que genera el propio exportador de cobertura.
# Al re-subir el Excel resultado como entrada, estas columnas se detectan
# automáticamente y se descartan para evitar duplicados en el nuevo reporte.
_COVERAGE_COL_NAMES: frozenset[str] = frozenset([
    # ── Formato antiguo (export Scopus re-procesado) ──
    "Revista en Scopus",
    "Título oficial (Scopus)",
    "Editorial (Scopus)",
    "Estado revista",
    "Periodos de cobertura",
    "¿En cobertura?",
    "ISSN resuelto (Scopus)",
    "E-ISSN resuelto (Scopus)",
    # ── Formato nuevo (hoja 'Cobertura' del propio reporte) ──
    "Fuente",               # columna _source
    "En Scopus",            # journal_found
    "¿En cobertura?",       # in_coverage  (ya arriba, pero por claridad)
    "Estado revista",       # journal_status  (ídem)
    "Revista (Scopus)",     # scopus_journal_title
    "Editorial",            # scopus_publisher
    "Periodos cobertura",   # coverage_periods_str
    "Áreas temáticas",      # journal_subject_areas
    "Encontrado vía",       # journal_found_via
    "ISSN resuelto",        # resolved_issn
    "E-ISSN resuelto",      # resolved_eissn
    # ── número de fila (columna '#') ──
    "#",
])


# Mapeo flexible de nombres de columna del export Scopus → clave interna
# Cada lista incluye primero el nombre del export de Scopus original y luego
# los nombres que usa el propio reporte generado (para re-procesamiento).
_SCOPUS_COL_MAP = {
    "title":          ["title",        "título del artículo"],
    "year":           ["year",          "año"],
    "source_title":   ["source title",  "revista (scopus)"],
    "issn":           ["issn",          "issn resuelto"],
    "eissn":          ["eissn", "e-issn", "electronic issn", "e-issn resuelto"],
    "isbn":           ["isbn"],
    "doi":            ["doi"],
    "document_type":  ["document type", "tipo de publicación"],
    "authors":        ["authors"],
    "eid":            ["eid"],
    "link":           ["link"],
    "language":       ["language of original document", "language"],
    "open_access":    ["open access"],
    "cited_by":       ["cited by"],
    "publisher":      ["publisher"],
}


def _normalize_header(h: str) -> str:
    return str(h).strip().lower()


def _deduplicate_headers(headers: list[str]) -> list[str]:
    """Añade sufijo _2, _3… a nombres de columna duplicados."""
    seen: dict[str, int] = {}
    result = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            result.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 1
            result.append(h)
    return result


def _looks_like_header_row(row_values: tuple) -> bool:
    """
    Devuelve True si la fila parece contener encabezados de columna reales.
    Comprueba que al menos una celda coincida con palabras clave conocidas
    de exportaciones Scopus o del propio reporte generado.
    """
    _KNOWN_HEADERS = {
        "title", "year", "issn", "isbn", "doi", "eid",
        "source title", "document type", "authors", "publisher",
        "cited by", "open access", "language",
        # columnas del reporte antiguo (re-procesamiento)
        "revista en scopus", "estado revista", "¿en cobertura?",
        "título oficial (scopus)", "editorial (scopus)",
        # columnas del reporte nuevo (hoja 'Cobertura')
        "fuente", "en scopus", "título del artículo", "año",
        "tipo de publicación",
        "revista (scopus)", "periodos cobertura", "encontrado vía",
        "áreas temáticas", "issn resuelto", "e-issn resuelto",
    }
    non_empty = [str(c).strip().lower() for c in row_values if c is not None and str(c).strip()]
    return any(v in _KNOWN_HEADERS for v in non_empty)


def read_publications_from_excel(file_bytes: bytes) -> tuple[list[str], list[dict]]:
    """
    Lee un Excel de exportación de Scopus (o el Excel resultado del propio
    endpoint) y retorna los datos de publicaciones.

    Detecta automáticamente si la fila 1 es un título del reporte (celda
    fusionada) y en ese caso usa la fila 2 como encabezados reales.

    Columnas clave detectadas automáticamente (case-insensitive):
      Title, Year, Source title, ISSN, DOI, Document Type, Authors, EID…

    Args:
        file_bytes: Contenido binario del .xlsx.

    Returns:
        Tuple (headers: list[str], rows: list[dict]):
          - headers: nombres originales de columna en orden
            (sin las columnas de cobertura si el archivo es un resultado previo)
          - rows: cada fila como dict con TODOS los valores originales,
                  MÁS claves internas normalizadas (title, year, issn,
                  source_title, doi, document_type, authors)

    Raises:
        ValueError: Si el archivo no es válido o no tiene encabezados.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"No se pudo leer el archivo Excel: {e}")

    ws = wb.active

    # Leer todas las filas en memoria para poder decide cuál es la cabecera
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise ValueError("El archivo Excel está vacío.")

    # Detectar fila de encabezados: si la fila 0 parece un título del reporte
    # (pocas celdas no vacías o no contiene palabras clave), usar la fila 1.
    header_row_idx = 0
    if not _looks_like_header_row(all_rows[0]):
        if len(all_rows) > 1 and _looks_like_header_row(all_rows[1]):
            header_row_idx = 1
            logger_excel.info(
                "[read_publications] Fila 1 detectada como título del reporte; "
                "usando fila 2 como encabezados."
            )
        else:
            logger_excel.warning(
                "[read_publications] No se reconoció una fila de encabezados estándar; "
                "se usará la primera fila como encabezados."
            )

    raw_headers: list[str] = [
        str(c).strip() if c is not None else f"Col_{i}"
        for i, c in enumerate(all_rows[header_row_idx])
    ]
    raw_headers = _deduplicate_headers(raw_headers)

    logger_excel.info(
        f"[read_publications] header_row_idx={header_row_idx}, "
        f"total filas de datos={len(all_rows) - header_row_idx - 1}, "
        f"headers detectados ({len(raw_headers)}): {raw_headers[:8]}{'...' if len(raw_headers)>8 else ''}"
    )

    rows: list[dict] = []
    norm_headers = {_normalize_header(h): h for h in raw_headers}

    for row in all_rows[header_row_idx + 1:]:
        # Skip filas completamente vacías
        if all(c is None for c in row):
            continue

        row_dict: dict = {}
        # Guardar todos los valores originales
        for col_name, val in zip(raw_headers, row):
            row_dict[col_name] = val if val is not None else ""

        # Añadir claves internas normalizadas
        for internal_key, candidates in _SCOPUS_COL_MAP.items():
            for cand in candidates:
                orig_col = norm_headers.get(cand)
                if orig_col:
                    row_dict[f"__{internal_key}"] = row_dict.get(orig_col, "")
                    break
            else:
                row_dict[f"__{internal_key}"] = ""

        rows.append(row_dict)

    if not raw_headers:
        raise ValueError("El archivo Excel no tiene encabezados reconocibles.")
    if not rows:
        raise ValueError("El archivo Excel no contiene filas de datos.")

    # Si el Excel proviene de una ejecución anterior del exportador,
    # descartar las columnas de cobertura para que el nuevo reporte
    # las genere limpias (no aparecen duplicadas).
    headers = [h for h in raw_headers if h not in _COVERAGE_COL_NAMES]
    if len(headers) < len(raw_headers):
        dropped = [h for h in raw_headers if h in _COVERAGE_COL_NAMES]
        logger_excel.info(
            f"[read_publications] Excel resultado re-procesado: "
            f"se ignoraron {len(dropped)} columnas previas de cobertura: {dropped}"
        )

    return headers, rows


# ── Generador del Excel de verificación de cobertura ─────────────────────────

# Columnas nuevas que se añaden al reporte (hoja principal, sin áreas temáticas)
# Los nombres de display DEBEN coincidir con _COVERAGE_COL_NAMES (para la detección
# automática al re-procesar el Excel resultado).
_COVERAGE_NEW_COLS = [
    ("Revista en Scopus",           "journal_found",        10),
    ("Título oficial (Scopus)",     "scopus_journal_title", 40),
    ("Editorial (Scopus)",          "scopus_publisher",     28),
    ("Estado revista",              "journal_status",       16),
    ("Periodos de cobertura",       "coverage_periods_str", 34),
    ("¿En cobertura?",              "in_coverage",          26),
    ("ISSN resuelto (Scopus)",      "resolved_issn",        18),
    ("E-ISSN resuelto (Scopus)",    "resolved_eissn",       18),
]

# Columnas fijas de la hoja "Cobertura" (vista limpia)
# (label, row_key, width, halign, wrap_text)
_CLEAN_MAIN_COLS = [
    ("#",                    "_row_num",              5,   "center", False),
    ("Fuente",               "_source",              14,  "center", False),
    ("En Scopus",            "journal_found",         11,  "center", False),
    ("¿En cobertura?",       "in_coverage",           22,  "center", False),
    ("Estado revista",       "journal_status",        15,  "center", False),
    ("Título del artículo",  "__title",               52,  "left",   True),
    ("Año",                  "__year",                 7,  "center", False),
    ("Tipo de publicación",   "__document_type",       20,  "center", False),
    ("Revista (Scopus)",     "scopus_journal_title",  36,  "left",   False),
    ("Editorial",            "scopus_publisher",      24,  "left",   False),
    ("Periodos cobertura",   "coverage_periods_str",  28,  "center", False),
    ("Áreas temáticas",      "journal_subject_areas", 34,  "left",   True),
    ("Encontrado vía",       "journal_found_via",     14,  "center", False),
    ("DOI",                  "__doi",                 40,  "left",   False),
    ("EID",                  "__eid",                 22,  "left",   False),
    ("ISSN resuelto",        "resolved_issn",         14,  "center", False),
    ("E-ISSN resuelto",      "resolved_eissn",        14,  "center", False),
]


def _in_cov_cell_color(in_cov: str) -> tuple[str, str]:
    """Retorna (bg_color, font_color) según el valor de ¿En cobertura?."""
    v = str(in_cov).strip().lower()
    if v == "sí":
        return ("1E6B2F", "FFFFFF")
    if v.startswith("no"):
        return ("922B21", "FFFFFF")
    if v == "sin datos":
        return ("7D6608", "FFFFFF")
    return ("595959", "FFFFFF")


# Palabras clave para detectar columnas de autores → van a la hoja "Autores"
_AUTHOR_COL_KEYWORDS = ("author", "affiliat", "correspondence")

COLOR_IN_COV      = "C6EFCE"   # verde
COLOR_OUT_COV     = "FFCCCC"   # rojo
COLOR_NO_DATA     = "FFF2CC"   # amarillo
COLOR_NOT_FOUND   = "E0E0E0"   # gris
COLOR_FALLBACK    = "FFFF99"   # amarillo vivo — resuelto por fallback (sin ISSN directo)


def _is_author_col(col_name: str) -> bool:
    """True si el encabezado corresponde a una columna de autores/afiliaciones."""
    norm = _normalize_header(col_name)
    return any(kw in norm for kw in _AUTHOR_COL_KEYWORDS)


def _write_sheet_header(ws, all_headers: list[str], title_text: str):
    """Escribe la fila de título (fila 1) y la fila de encabezados (fila 2)."""
    num_cols = len(all_headers)
    header_fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)

    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill = header_fill
    ws.row_dimensions[1].height = 22
    if num_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)

    for col_idx, col_name in enumerate(all_headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 28


def generate_publications_coverage_excel(
    headers: list[str],
    rows: list[dict],
) -> bytes:
    """
    Genera un Excel de verificación de cobertura con las siguientes hojas:
      1. "Cobertura"        – una fila por publicación con datos del artículo/revista
      2. "Autores"          – una fila por AUTOR: publicación combinada + autores Scopus/OpenAlex
      3. "Datos originales" – columnas originales del archivo fuente sin modificar
      4. "Descontinuadas"   – resumen de revistas descontinuadas (condicional)
      5. "Descont. OpenAlex"– detalle de publicaciones en revistas descontinuadas ×OpenAlex (condicional)
      6. "Resumen"          – estadísticas del cruce

    Args:
        headers: Columnas originales del archivo fuente (en orden).
        rows:    Filas enriquecidas por SerialTitleExtractor.check_publications_coverage().

    Returns:
        bytes del .xlsx listo para enviar como respuesta HTTP.
    """
    wb = openpyxl.Workbook()

    # Separar columnas de autores de columnas de artículo/revista
    article_headers = [h for h in headers if not _is_author_col(h)]
    author_headers  = [h for h in headers if _is_author_col(h)]

    # Columnas de cobertura que se agregan a la hoja principal
    cov_col_names = [col[0] for col in _COVERAGE_NEW_COLS]

    logger_excel.info(f"[Excel] Iniciando generación: {len(rows)} filas, {len(headers)} columnas originales")

    # Pre-calcular texto de periodos para cada fila
    for row in rows:
        periods: list = row.get("coverage_periods") or []
        if periods:
            parts = []
            for s, e in periods:
                parts.append(str(s) if s == e else f"{s}–{e}")
            row["coverage_periods_str"] = "  |  ".join(parts)
        else:
            # fallback: si no hay lista pero sí hay from/to
            cf = row.get("coverage_from")
            ct = row.get("coverage_to")
            if cf and ct:
                row["coverage_periods_str"] = f"{cf}–{ct}"
            elif cf:
                row["coverage_periods_str"] = f"{cf}–actual"
            else:
                # Preservar valor previo si ya venía del Excel anterior;
                # solo poner "—" si realmente no hay nada.
                prev = str(row.get("coverage_periods_str") or "").strip()
                row["coverage_periods_str"] = prev if prev and prev != "—" else "—"

    # Pre-crear objetos de estilo reutilizables para evitar crear miles de instancias
    _fill_pool: dict[str, PatternFill] = {}
    def _fill(color: str) -> PatternFill:
        if color not in _fill_pool:
            _fill_pool[color] = PatternFill(fill_type="solid", fgColor=color)
        return _fill_pool[color]

    _align_pool: dict[tuple, Alignment] = {}
    def _align(h: str, v: str = "center", wrap: bool = False) -> Alignment:
        k = (h, v, wrap)
        if k not in _align_pool:
            _align_pool[k] = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
        return _align_pool[k]

    _font_pool: dict[tuple, Font] = {}
    def _font(bold: bool = False, color: str = "000000", size: int = 10, italic: bool = False) -> Font:
        k = (bold, color, size, italic)
        if k not in _font_pool:
            _font_pool[k] = Font(bold=bold, color=color, size=size, italic=italic)
        return _font_pool[k]

    # ── Hoja 1: Cobertura ─────────────────────────────────────────────────────
    logger_excel.info(f"[Excel] Escribiendo hoja 'Cobertura' ({len(rows)} filas)...")
    ws = wb.active
    ws.title = "Cobertura"

    # ── Hoja 1: Cobertura (vista limpia) ─────────────────────────────────────
    clean_col_labels = [c[0] for c in _CLEAN_MAIN_COLS]
    _write_sheet_header(
        ws, clean_col_labels,
        f"Verificación de Cobertura Scopus  —  "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  —  "
        f"{len(rows)} publicaciones",
    )

    # Encabezados de columnas «resultado» con fondo azul más oscuro para destacarlas
    _res_keys   = {"journal_found", "in_coverage", "journal_status"}
    _dark_hdr   = PatternFill(fill_type="solid", fgColor="1A3A4A")
    for col_idx, (_, key, *_) in enumerate(_CLEAN_MAIN_COLS, start=1):
        if key in _res_keys:
            ws.cell(row=2, column=col_idx).fill = _dark_hdr

    for row_idx, row in enumerate(rows, start=3):
        in_cov    = str(row.get("in_coverage",   "") or "")
        jstatus   = str(row.get("journal_status", "") or "")
        found     = bool(row.get("journal_found", False))
        found_via = str(row.get("journal_found_via") or "issn")
        base_color = _coverage_row_color(in_cov, found, row_idx, found_via=found_via)
        base_fill  = _fill(base_color)

        for col_idx, (_, col_key, _, halign, wrap) in enumerate(_CLEAN_MAIN_COLS, start=1):
            if col_key == "_row_num":
                val = row_idx - 2
            elif col_key == "journal_found":
                val = "Sí" if found else "No"
            else:
                raw = row.get(col_key)
                val = "—" if (raw is None or raw == "") else raw

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = _align(halign, "center", wrap)

            # Color específico por tipo de columna
            if col_key == "in_coverage":
                bg, fg = _in_cov_cell_color(in_cov)
                cell.fill = _fill(bg)
                cell.font = _font(bold=True, color=fg)
            elif col_key == "journal_status":
                sl = jstatus.strip().lower()
                if sl in ("discontinued", "inactive"):
                    cell.fill = _fill("C0392B")
                    cell.font = _font(bold=True, color="FFFFFF")
                elif sl == "inactiva":
                    cell.fill = _fill("CA6F1E")
                    cell.font = _font(bold=True, color="FFFFFF")
                elif sl == "active":
                    cell.fill = _fill("1E8449")
                    cell.font = _font(bold=True, color="FFFFFF")
                else:
                    cell.fill = _fill("797D7F")
                    cell.font = _font(color="FFFFFF")
            elif col_key == "journal_found":
                cell.fill = _fill("1A5276") if found else _fill("922B21")
                cell.font = _font(bold=True, color="FFFFFF")
            elif col_key in ("resolved_issn", "resolved_eissn"):
                if val not in ("—", "", None):
                    cell.fill = _fill(COLOR_FALLBACK)
                    cell.font = _font(bold=True)
                else:
                    cell.fill = _fill("F2F2F2")
                    cell.font = _font(color="AAAAAA")
            elif col_key == "_row_num":
                cell.fill = _fill("D5D8DC")
                cell.font = _font(color="555555", size=9)
            elif col_key == "_source":
                src = str(val or "").strip()
                if src == "OpenAlex BD":
                    cell.fill = _fill("1A5276")
                    cell.font = _font(bold=True, color="FFFFFF", size=9)
                else:
                    cell.fill = _fill("145A32")
                    cell.font = _font(bold=True, color="FFFFFF", size=9)
            elif col_key == "__document_type":
                dt = str(val or "").strip().lower()
                if dt in ("article", "review", "short survey"):
                    cell.fill = _fill("0B5345")
                    cell.font = _font(bold=True, color="FFFFFF", size=9)
                elif "conference" in dt or "proceedings" in dt:
                    cell.fill = _fill("154360")
                    cell.font = _font(bold=True, color="FFFFFF", size=9)
                elif "book" in dt or "chapter" in dt:
                    cell.fill = _fill("4A235A")
                    cell.font = _font(bold=True, color="FFFFFF", size=9)
                elif dt not in ("", "—"):
                    cell.fill = _fill("424949")
                    cell.font = _font(bold=True, color="FFFFFF", size=9)
                else:
                    cell.fill = _fill("F2F2F2")
                    cell.font = _font(color="AAAAAA")
            else:
                cell.fill = base_fill

        ws.row_dimensions[row_idx].height = 20

    # Anchos de columna
    for col_idx, (_, _, width, *_) in enumerate(_CLEAN_MAIN_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Leyenda embebida a la derecha — columna separada del área de datos
    _leg_col = len(_CLEAN_MAIN_COLS) + 2
    _leg_col_letter = get_column_letter(_leg_col)
    ws.column_dimensions[_leg_col_letter].width = 36
    _legend_items = [
        ("Leyenda  ¿En cobertura?",            None,           None,     True),
        ("✓  Sí  — publicación cubierta",      "1E6B2F",       "FFFFFF", False),
        ("✗  No  — fuera de cobertura",        "922B21",       "FFFFFF", False),
        ("?  Sin datos suficientes",            "7D6608",       "FFFFFF", False),
        ("—  No encontrada en Scopus",          "595959",       "FFFFFF", False),
        ("",                                    None,           None,     False),
        ("Leyenda  Estado revista",             None,           None,     True),
        ("Active",                              "1E8449",       "FFFFFF", False),
        ("Inactiva (sin confirmar activa)",      "CA6F1E",       "FFFFFF", False),
        ("Discontinued / Inactive",             "C0392B",       "FFFFFF", False),
        ("",                                    None,           None,     False),
        ("⚠ Amarillo = resuelto sin ISSN",     COLOR_FALLBACK, "000000", False),
        ("(por título / DOI / EID — verificar)",COLOR_FALLBACK, "000000", False),
    ]
    for li, (txt, bg, fg, is_hdr) in enumerate(_legend_items, start=2):
        lc = ws.cell(row=li, column=_leg_col, value=txt)
        lc.border = THIN_BORDER
        lc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if is_hdr:
            lc.font = Font(bold=True, color="FFFFFF")
            lc.fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
        elif bg:
            lc.fill = PatternFill(fill_type="solid", fgColor=bg)
            lc.font = Font(color=fg)

    ws.freeze_panes = "E3"   # fija columnas #, En Scopus, ¿En cobertura?, Estado revista

    # ── Hoja 2: Autores ───────────────────────────────────────────────────────
    logger_excel.info(f"[Excel] Escribiendo hoja 'Autores' ({len(rows)} publicaciones)...")
    ws_auth = wb.create_sheet("Autores")

    # Columnas de publicación que se combinarán verticalmente
    _AUTH_PUB_COLS = [
        # (label, key, width, halign, wrap)
        ("#",                    "_row_num",              5,   "center", False),
        ("Fuente",               "_source",              14,  "center", False),
        ("¿En cobertura?",       "in_coverage",          22,  "center", False),
        ("Estado revista",       "journal_status",       16,  "center", False),
        ("Título del artículo",  "__title",              50,  "left",   True ),
        ("Año",                  "__year",                7,  "center", False),
        ("Tipo",                 "__document_type",      20,  "center", False),
        ("Revista (Scopus)",     "scopus_journal_title", 34,  "left",   False),
        ("DOI",                  "__doi",                38,  "left",   False),
    ]
    _n_pc          = len(_AUTH_PUB_COLS)
    _col_auth_num  = _n_pc + 1   # "# Autor"
    _col_auth_name = _n_pc + 2   # "Nombre autor (Scopus)"
    _col_auth_afil = _n_pc + 3   # "Afiliación (Scopus)"  — opcional
    _col_auth_oa   = _n_pc + 4   # "Autores (OpenAlex)"  — combinada por pub

    # Detectar si el fuente tiene columna "Authors with affiliations"
    _afil_header = next(
        (h for h in headers if "with affiliation" in _normalize_header(h)), None
    )
    _auth_labels = (
        [c[0] for c in _AUTH_PUB_COLS]
        + ["# Autor", "Nombre autor (Scopus)", "Afiliación (Scopus)", "Autores (OpenAlex)"]
    )
    _write_sheet_header(
        ws_auth, _auth_labels,
        f"Autores de publicaciones  —  {len(rows)} publicaciones  —  "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
    )

    # Encabezados de columnas de autor con color diferenciado
    _auth_hdr_scopus = PatternFill(fill_type="solid", fgColor="2C3E50")
    _auth_hdr_oa     = PatternFill(fill_type="solid", fgColor="16537e")
    for _ci in (_col_auth_num, _col_auth_name, _col_auth_afil):
        ws_auth.cell(row=2, column=_ci).fill = _auth_hdr_scopus
    ws_auth.cell(row=2, column=_col_auth_oa).fill = _auth_hdr_oa

    # Anchos de columna
    for _ci, (_, _, _w, _, _) in enumerate(_AUTH_PUB_COLS, start=1):
        ws_auth.column_dimensions[get_column_letter(_ci)].width = _w
    ws_auth.column_dimensions[get_column_letter(_col_auth_num )].width = 8
    ws_auth.column_dimensions[get_column_letter(_col_auth_name)].width = 35
    ws_auth.column_dimensions[get_column_letter(_col_auth_afil)].width = 50
    ws_auth.column_dimensions[get_column_letter(_col_auth_oa  )].width = 46

    _cur_auth_row = 3

    for _pub_idx, _arow in enumerate(rows, start=1):
        # --- Autores Scopus (del export original) ---
        _raw_authors = str(_arow.get("__authors") or "")
        _scopus_names = [p.strip() for p in _raw_authors.split(";") if p.strip()]

        # --- Afiliaciones Scopus (columna "Authors with affiliations") ---
        if _afil_header:
            _raw_afil    = str(_arow.get(_afil_header, "") or "")
            _afil_parts  = [p.strip() for p in _raw_afil.split(";") if p.strip()]
        else:
            _afil_parts  = []

        if not _scopus_names:
            _scopus_names = ["—"]

        _n_auth  = len(_scopus_names)
        _start_r = _cur_auth_row
        _end_r   = _cur_auth_row + _n_auth - 1

        # Colores basados en cobertura (iguales a hoja Cobertura)
        _inc  = str(_arow.get("in_coverage",    "") or "")
        _jst  = str(_arow.get("journal_status", "") or "")
        _fnd  = bool(_arow.get("journal_found", False))
        _fvia = str(_arow.get("journal_found_via") or "issn")
        _bcol = _coverage_row_color(_inc, _fnd, _pub_idx, found_via=_fvia)
        _bfil = _fill(_bcol)

        # ── Columnas de publicación (combinadas verticalmente si >1 autor) ──
        for _ci, (_, _ck, _, _ha, _wr) in enumerate(_AUTH_PUB_COLS, start=1):
            if _ck == "_row_num":
                _val = _pub_idx
            else:
                _raw = _arow.get(_ck)
                _val = "—" if (_raw is None or _raw == "") else _raw

            _cell = ws_auth.cell(row=_start_r, column=_ci, value=_val)
            _cell.border    = THIN_BORDER
            _cell.alignment = _align(_ha, "center", _wr)

            # Estilos idénticos a la hoja "Cobertura"
            if _ck == "in_coverage":
                _bg, _fg = _in_cov_cell_color(_inc)
                _cell.fill = _fill(_bg); _cell.font = _font(bold=True, color=_fg)
            elif _ck == "journal_status":
                _sl = _jst.strip().lower()
                if _sl in ("discontinued", "inactive"):
                    _cell.fill = _fill("C0392B"); _cell.font = _font(bold=True, color="FFFFFF")
                elif _sl == "inactiva":
                    _cell.fill = _fill("CA6F1E"); _cell.font = _font(bold=True, color="FFFFFF")
                elif _sl == "active":
                    _cell.fill = _fill("1E8449"); _cell.font = _font(bold=True, color="FFFFFF")
                else:
                    _cell.fill = _fill("797D7F"); _cell.font = _font(color="FFFFFF")
            elif _ck == "_row_num":
                _cell.fill = _fill("D5D8DC"); _cell.font = _font(color="555555", size=9)
            elif _ck == "_source":
                _src = str(_val or "").strip()
                if _src == "OpenAlex BD":
                    _cell.fill = _fill("1A5276"); _cell.font = _font(bold=True, color="FFFFFF", size=9)
                else:
                    _cell.fill = _fill("145A32"); _cell.font = _font(bold=True, color="FFFFFF", size=9)
            elif _ck == "__document_type":
                _dt = str(_val or "").strip().lower()
                if _dt in ("article", "review", "short survey"):
                    _cell.fill = _fill("0B5345"); _cell.font = _font(bold=True, color="FFFFFF", size=9)
                elif "conference" in _dt or "proceedings" in _dt:
                    _cell.fill = _fill("154360"); _cell.font = _font(bold=True, color="FFFFFF", size=9)
                elif "book" in _dt or "chapter" in _dt:
                    _cell.fill = _fill("4A235A"); _cell.font = _font(bold=True, color="FFFFFF", size=9)
                elif _dt not in ("", "—"):
                    _cell.fill = _fill("424949"); _cell.font = _font(bold=True, color="FFFFFF", size=9)
                else:
                    _cell.fill = _bfil
            else:
                _cell.fill = _bfil

            if _n_auth > 1:
                ws_auth.merge_cells(
                    start_row=_start_r, start_column=_ci,
                    end_row=_end_r,     end_column=_ci,
                )

        # ── Columna "Autores (OpenAlex)" — combinada por publicación ──
        _oa_d   = _arow.get("_openalex") or {}
        _oa_str = str(_oa_d.get("oa_authors") or "")
        _oac    = ws_auth.cell(row=_start_r, column=_col_auth_oa,
                               value=_oa_str if _oa_str else "—")
        _oac.border    = THIN_BORDER
        _oac.alignment = _align("left", "center", True)
        _oac.fill      = _fill("E8F4FD") if _oa_str else _fill("F2F2F2")
        _oac.font      = _font() if _oa_str else _font(color="AAAAAA", italic=True)
        if _n_auth > 1:
            ws_auth.merge_cells(
                start_row=_start_r, start_column=_col_auth_oa,
                end_row=_end_r,     end_column=_col_auth_oa,
            )

        # ── Filas individuales de autores ──
        for _ai, _aname in enumerate(_scopus_names, start=1):
            _r = _start_r + _ai - 1

            # # Autor
            _nc = ws_auth.cell(row=_r, column=_col_auth_num, value=_ai)
            _nc.fill = _fill("D5D8DC"); _nc.font = _font(bold=True, color="333333", size=9)
            _nc.alignment = _align("center"); _nc.border = THIN_BORDER

            # Nombre autor (Scopus)
            _ac = ws_auth.cell(row=_r, column=_col_auth_name, value=_aname)
            _ac.fill = _bfil; _ac.font = _font()
            _ac.alignment = _align("left", "center", False); _ac.border = THIN_BORDER

            # Afiliación (Scopus) — si existe la columna
            _afval = _afil_parts[_ai - 1] if (_ai - 1) < len(_afil_parts) else "—"
            _afc   = ws_auth.cell(row=_r, column=_col_auth_afil, value=_afval)
            _afc.fill = _bfil; _afc.font = _font(size=9)
            _afc.alignment = _align("left", "center", True); _afc.border = THIN_BORDER

            ws_auth.row_dimensions[_r].height = 18

        _cur_auth_row = _end_r + 1

    ws_auth.freeze_panes = "D3"   # fija columnas #, Fuente, ¿En cobertura?

    # ── Hoja 3 (antes 2): Datos originales ────────────────────────────────────
    logger_excel.info(f"[Excel] Escribiendo hoja 'Datos originales' ({len(rows)} filas)...")
    # Todas las columnas originales del Excel fuente (artículo + autores/afiliaciones)
    ws_orig = wb.create_sheet("Datos originales")
    _orig_headers = ["#"] + list(headers)
    _write_sheet_header(
        ws_orig, _orig_headers,
        f"Datos originales del archivo fuente  —  {len(rows)} publicaciones",
    )

    ws_orig.column_dimensions["A"].width = 5
    for col_idx, col_name in enumerate(headers, start=2):
        col_letter = get_column_letter(col_idx)
        norm = _normalize_header(col_name)
        if "title" in norm and "source" not in norm:
            ws_orig.column_dimensions[col_letter].width = 52
        elif "source title" in norm:
            ws_orig.column_dimensions[col_letter].width = 30
        elif norm in ("year", "volume", "issue", "cited by", "art. no."):
            ws_orig.column_dimensions[col_letter].width = 9
        elif norm in ("doi", "link", "eid"):
            ws_orig.column_dimensions[col_letter].width = 40
        elif "with affiliation" in norm:
            ws_orig.column_dimensions[col_letter].width = 55
        elif "affiliation" in norm:
            ws_orig.column_dimensions[col_letter].width = 45
        elif "author" in norm:
            ws_orig.column_dimensions[col_letter].width = 40
        elif norm in ("issn", "isbn", "coden", "eissn"):
            ws_orig.column_dimensions[col_letter].width = 14
        else:
            ws_orig.column_dimensions[col_letter].width = 20

    _num_fill  = _fill("D5D8DC")
    _even_fill = _fill("EBF3FB")
    _odd_fill  = _fill("FFFFFF")
    _num_font  = _font(color="555555", size=9)
    _align_ctr = _align("center")
    _align_lft = _align("left")
    for row_idx, row in enumerate(rows, start=3):
        nc = ws_orig.cell(row=row_idx, column=1, value=row_idx - 2)
        nc.fill  = _num_fill
        nc.font  = _num_font
        nc.alignment = _align_ctr
        nc.border = THIN_BORDER

        alt = _even_fill if row_idx % 2 == 0 else _odd_fill
        for col_idx, col_name in enumerate(headers, start=2):
            val  = row.get(col_name, "")
            cell = ws_orig.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = alt
            cell.border = THIN_BORDER
            cell.alignment = _align_lft
        ws_orig.row_dimensions[row_idx].height = 16

    ws_orig.freeze_panes = "B3"

    # Letras de columnas clave en la hoja "Cobertura" (determinadas por _CLEAN_MAIN_COLS)
    def _col_for_key(k: str) -> str:
        for i, (_, ck, *_) in enumerate(_CLEAN_MAIN_COLS, start=1):
            if ck == k:
                return get_column_letter(i)
        return "A"

    col_revista   = _col_for_key("journal_found")   # "En Scopus" → Sí/No
    col_status    = _col_for_key("journal_status")  # "Estado revista"
    col_cobertura = _col_for_key("in_coverage")     # "¿En cobertura?"
    col_source    = _col_for_key("_source")         # "Fuente" → Scopus Export / OpenAlex BD
    last_data_row = 2 + len(rows)                   # fila 1=título, 2=encabezados, 3..N=datos

    # ── Hoja 4: Descontinuadas ────────────────────────────────────────────────
    # Una fila por REVISTA única descontinuada (no por artículo)
    _DISC_STATUSES = {"discontinued", "inactive", "inactiva"}

    # Log de diagnóstico
    status_counts: dict[str, int] = {}
    for r in rows:
        s = str(r.get("journal_status", "") or "").strip()
        status_counts[s] = status_counts.get(s, 0) + 1
    logger_excel.info(f"[Excel] Distribución journal_status: {status_counts}")

    # Deduplicar revistas descontinuadas: clave = título Scopus (o ISSN como fallback)
    _seen_disc: dict[str, dict] = {}   # clave → datos de la revista
    _disc_pub_count: dict[str, int] = {}  # clave → # publicaciones afectadas
    for r in rows:
        if str(r.get("journal_status", "")).strip().lower() not in _DISC_STATUSES:
            continue
        key = (
            str(r.get("scopus_journal_title") or "").strip().lower()
            or str(r.get("__issn") or r.get("issn") or "").strip()
            or str(r.get("__source_title") or "").strip().lower()
        )
        if not key:
            continue
        _disc_pub_count[key] = _disc_pub_count.get(key, 0) + 1
        if key not in _seen_disc:
            _seen_disc[key] = {
                "issn":              str(r.get("__issn") or r.get("issn") or "—"),
                "titulo_scopus":     r.get("scopus_journal_title") or r.get("__source_title") or "—",
                "editorial":         r.get("scopus_publisher") or "—",
                "estado":            r.get("journal_status") or "Discontinued",
                "periodos":          r.get("coverage_periods_str") or "—",
                "areas":             r.get("journal_subject_areas") or "—",
            }

    _disc_journals = [
        {**datos, "publicaciones_afectadas": _disc_pub_count[clave]}
        for clave, datos in _seen_disc.items()
    ]
    # Ordenar por título
    _disc_journals.sort(key=lambda x: str(x.get("titulo_scopus", "")).lower())

    logger_excel.info(f"[Excel] Revistas descontinuadas únicas: {len(_disc_journals)} / {len(_seen_disc)} (sobre {len(rows)} filas)")

    if _disc_journals:
        logger_excel.info(f"[Excel] Escribiendo hoja 'Descontinuadas' ({len(_disc_journals)} revistas)...")
        ws_disc = wb.create_sheet("Descontinuadas")
        _DISC_COLS = [
            ("ISSN",                   "issn",                   14),
            ("Título oficial (Scopus)", "titulo_scopus",          44),
            ("Editorial",              "editorial",               28),
            ("Estado",                 "estado",                  16),
            ("Periodos de cobertura",  "periodos",                34),
            ("Áreas temáticas",        "areas",                   46),
            ("# Publicaciones afectadas", "publicaciones_afectadas", 20),
        ]
        disc_col_names = [c[0] for c in _DISC_COLS]
        _write_sheet_header(
            ws_disc, disc_col_names,
            f"Revistas Descontinuadas / Inactivas en Scopus  —  "
            f"{len(_disc_journals)} revistas únicas  —  "
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        )
        disc_fill = PatternFill(fill_type="solid", fgColor=COLOR_DISCONT)
        alt_fill  = PatternFill(fill_type="solid", fgColor="FFE8E8")

        for row_idx, jrn in enumerate(_disc_journals, start=3):
            row_fill = disc_fill if row_idx % 2 == 0 else alt_fill
            for col_idx, (_, col_key, _) in enumerate(_DISC_COLS, start=1):
                val = jrn.get(col_key, "—")
                cell = ws_disc.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = row_fill
                cell.border = THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="center" if col_key == "publicaciones_afectadas" else "left",
                    vertical="center",
                    wrap_text=True,
                )
            ws_disc.row_dimensions[row_idx].height = 20

        for col_idx, (_, _, width) in enumerate(_DISC_COLS, start=1):
            ws_disc.column_dimensions[get_column_letter(col_idx)].width = width
        ws_disc.freeze_panes = "A3"

    # ── Hoja 5: Descontinuadas detalle + OpenAlex ─────────────────────────────
    # Una fila por PUBLICACIÓN (no por revista) en revista descontinuada,
    # cruzada con openalex_records por DOI (campo _openalex adjuntado en pipeline).
    _OA_SHEET_COLS = [
        # (label,                       key_en_row,            key_en_oa,             width)
        # --- datos del Excel original / Scopus ---
        ("DOI",                          "__doi",               None,                  38),
        ("Título",                       "__title",             None,                  50),
        ("Año",                          "__year",              None,                   8),
        ("Revista (Scopus)",             "scopus_journal_title",None,                  34),
        ("Estado revista",               "journal_status",      None,                  16),
        ("Periodos de cobertura",        "coverage_periods_str",None,                  28),
        ("¿En cobertura?",               "in_coverage",         None,                  16),
        # --- datos OpenAlex ---
        ("OpenAlex ID",                  None,                  "oa_work_id",          36),
        ("Título (OpenAlex)",            None,                  "oa_title",            50),
        ("Año (OpenAlex)",               None,                  "oa_year",              8),
        ("Autores",                      None,                  "oa_authors",          50),
        ("Acceso Abierto",               None,                  "oa_open_access",      14),
        ("Estado OA",                    None,                  "oa_oa_status",        16),
        ("Citas (OpenAlex)",             None,                  "oa_citations",        14),
        ("URL",                          None,                  "oa_url",              38),
    ]

    # Filas para la hoja: publicaciones en revistas descontinuadas
    _disc_pub_rows = [
        r for r in rows
        if str(r.get("journal_status", "")).strip().lower() in _DISC_STATUSES
    ]
    _disc_pub_rows.sort(key=lambda r: (
        str(r.get("scopus_journal_title") or r.get("__source_title") or "").lower(),
        str(r.get("__year") or ""),
    ))

    if _disc_pub_rows:
        logger_excel.info(f"[Excel] Escribiendo hoja 'Descont. OpenAlex' ({len(_disc_pub_rows)} filas)...")
        ws_oa = wb.create_sheet("Descont. OpenAlex")
        n_matched_oa = sum(1 for r in _disc_pub_rows if r.get("_openalex"))

        _write_sheet_header(
            ws_oa,
            [c[0] for c in _OA_SHEET_COLS],
            (
                f"Publicaciones en Revistas Descontinuadas/Inactivas — Cruce OpenAlex  —  "
                f"{len(_disc_pub_rows)} publicaciones  |  "
                f"{n_matched_oa} con datos OpenAlex  —  "
                f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ),
        )

        # Columna pivote: separador entre datos Scopus/Excel y datos OpenAlex
        _SPL_COL = 7  # índice (1-based) de la última col Scopus = "¿En cobertura?"
        _oa_header_fill = PatternFill(fill_type="solid", fgColor="16537e")  # azul OpenAlex

        # Colorear encabezados de columnas OpenAlex con un tono diferente
        for col_idx, (label, _, oa_key, _) in enumerate(_OA_SHEET_COLS, start=1):
            if oa_key is not None:
                cell = ws_oa.cell(row=2, column=col_idx)
                cell.fill = _oa_header_fill

        disc_row_fill  = PatternFill(fill_type="solid", fgColor=COLOR_DISCONT)
        disc_alt_fill  = PatternFill(fill_type="solid", fgColor="FFE8E8")
        oa_match_fill  = PatternFill(fill_type="solid", fgColor="E8F4FD")   # azul muy claro → tiene OA
        oa_no_fill     = PatternFill(fill_type="solid", fgColor="F5F5F5")   # gris → sin OA

        for row_idx, row in enumerate(_disc_pub_rows, start=3):
            oa = row.get("_openalex")  # dict con datos OpenAlex o None
            base_fill = disc_row_fill if row_idx % 2 == 0 else disc_alt_fill

            for col_idx, (_, row_key, oa_key, _) in enumerate(_OA_SHEET_COLS, start=1):
                if row_key is not None:
                    # Columna del Excel original / Scopus
                    val = row.get(row_key, "")
                    if val is None or val == "":
                        val = "—"
                    cell = ws_oa.cell(row=row_idx, column=col_idx, value=val)
                    cell.fill = base_fill
                else:
                    # Columna de OpenAlex
                    val = oa.get(oa_key, "") if oa else ""
                    if val is None or val == "":
                        val = "—"
                    cell = ws_oa.cell(row=row_idx, column=col_idx, value=val)
                    cell.fill = oa_match_fill if oa else oa_no_fill

                cell.border = THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=(col_idx in (2, 10, 11)),  # Títulos y Autores
                )
                if row_key == "in_coverage" or row_key == "journal_status":
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if oa_key in ("oa_year", "oa_citations"):
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws_oa.row_dimensions[row_idx].height = 20

        # Anchos
        for col_idx, (_, _, _, width) in enumerate(_OA_SHEET_COLS, start=1):
            ws_oa.column_dimensions[get_column_letter(col_idx)].width = width

        ws_oa.freeze_panes = "B3"

        # Nota informativa en una celda por debajo de los datos
        note_row = len(_disc_pub_rows) + 3
        note_cell = ws_oa.cell(
            row=note_row, column=1,
            value=(
                f"ℹ Filas con fondo azul claro = publicación encontrada en openalex_records por DOI. "
                f"Filas con fondo gris = no encontrada en BD OpenAlex. "
                f"Total: {n_matched_oa}/{len(_disc_pub_rows)} emparejadas."
            )
        )
        note_cell.font = Font(italic=True, color="444444", size=9)
        note_cell.alignment = Alignment(wrap_text=True)
        ws_oa.merge_cells(
            start_row=note_row, start_column=1,
            end_row=note_row, end_column=len(_OA_SHEET_COLS)
        )
        ws_oa.row_dimensions[note_row].height = 28

    # ── Hoja 6: Resumen ───────────────────────────────────────────────────────
    logger_excel.info(f"[Excel] Escribiendo hoja 'Resumen'...")
    ws_sum = wb.create_sheet("Resumen")
    _write_publications_summary(ws_sum, col_revista, col_status, col_cobertura, col_source, last_data_row)

    logger_excel.info(f"[Excel] Serializando workbook a bytes...")
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    data = buffer.read()
    logger_excel.info(f"[Excel] Generación completa: {len(data):,} bytes")
    return data


def _coverage_row_color(
    in_cov: str, found: bool, row_idx: int, found_via: str = "issn"
) -> str:
    """Color de fondo para la fila según el resultado de cobertura.

    - Verde  : encontrada vía ISSN y en cobertura
    - Rojo   : encontrada vía ISSN y fuera de cobertura
    - Amarillo vivo (FFFF99) : encontrada por fallback (title/doi/eid) — puede ser inexacta
    - Amarillo suave : sin datos de cobertura
    - Gris   : no encontrada
    """
    if not found:
        return COLOR_NOT_FOUND
    # Fallback: resuelto sin ISSN directo → amarillo para indicar menor confianza
    if found_via and found_via not in ("issn", ""):
        return COLOR_FALLBACK if row_idx % 2 == 0 else "FFFFCC"
    if in_cov == "Sí":
        return COLOR_IN_COV if row_idx % 2 == 0 else "D9F0DD"
    if in_cov.startswith("No"):
        return COLOR_OUT_COV if row_idx % 2 == 0 else "FFE0E0"
    return COLOR_NO_DATA if row_idx % 2 == 0 else "FFFBE6"


def _write_publications_summary(
    ws,
    col_revista: str,
    col_status: str,
    col_cobertura: str,
    col_source: str,
    last_data_row: int,
):
    """
    Hoja de resumen con fórmulas COUNTIF/COUNTIFS que apuntan a la hoja 'Cobertura'.
    Al modificar valores manualmente en esa hoja, los totales se actualizan solos.

    Args:
        col_revista:    Letra de la columna 'Revista en Scopus' en la hoja Cobertura.
        col_status:     Letra de la columna 'Estado revista' en la hoja Cobertura.
        col_cobertura:  Letra de la columna '¿En cobertura?' en la hoja Cobertura.
        col_source:     Letra de la columna 'Fuente' en la hoja Cobertura.
        last_data_row:  Última fila con datos en la hoja Cobertura.
    """
    header_fill  = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
    data_range_r = f"'Cobertura'!{col_revista}{3}:{col_revista}{last_data_row}"
    data_range_s = f"'Cobertura'!{col_status}{3}:{col_status}{last_data_row}"
    data_range_c = f"'Cobertura'!{col_cobertura}{3}:{col_cobertura}{last_data_row}"
    data_range_f = f"'Cobertura'!{col_source}{3}:{col_source}{last_data_row}"

    # (etiqueta, fórmula_o_valor,  color_fill)
    rows_def = [
        # ─ Encabezado ─────────────────────────────────────────────────────
        ("Métrica",                              "Valor",                    "__header"),
        # ─ Totales ────────────────────────────────────────────────────
        ("Total publicaciones analizadas",        f"=COUNTA({data_range_c})",  None),
        ("Revistas encontradas en Scopus",        f'=COUNTIF({data_range_r},"S\u00ed")',    None),
        ("   de las cuales: activas",             f'=COUNTIF(\'Cobertura\'!{col_status}{3}:{col_status}{last_data_row},"Active")', None),
        ("Revistas NO encontradas",               f'=COUNTIF({data_range_r},"No")',   None),
        ("",                                      "",                         None),
        # ─ Cobertura por estado ───────────────────────────────────────
        ("✓  Publicación EN cobertura",            f'=COUNTIF({data_range_c},"S\u00ed")',             COLOR_IN_COV),
        ("✗  ANTES de cobertura",                 f'=COUNTIF({data_range_c},"No (antes de cobertura)")',       COLOR_OUT_COV),
        ("✗  DESPUÉS de cobertura",               f'=COUNTIF({data_range_c},"No (despu\u00e9s de cobertura)")',     COLOR_OUT_COV),
        ("✗  LAGUNA de cobertura",                f'=COUNTIF({data_range_c},"No (laguna de cobertura)")',       COLOR_OUT_COV),
        ("?  Sin datos suficientes",               f'=COUNTIF({data_range_c},"Sin datos")',                      COLOR_NO_DATA),
        ("",                                      "",                         None),
        # ─ % cobertura ────────────────────────────────────────────────────
        ("% en cobertura (sobre total)",          '=IFERROR(B7/B2,0)',         None),
        ("",                                      "",                         None),        # ─ Revistas descontinuadas ────────────────────────────────────────────
        ("Revistas descontinuadas en Scopus",
             f'=COUNTIF({data_range_s},"Discontinued")+COUNTIF({data_range_s},"Inactive")',
             COLOR_DISCONT),
        ("   con publicación dentro de cobertura",
             f'=COUNTIFS({data_range_s},"Discontinued",{data_range_c},"S\u00ed")'
             f'+COUNTIFS({data_range_s},"Inactive",{data_range_c},"S\u00ed")',
             COLOR_DISCONT),
        ("",                                      "",                         None),        # ─ Metadatos ────────────────────────────────────────────────────
        ("Fecha generación",                     datetime.now().strftime("%d/%m/%Y %H:%M"), None),
        ("",                                      "",                         None),
        # ─ Por fuente ─────────────────────────────────────────────────────────
        ("Publicaciones del Excel de Scopus",
             f'=COUNTIF({data_range_f},"Scopus Export")',
             "145A32"),
        ("Publicaciones de OpenAlex BD",
             f'=COUNTIF({data_range_f},"OpenAlex BD")',
             "1A5276"),
        ("   OA BD · en cobertura",
             f'=COUNTIFS({data_range_f},"OpenAlex BD",{data_range_c},"S\u00ed")',
             "1A5276"),
        ("   OA BD · NO encontradas en Scopus",
             f'=COUNTIFS({data_range_f},"OpenAlex BD",{data_range_r},"No")',
             "1A5276"),
    ]

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 58

    # Encabezado de la tercera columna (leyenda de valores válidos)
    legend_header = ws.cell(row=1, column=3, value="Valores válidos en '¿En cobertura?'")
    legend_header.font  = Font(bold=True, color=COLOR_HEADER_FONT)
    legend_header.fill  = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
    legend_header.alignment = Alignment(horizontal="center")
    legend_header.border = THIN_BORDER

    legend_items = [
        ("Sí",                         COLOR_IN_COV),
        ("No (antes de cobertura)",     COLOR_OUT_COV),
        ("No (después de cobertura)",   COLOR_OUT_COV),
        ("No (laguna de cobertura)",    COLOR_OUT_COV),
        ("Sin datos",                   COLOR_NO_DATA),
        ("—",                          COLOR_NOT_FOUND),
        ("[Amarillo] Sin ISSN – resuelto por título/DOI/EID. Verificar manualmente.", COLOR_FALLBACK),
    ]

    for row_idx, (label, formula, color) in enumerate(rows_def, start=1):
        ca = ws.cell(row=row_idx, column=1, value=label)
        cb = ws.cell(row=row_idx, column=2, value=formula)

        if color == "__header":
            for c in (ca, cb):
                c.font      = Font(bold=True, color=COLOR_HEADER_FONT)
                c.fill      = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
                c.alignment = Alignment(horizontal="center")
                c.border    = THIN_BORDER
        elif label:
            ca.font   = Font(bold=True)
            ca.border = THIN_BORDER
            cb.border = THIN_BORDER
            cb.alignment = Alignment(horizontal="center")
            # Formato porcentaje para la fila de %
            if "%)" in label or "%" in label:
                cb.number_format = "0.0%"
            if color:
                fill = PatternFill(fill_type="solid", fgColor=color)
                ca.fill = fill
                cb.fill = fill

        # Leyenda en columna C (solo las primeras filas)
        if row_idx == 1:
            pass  # ya escrito arriba
        elif 1 <= row_idx - 1 <= len(legend_items):
            leg_val, leg_color = legend_items[row_idx - 2] if row_idx >= 2 and (row_idx - 2) < len(legend_items) else ("", None)
            cl = ws.cell(row=row_idx, column=3, value=leg_val)
            cl.border = THIN_BORDER
            cl.alignment = Alignment(horizontal="left", wrap_text=True)
            if leg_color:
                cl.fill = PatternFill(fill_type="solid", fgColor=leg_color)

    # Nota al pie
    note_row = len(rows_def) + 2
    note = ws.cell(
        row=note_row, column=1,
        value=(
            "⚠ Puede cambiar manualmente los valores en la columna '¿En cobertura?' "
            "de la hoja 'Cobertura' y este resumen se actualizará automáticamente."
        )
    )
    note.font      = Font(italic=True, color="666666", size=9)
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)
    ws.row_dimensions[note_row].height = 30


def get_column_letter_offset(col_letter: str, offset: int) -> str:
    """Devuelve la letra de columna desplazada 'offset' posiciones desde col_letter."""
    col_idx = 0
    for ch in col_letter.upper():
        col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
    return get_column_letter(col_idx + offset)







