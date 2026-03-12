"""
Genera el Excel de verificación de cobertura por publicación.

Función pública:
  generate_publications_coverage_excel(headers, rows) → bytes
  get_column_letter_offset(col_letter, offset) → str
"""
import io
import logging
from datetime import datetime
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ._styles import (
    COLOR_HEADER_BG,
    COLOR_HEADER_FONT,
    COLOR_DISCONT,
    COLOR_IN_COV,
    COLOR_OUT_COV,
    COLOR_NO_DATA,
    COLOR_NOT_FOUND,
    COLOR_FALLBACK,
    THIN_BORDER,
    _fill,
    _align,
    _font,
    _in_cov_cell_color,
    _coverage_row_color,
    _is_author_col,
    _normalize_header,
    _write_sheet_header,
)

logger = logging.getLogger("excel")

# ── Columnas de cobertura añadidas al reporte ─────────────────────────────────

_COVERAGE_NEW_COLS = [
    ("Revista en Scopus",           "journal_found",        10),
    ("Título oficial (Scopus)",     "scopus_journal_title", 40),
    ("Editorial (Scopus)",          "scopus_publisher",     28),
    ("Estado revista",              "journal_status",       16),
    ("Periodos de cobertura",       "coverage_periods_str", 34),
    ("¿En cobertura?",              "in_coverage",          26),
    ("ISSN resuelto (Scopus)",      "resolved_issn",        18),
    ("E-ISSN resuelto (Scopus)",    "resolved_eissn",       18),
]

# ── Columnas de la hoja "Cobertura" (vista limpia) ────────────────────────────

_CLEAN_MAIN_COLS = [
    ("#",                    "_row_num",              5,   "center", False),
    ("Fuente",               "_source",              14,  "center", False),
    ("En Scopus",            "journal_found",         11,  "center", False),
    ("¿En cobertura?",       "in_coverage",           22,  "center", False),
    ("Estado revista",       "journal_status",        15,  "center", False),
    ("Título del artículo",  "__title",               52,  "left",   True),
    ("Año",                  "__year",                 7,  "center", False),
    ("Tipo de publicación",  "__document_type",        20,  "center", False),
    ("Revista (Scopus)",     "scopus_journal_title",  36,  "left",   False),
    ("Editorial",            "scopus_publisher",      24,  "left",   False),
    ("Periodos cobertura",   "coverage_periods_str",  28,  "center", False),
    ("Áreas temáticas",      "journal_subject_areas", 34,  "left",   True),
    ("Encontrado vía",       "journal_found_via",     14,  "center", False),
    ("DOI",                  "__doi",                 40,  "left",   False),
    ("EID",                  "__eid",                 22,  "left",   False),
    ("ISSN resuelto",        "resolved_issn",         14,  "center", False),
    ("E-ISSN resuelto",      "resolved_eissn",        14,  "center", False),
]

# ── Columnas de la hoja "Autores" ─────────────────────────────────────────────

_AUTH_PUB_COLS = [
    ("#",                    "_row_num",              5,   "center", False),
    ("Fuente",               "_source",              14,  "center", False),
    ("¿En cobertura?",       "in_coverage",          22,  "center", False),
    ("Estado revista",       "journal_status",       16,  "center", False),
    ("Título del artículo",  "__title",              50,  "left",   True),
    ("Año",                  "__year",                7,  "center", False),
    ("Tipo",                 "__document_type",      20,  "center", False),
    ("Revista (Scopus)",     "scopus_journal_title", 34,  "left",   False),
    ("DOI",                  "__doi",                38,  "left",   False),
]

# ── Columnas de la hoja "Descont. OpenAlex" ───────────────────────────────────

_OA_SHEET_COLS = [
    ("DOI",                   "__doi",                None,                  38),
    ("Título",                "__title",              None,                  50),
    ("Año",                   "__year",               None,                   8),
    ("Revista (Scopus)",      "scopus_journal_title", None,                  34),
    ("Estado revista",        "journal_status",       None,                  16),
    ("Periodos de cobertura", "coverage_periods_str", None,                  28),
    ("¿En cobertura?",        "in_coverage",          None,                  16),
    ("OpenAlex ID",           None,                   "oa_work_id",          36),
    ("Título (OpenAlex)",     None,                   "oa_title",            50),
    ("Año (OpenAlex)",        None,                   "oa_year",              8),
    ("Autores",               None,                   "oa_authors",          50),
    ("Acceso Abierto",        None,                   "oa_open_access",      14),
    ("Estado OA",             None,                   "oa_oa_status",        16),
    ("Citas (OpenAlex)",      None,                   "oa_citations",        14),
    ("URL",                   None,                   "oa_url",              38),
]

# ── Estados de revistas descontinuadas ───────────────────────────────────────

_DISC_STATUSES = {"discontinued", "inactive", "inactiva"}


# ── Función principal ─────────────────────────────────────────────────────────

def generate_publications_coverage_excel(
    headers: List[str],
    rows: List[dict],
) -> bytes:
    """
    Genera un Excel de verificación de cobertura con las siguientes hojas:
      1. "Cobertura"         — una fila por publicación (vista limpia)
      2. "Autores"           — una fila por autor: publicación + autores
      3. "Datos originales"  — columnas originales del archivo fuente
      4. "Descontinuadas"    — resumen de revistas descontinuadas (condicional)
      5. "Descont. OpenAlex" — publicaciones en revistas descont. × OpenAlex
      6. "Resumen"           — estadísticas con fórmulas COUNTIF

    Args:
        headers: Columnas originales del archivo fuente (en orden).
        rows:    Filas enriquecidas por check_publications_coverage().

    Returns:
        bytes del .xlsx listo para enviar como respuesta HTTP.
    """
    wb = openpyxl.Workbook()

    logger.info(
        f"[Excel] Iniciando generación: {len(rows)} filas, "
        f"{len(headers)} columnas originales"
    )

    # Pre-calcular texto de periodos para cada fila
    for row in rows:
        periods = row.get("coverage_periods") or []
        if periods:
            parts = [str(s) if s == e else f"{s}–{e}" for s, e in periods]
            row["coverage_periods_str"] = "  |  ".join(parts)
        else:
            cf  = row.get("coverage_from")
            ct  = row.get("coverage_to")
            if cf and ct:
                row["coverage_periods_str"] = f"{cf}–{ct}"
            elif cf:
                row["coverage_periods_str"] = f"{cf}–actual"
            else:
                prev = str(row.get("coverage_periods_str") or "").strip()
                row["coverage_periods_str"] = prev if prev and prev != "—" else "—"

    # ── Hoja 1: Cobertura ─────────────────────────────────────────────────────
    logger.info(f"[Excel] Escribiendo hoja 'Cobertura' ({len(rows)} filas)...")
    ws = wb.active
    ws.title = "Cobertura"

    clean_col_labels = [c[0] for c in _CLEAN_MAIN_COLS]
    _write_sheet_header(
        ws, clean_col_labels,
        f"Verificación de Cobertura Scopus  —  "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  —  "
        f"{len(rows)} publicaciones",
    )

    _res_keys = {"journal_found", "in_coverage", "journal_status"}
    _dark_hdr = PatternFill(fill_type="solid", fgColor="1A3A4A")
    for col_idx, (_, key, *_) in enumerate(_CLEAN_MAIN_COLS, start=1):
        if key in _res_keys:
            ws.cell(row=2, column=col_idx).fill = _dark_hdr

    for row_idx, row in enumerate(rows, start=3):
        in_cov    = str(row.get("in_coverage",    "") or "")
        jstatus   = str(row.get("journal_status", "") or "")
        found     = bool(row.get("journal_found", False))
        found_via = str(row.get("journal_found_via") or "issn")
        base_color = _coverage_row_color(in_cov, found, row_idx, found_via=found_via)
        base_fill  = _fill(base_color)

        for col_idx, (_, col_key, _, halign, wrap) in enumerate(_CLEAN_MAIN_COLS, start=1):
            if col_key == "_row_num":
                val = row_idx - 2
            elif col_key == "journal_found":
                val = "Sí" if found else "No"
            else:
                raw = row.get(col_key)
                val = "—" if (raw is None or raw == "") else raw

            cell           = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = THIN_BORDER
            cell.alignment = _align(halign, "center", wrap)

            if col_key == "in_coverage":
                bg, fg = _in_cov_cell_color(in_cov)
                cell.fill = _fill(bg)
                cell.font = _font(bold=True, color=fg)
            elif col_key == "journal_status":
                sl = jstatus.strip().lower()
                if sl in ("discontinued", "inactive"):
                    cell.fill = _fill("C0392B"); cell.font = _font(bold=True, color="FFFFFF")
                elif sl == "inactiva":
                    cell.fill = _fill("CA6F1E"); cell.font = _font(bold=True, color="FFFFFF")
                elif sl == "active":
                    cell.fill = _fill("1E8449"); cell.font = _font(bold=True, color="FFFFFF")
                else:
                    cell.fill = _fill("797D7F"); cell.font = _font(color="FFFFFF")
            elif col_key == "journal_found":
                cell.fill = _fill("1A5276") if found else _fill("922B21")
                cell.font = _font(bold=True, color="FFFFFF")
            elif col_key in ("resolved_issn", "resolved_eissn"):
                if val not in ("—", "", None):
                    cell.fill = _fill(COLOR_FALLBACK); cell.font = _font(bold=True)
                else:
                    cell.fill = _fill("F2F2F2"); cell.font = _font(color="AAAAAA")
            elif col_key == "_row_num":
                cell.fill = _fill("D5D8DC"); cell.font = _font(color="555555", size=9)
            elif col_key == "_source":
                src = str(val or "").strip()
                if src == "OpenAlex BD":
                    cell.fill = _fill("1A5276"); cell.font = _font(bold=True, color="FFFFFF", size=9)
                else:
                    cell.fill = _fill("145A32"); cell.font = _font(bold=True, color="FFFFFF", size=9)
            elif col_key == "__document_type":
                dt = str(val or "").strip().lower()
                if dt in ("article", "review", "short survey"):
                    cell.fill = _fill("0B5345"); cell.font = _font(bold=True, color="FFFFFF", size=9)
                elif "conference" in dt or "proceedings" in dt:
                    cell.fill = _fill("154360"); cell.font = _font(bold=True, color="FFFFFF", size=9)
                elif "book" in dt or "chapter" in dt:
                    cell.fill = _fill("4A235A"); cell.font = _font(bold=True, color="FFFFFF", size=9)
                elif dt not in ("", "—"):
                    cell.fill = _fill("424949"); cell.font = _font(bold=True, color="FFFFFF", size=9)
                else:
                    cell.fill = _fill("F2F2F2"); cell.font = _font(color="AAAAAA")
            else:
                cell.fill = base_fill

        ws.row_dimensions[row_idx].height = 20

    for col_idx, (_, _, width, *_) in enumerate(_CLEAN_MAIN_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Leyenda embebida a la derecha
    _leg_col = len(_CLEAN_MAIN_COLS) + 2
    _leg_col_letter = get_column_letter(_leg_col)
    ws.column_dimensions[_leg_col_letter].width = 36
    _legend_items = [
        ("Leyenda  ¿En cobertura?",             None,           None,     True),
        ("✓  Sí  — publicación cubierta",       "1E6B2F",       "FFFFFF", False),
        ("✗  No  — fuera de cobertura",         "922B21",       "FFFFFF", False),
        ("?  Sin datos suficientes",             "7D6608",       "FFFFFF", False),
        ("—  No encontrada en Scopus",           "595959",       "FFFFFF", False),
        ("",                                     None,           None,     False),
        ("Leyenda  Estado revista",              None,           None,     True),
        ("Active",                               "1E8449",       "FFFFFF", False),
        ("Inactiva (sin confirmar activa)",       "CA6F1E",       "FFFFFF", False),
        ("Discontinued / Inactive",              "C0392B",       "FFFFFF", False),
        ("",                                     None,           None,     False),
        ("⚠ Amarillo = resuelto sin ISSN",      COLOR_FALLBACK, "000000", False),
        ("(por título / DOI / EID — verificar)", COLOR_FALLBACK, "000000", False),
    ]
    for li, (txt, bg, fg, is_hdr) in enumerate(_legend_items, start=2):
        lc = ws.cell(row=li, column=_leg_col, value=txt)
        lc.border    = THIN_BORDER
        lc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if is_hdr:
            lc.font = Font(bold=True, color="FFFFFF")
            lc.fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
        elif bg:
            lc.fill = PatternFill(fill_type="solid", fgColor=bg)
            lc.font = Font(color=fg)

    ws.freeze_panes = "E3"

    # ── Hoja 2: Autores ───────────────────────────────────────────────────────
    logger.info(f"[Excel] Escribiendo hoja 'Autores' ({len(rows)} publicaciones)...")
    ws_auth = wb.create_sheet("Autores")

    _n_pc          = len(_AUTH_PUB_COLS)
    _col_auth_num  = _n_pc + 1
    _col_auth_name = _n_pc + 2
    _col_auth_afil = _n_pc + 3
    _col_auth_oa   = _n_pc + 4

    _afil_header = next(
        (h for h in headers if "with affiliation" in _normalize_header(h)), None
    )
    _auth_labels = (
        [c[0] for c in _AUTH_PUB_COLS]
        + ["# Autor", "Nombre autor (Scopus)", "Afiliación (Scopus)", "Autores (OpenAlex)"]
    )
    _write_sheet_header(
        ws_auth, _auth_labels,
        f"Autores de publicaciones  —  {len(rows)} publicaciones  —  "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
    )

    _auth_hdr_scopus = PatternFill(fill_type="solid", fgColor="2C3E50")
    _auth_hdr_oa     = PatternFill(fill_type="solid", fgColor="16537e")
    for _ci in (_col_auth_num, _col_auth_name, _col_auth_afil):
        ws_auth.cell(row=2, column=_ci).fill = _auth_hdr_scopus
    ws_auth.cell(row=2, column=_col_auth_oa).fill = _auth_hdr_oa

    for _ci, (_, _, _w, _, _) in enumerate(_AUTH_PUB_COLS, start=1):
        ws_auth.column_dimensions[get_column_letter(_ci)].width = _w
    ws_auth.column_dimensions[get_column_letter(_col_auth_num )].width = 8
    ws_auth.column_dimensions[get_column_letter(_col_auth_name)].width = 35
    ws_auth.column_dimensions[get_column_letter(_col_auth_afil)].width = 50
    ws_auth.column_dimensions[get_column_letter(_col_auth_oa  )].width = 46

    _cur_auth_row = 3

    for _pub_idx, _arow in enumerate(rows, start=1):
        _raw_authors  = str(_arow.get("__authors") or "")
        _scopus_names = [p.strip() for p in _raw_authors.split(";") if p.strip()]

        if _afil_header:
            _raw_afil   = str(_arow.get(_afil_header, "") or "")
            _afil_parts = [p.strip() for p in _raw_afil.split(";") if p.strip()]
        else:
            _afil_parts = []

        if not _scopus_names:
            _scopus_names = ["—"]

        _n_auth  = len(_scopus_names)
        _start_r = _cur_auth_row
        _end_r   = _cur_auth_row + _n_auth - 1

        _inc  = str(_arow.get("in_coverage",    "") or "")
        _jst  = str(_arow.get("journal_status", "") or "")
        _fnd  = bool(_arow.get("journal_found", False))
        _fvia = str(_arow.get("journal_found_via") or "issn")
        _bcol = _coverage_row_color(_inc, _fnd, _pub_idx, found_via=_fvia)
        _bfil = _fill(_bcol)

        for _ci, (_, _ck, _, _ha, _wr) in enumerate(_AUTH_PUB_COLS, start=1):
            if _ck == "_row_num":
                _val = _pub_idx
            else:
                _raw = _arow.get(_ck)
                _val = "—" if (_raw is None or _raw == "") else _raw

            _cell           = ws_auth.cell(row=_start_r, column=_ci, value=_val)
            _cell.border    = THIN_BORDER
            _cell.alignment = _align(_ha, "center", _wr)

            if _ck == "in_coverage":
                _bg, _fg = _in_cov_cell_color(_inc)
                _cell.fill = _fill(_bg); _cell.font = _font(bold=True, color=_fg)
            elif _ck == "journal_status":
                _sl = _jst.strip().lower()
                if _sl in ("discontinued", "inactive"):
                    _cell.fill = _fill("C0392B"); _cell.font = _font(bold=True, color="FFFFFF")
                elif _sl == "inactiva":
                    _cell.fill = _fill("CA6F1E"); _cell.font = _font(bold=True, color="FFFFFF")
                elif _sl == "active":
                    _cell.fill = _fill("1E8449"); _cell.font = _font(bold=True, color="FFFFFF")
                else:
                    _cell.fill = _fill("797D7F"); _cell.font = _font(color="FFFFFF")
            elif _ck == "_row_num":
                _cell.fill = _fill("D5D8DC"); _cell.font = _font(color="555555", size=9)
            elif _ck == "_source":
                _src = str(_val or "").strip()
                if _src == "OpenAlex BD":
                    _cell.fill = _fill("1A5276"); _cell.font = _font(bold=True, color="FFFFFF", size=9)
                else:
                    _cell.fill = _fill("145A32"); _cell.font = _font(bold=True, color="FFFFFF", size=9)
            elif _ck == "__document_type":
                _dt = str(_val or "").strip().lower()
                if _dt in ("article", "review", "short survey"):
                    _cell.fill = _fill("0B5345"); _cell.font = _font(bold=True, color="FFFFFF", size=9)
                elif "conference" in _dt or "proceedings" in _dt:
                    _cell.fill = _fill("154360"); _cell.font = _font(bold=True, color="FFFFFF", size=9)
                elif "book" in _dt or "chapter" in _dt:
                    _cell.fill = _fill("4A235A"); _cell.font = _font(bold=True, color="FFFFFF", size=9)
                elif _dt not in ("", "—"):
                    _cell.fill = _fill("424949"); _cell.font = _font(bold=True, color="FFFFFF", size=9)
                else:
                    _cell.fill = _bfil
            else:
                _cell.fill = _bfil

            if _n_auth > 1:
                ws_auth.merge_cells(
                    start_row=_start_r, start_column=_ci,
                    end_row=_end_r,     end_column=_ci,
                )

        _oa_d   = _arow.get("_openalex") or {}
        _oa_str = str(_oa_d.get("oa_authors") or "")
        _oac    = ws_auth.cell(row=_start_r, column=_col_auth_oa,
                                value=_oa_str if _oa_str else "—")
        _oac.border    = THIN_BORDER
        _oac.alignment = _align("left", "center", True)
        _oac.fill      = _fill("E8F4FD") if _oa_str else _fill("F2F2F2")
        _oac.font      = _font() if _oa_str else _font(color="AAAAAA", italic=True)
        if _n_auth > 1:
            ws_auth.merge_cells(
                start_row=_start_r, start_column=_col_auth_oa,
                end_row=_end_r,     end_column=_col_auth_oa,
            )

        for _ai, _aname in enumerate(_scopus_names, start=1):
            _r   = _start_r + _ai - 1
            _nc  = ws_auth.cell(row=_r, column=_col_auth_num, value=_ai)
            _nc.fill = _fill("D5D8DC"); _nc.font = _font(bold=True, color="333333", size=9)
            _nc.alignment = _align("center"); _nc.border = THIN_BORDER

            _ac  = ws_auth.cell(row=_r, column=_col_auth_name, value=_aname)
            _ac.fill = _bfil; _ac.font = _font()
            _ac.alignment = _align("left", "center", False); _ac.border = THIN_BORDER

            _afval = _afil_parts[_ai - 1] if (_ai - 1) < len(_afil_parts) else "—"
            _afc   = ws_auth.cell(row=_r, column=_col_auth_afil, value=_afval)
            _afc.fill = _bfil; _afc.font = _font(size=9)
            _afc.alignment = _align("left", "center", True); _afc.border = THIN_BORDER

            ws_auth.row_dimensions[_r].height = 18

        _cur_auth_row = _end_r + 1

    ws_auth.freeze_panes = "D3"

    # ── Hoja 3: Datos originales ──────────────────────────────────────────────
    logger.info(f"[Excel] Escribiendo hoja 'Datos originales' ({len(rows)} filas)...")
    ws_orig = wb.create_sheet("Datos originales")
    _orig_headers = ["#"] + list(headers)
    _write_sheet_header(
        ws_orig, _orig_headers,
        f"Datos originales del archivo fuente  —  {len(rows)} publicaciones",
    )

    ws_orig.column_dimensions["A"].width = 5
    for col_idx, col_name in enumerate(headers, start=2):
        col_letter = get_column_letter(col_idx)
        norm = _normalize_header(col_name)
        if "title" in norm and "source" not in norm:
            ws_orig.column_dimensions[col_letter].width = 52
        elif "source title" in norm:
            ws_orig.column_dimensions[col_letter].width = 30
        elif norm in ("year", "volume", "issue", "cited by", "art. no."):
            ws_orig.column_dimensions[col_letter].width = 9
        elif norm in ("doi", "link", "eid"):
            ws_orig.column_dimensions[col_letter].width = 40
        elif "with affiliation" in norm:
            ws_orig.column_dimensions[col_letter].width = 55
        elif "affiliation" in norm:
            ws_orig.column_dimensions[col_letter].width = 45
        elif "author" in norm:
            ws_orig.column_dimensions[col_letter].width = 40
        elif norm in ("issn", "isbn", "coden", "eissn"):
            ws_orig.column_dimensions[col_letter].width = 14
        else:
            ws_orig.column_dimensions[col_letter].width = 20

    _num_fill  = _fill("D5D8DC")
    _even_fill = _fill("EBF3FB")
    _odd_fill  = _fill("FFFFFF")
    _num_font  = _font(color="555555", size=9)
    _align_ctr = _align("center")
    _align_lft = _align("left")

    for row_idx, row in enumerate(rows, start=3):
        nc           = ws_orig.cell(row=row_idx, column=1, value=row_idx - 2)
        nc.fill      = _num_fill
        nc.font      = _num_font
        nc.alignment = _align_ctr
        nc.border    = THIN_BORDER

        alt = _even_fill if row_idx % 2 == 0 else _odd_fill
        for col_idx, col_name in enumerate(headers, start=2):
            val  = row.get(col_name, "")
            cell = ws_orig.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = alt
            cell.border    = THIN_BORDER
            cell.alignment = _align_lft
        ws_orig.row_dimensions[row_idx].height = 16

    ws_orig.freeze_panes = "B3"

    # Letras de columnas clave (hoja "Cobertura"), para fórmulas del Resumen
    def _col_for_key(k: str) -> str:
        for i, (_, ck, *_) in enumerate(_CLEAN_MAIN_COLS, start=1):
            if ck == k:
                return get_column_letter(i)
        return "A"

    col_revista   = _col_for_key("journal_found")
    col_status    = _col_for_key("journal_status")
    col_cobertura = _col_for_key("in_coverage")
    col_source    = _col_for_key("_source")
    last_data_row = 2 + len(rows)

    # ── Hoja 4: Descontinuadas (una fila por REVISTA única) ───────────────────
    _seen_disc: dict = {}
    _disc_pub_count: dict = {}
    for r in rows:
        if str(r.get("journal_status", "")).strip().lower() not in _DISC_STATUSES:
            continue
        key = (
            str(r.get("scopus_journal_title") or "").strip().lower()
            or str(r.get("__issn") or r.get("issn") or "").strip()
            or str(r.get("__source_title") or "").strip().lower()
        )
        if not key:
            continue
        _disc_pub_count[key] = _disc_pub_count.get(key, 0) + 1
        if key not in _seen_disc:
            _seen_disc[key] = {
                "issn":          str(r.get("__issn") or r.get("issn") or "—"),
                "titulo_scopus": r.get("scopus_journal_title") or r.get("__source_title") or "—",
                "editorial":     r.get("scopus_publisher") or "—",
                "estado":        r.get("journal_status") or "Discontinued",
                "periodos":      r.get("coverage_periods_str") or "—",
                "areas":         r.get("journal_subject_areas") or "—",
            }

    _disc_journals = [
        {**datos, "publicaciones_afectadas": _disc_pub_count[clave]}
        for clave, datos in _seen_disc.items()
    ]
    _disc_journals.sort(key=lambda x: str(x.get("titulo_scopus", "")).lower())

    status_counts: dict = {}
    for r in rows:
        s = str(r.get("journal_status", "") or "").strip()
        status_counts[s] = status_counts.get(s, 0) + 1
    logger.info(f"[Excel] Distribución journal_status: {status_counts}")
    logger.info(f"[Excel] Revistas descontinuadas únicas: {len(_disc_journals)}")

    if _disc_journals:
        logger.info(
            f"[Excel] Escribiendo hoja 'Descontinuadas' ({len(_disc_journals)} revistas)..."
        )
        ws_disc = wb.create_sheet("Descontinuadas")
        _DISC_COLS = [
            ("ISSN",                       "issn",                   14),
            ("Título oficial (Scopus)",     "titulo_scopus",          44),
            ("Editorial",                  "editorial",               28),
            ("Estado",                     "estado",                  16),
            ("Periodos de cobertura",       "periodos",                34),
            ("Áreas temáticas",            "areas",                   46),
            ("# Publicaciones afectadas",  "publicaciones_afectadas", 20),
        ]
        disc_col_names = [c[0] for c in _DISC_COLS]
        _write_sheet_header(
            ws_disc, disc_col_names,
            f"Revistas Descontinuadas / Inactivas en Scopus  —  "
            f"{len(_disc_journals)} revistas únicas  —  "
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        )
        disc_fill = PatternFill(fill_type="solid", fgColor=COLOR_DISCONT)
        alt_fill  = PatternFill(fill_type="solid", fgColor="FFE8E8")

        for row_idx, jrn in enumerate(_disc_journals, start=3):
            row_fill = disc_fill if row_idx % 2 == 0 else alt_fill
            for col_idx, (_, col_key, _) in enumerate(_DISC_COLS, start=1):
                val  = jrn.get(col_key, "—")
                cell = ws_disc.cell(row=row_idx, column=col_idx, value=val)
                cell.fill      = row_fill
                cell.border    = THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="center" if col_key == "publicaciones_afectadas" else "left",
                    vertical="center",
                    wrap_text=True,
                )
            ws_disc.row_dimensions[row_idx].height = 20

        for col_idx, (_, _, width) in enumerate(_DISC_COLS, start=1):
            ws_disc.column_dimensions[get_column_letter(col_idx)].width = width
        ws_disc.freeze_panes = "A3"

    # ── Hoja 5: Descont. OpenAlex ─────────────────────────────────────────────
    _disc_pub_rows = [
        r for r in rows
        if str(r.get("journal_status", "")).strip().lower() in _DISC_STATUSES
    ]
    _disc_pub_rows.sort(key=lambda r: (
        str(r.get("scopus_journal_title") or r.get("__source_title") or "").lower(),
        str(r.get("__year") or ""),
    ))

    if _disc_pub_rows:
        logger.info(
            f"[Excel] Escribiendo hoja 'Descont. OpenAlex' "
            f"({len(_disc_pub_rows)} filas)..."
        )
        ws_oa = wb.create_sheet("Descont. OpenAlex")
        n_matched_oa = sum(1 for r in _disc_pub_rows if r.get("_openalex"))

        _write_sheet_header(
            ws_oa,
            [c[0] for c in _OA_SHEET_COLS],
            (
                f"Publicaciones en Revistas Descontinuadas/Inactivas — Cruce OpenAlex  —  "
                f"{len(_disc_pub_rows)} publicaciones  |  "
                f"{n_matched_oa} con datos OpenAlex  —  "
                f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ),
        )

        _oa_header_fill = PatternFill(fill_type="solid", fgColor="16537e")
        for col_idx, (_, _, oa_key, _) in enumerate(_OA_SHEET_COLS, start=1):
            if oa_key is not None:
                ws_oa.cell(row=2, column=col_idx).fill = _oa_header_fill

        disc_row_fill = PatternFill(fill_type="solid", fgColor=COLOR_DISCONT)
        disc_alt_fill = PatternFill(fill_type="solid", fgColor="FFE8E8")
        oa_match_fill = PatternFill(fill_type="solid", fgColor="E8F4FD")
        oa_no_fill    = PatternFill(fill_type="solid", fgColor="F5F5F5")

        for row_idx, row in enumerate(_disc_pub_rows, start=3):
            oa        = row.get("_openalex")
            base_fill = disc_row_fill if row_idx % 2 == 0 else disc_alt_fill

            for col_idx, (_, row_key, oa_key, _) in enumerate(_OA_SHEET_COLS, start=1):
                if row_key is not None:
                    val = row.get(row_key, "")
                    if val is None or val == "":
                        val = "—"
                    cell      = ws_oa.cell(row=row_idx, column=col_idx, value=val)
                    cell.fill = base_fill
                else:
                    val = oa.get(oa_key, "") if oa else ""
                    if val is None or val == "":
                        val = "—"
                    cell      = ws_oa.cell(row=row_idx, column=col_idx, value=val)
                    cell.fill = oa_match_fill if oa else oa_no_fill

                cell.border    = THIN_BORDER
                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=(col_idx in (2, 10, 11)),
                )
                if row_key in ("in_coverage", "journal_status") or oa_key in ("oa_year", "oa_citations"):
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            ws_oa.row_dimensions[row_idx].height = 20

        for col_idx, (_, _, _, width) in enumerate(_OA_SHEET_COLS, start=1):
            ws_oa.column_dimensions[get_column_letter(col_idx)].width = width

        ws_oa.freeze_panes = "B3"

        note_row  = len(_disc_pub_rows) + 3
        note_cell = ws_oa.cell(
            row=note_row, column=1,
            value=(
                f"ℹ Filas con fondo azul claro = publicación encontrada en openalex_records por DOI. "
                f"Filas con fondo gris = no encontrada en BD OpenAlex. "
                f"Total: {n_matched_oa}/{len(_disc_pub_rows)} emparejadas."
            ),
        )
        note_cell.font      = Font(italic=True, color="444444", size=9)
        note_cell.alignment = Alignment(wrap_text=True)
        ws_oa.merge_cells(
            start_row=note_row, start_column=1,
            end_row=note_row,   end_column=len(_OA_SHEET_COLS),
        )
        ws_oa.row_dimensions[note_row].height = 28

    # ── Hoja 6: Resumen ───────────────────────────────────────────────────────
    logger.info("[Excel] Escribiendo hoja 'Resumen'...")
    ws_sum = wb.create_sheet("Resumen")
    _write_publications_summary(
        ws_sum, col_revista, col_status, col_cobertura, col_source, last_data_row
    )

    logger.info("[Excel] Serializando workbook a bytes...")
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    data = buffer.read()
    logger.info(f"[Excel] Generación completa: {len(data):,} bytes")
    return data


# ── Resumen con fórmulas ──────────────────────────────────────────────────────

def _write_publications_summary(
    ws,
    col_revista:   str,
    col_status:    str,
    col_cobertura: str,
    col_source:    str,
    last_data_row: int,
) -> None:
    """
    Hoja de resumen con fórmulas COUNTIF/COUNTIFS que apuntan a la hoja 'Cobertura'.
    """
    header_fill  = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
    data_range_r = f"'Cobertura'!{col_revista}{3}:{col_revista}{last_data_row}"
    data_range_s = f"'Cobertura'!{col_status}{3}:{col_status}{last_data_row}"
    data_range_c = f"'Cobertura'!{col_cobertura}{3}:{col_cobertura}{last_data_row}"
    data_range_f = f"'Cobertura'!{col_source}{3}:{col_source}{last_data_row}"

    rows_def = [
        ("Métrica",                                  "Valor",                                                          "__header"),
        ("Total publicaciones analizadas",            f"=COUNTA({data_range_c})",                                       None),
        ("Revistas encontradas en Scopus",            f'=COUNTIF({data_range_r},"S\u00ed")',                            None),
        ("   de las cuales: activas",                 f'=COUNTIF(\'Cobertura\'!{col_status}3:{col_status}{last_data_row},"Active")', None),
        ("Revistas NO encontradas",                   f'=COUNTIF({data_range_r},"No")',                                  None),
        ("",                                          "",                                                               None),
        ("✓  Publicación EN cobertura",               f'=COUNTIF({data_range_c},"S\u00ed")',                            COLOR_IN_COV),
        ("✗  ANTES de cobertura",                    f'=COUNTIF({data_range_c},"No (antes de cobertura)")',            COLOR_OUT_COV),
        ("✗  DESPUÉS de cobertura",                  f'=COUNTIF({data_range_c},"No (despu\u00e9s de cobertura)")',     COLOR_OUT_COV),
        ("✗  LAGUNA de cobertura",                   f'=COUNTIF({data_range_c},"No (laguna de cobertura)")',           COLOR_OUT_COV),
        ("?  Sin datos suficientes",                  f'=COUNTIF({data_range_c},"Sin datos")',                          COLOR_NO_DATA),
        ("",                                          "",                                                               None),
        ("% en cobertura (sobre total)",              "=IFERROR(B7/B2,0)",                                              None),
        ("",                                          "",                                                               None),
        ("Revistas descontinuadas en Scopus",
             f'=COUNTIF({data_range_s},"Discontinued")+COUNTIF({data_range_s},"Inactive")',
             COLOR_DISCONT),
        ("   con publicación dentro de cobertura",
             f'=COUNTIFS({data_range_s},"Discontinued",{data_range_c},"S\u00ed")'
             f'+COUNTIFS({data_range_s},"Inactive",{data_range_c},"S\u00ed")',
             COLOR_DISCONT),
        ("",                                          "",                                                               None),
        ("Fecha generación",                         datetime.now().strftime("%d/%m/%Y %H:%M"),                        None),
        ("",                                          "",                                                               None),
        ("Publicaciones del Excel de Scopus",
             f'=COUNTIF({data_range_f},"Scopus Export")',
             "145A32"),
        ("Publicaciones de OpenAlex BD",
             f'=COUNTIF({data_range_f},"OpenAlex BD")',
             "1A5276"),
        ("   OA BD · en cobertura",
             f'=COUNTIFS({data_range_f},"OpenAlex BD",{data_range_c},"S\u00ed")',
             "1A5276"),
        ("   OA BD · NO encontradas en Scopus",
             f'=COUNTIFS({data_range_f},"OpenAlex BD",{data_range_r},"No")',
             "1A5276"),
    ]

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 58

    legend_header = ws.cell(row=1, column=3, value="Valores válidos en '¿En cobertura?'")
    legend_header.font      = Font(bold=True, color=COLOR_HEADER_FONT)
    legend_header.fill      = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
    legend_header.alignment = Alignment(horizontal="center")
    legend_header.border    = THIN_BORDER

    legend_items = [
        ("Sí",                                                          COLOR_IN_COV),
        ("No (antes de cobertura)",                                     COLOR_OUT_COV),
        ("No (después de cobertura)",                                   COLOR_OUT_COV),
        ("No (laguna de cobertura)",                                    COLOR_OUT_COV),
        ("Sin datos",                                                   COLOR_NO_DATA),
        ("—",                                                          COLOR_NOT_FOUND),
        ("[Amarillo] Sin ISSN – resuelto por título/DOI/EID. Verificar manualmente.", COLOR_FALLBACK),
    ]

    for row_idx, (label, formula, color) in enumerate(rows_def, start=1):
        ca = ws.cell(row=row_idx, column=1, value=label)
        cb = ws.cell(row=row_idx, column=2, value=formula)

        if color == "__header":
            for c in (ca, cb):
                c.font      = Font(bold=True, color=COLOR_HEADER_FONT)
                c.fill      = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
                c.alignment = Alignment(horizontal="center")
                c.border    = THIN_BORDER
        elif label:
            ca.font   = Font(bold=True)
            ca.border = THIN_BORDER
            cb.border = THIN_BORDER
            cb.alignment = Alignment(horizontal="center")
            if "%" in label:
                cb.number_format = "0.0%"
            if color:
                fill = PatternFill(fill_type="solid", fgColor=color)
                ca.fill = fill
                cb.fill = fill

        if row_idx == 1:
            pass
        elif 1 <= row_idx - 1 <= len(legend_items):
            leg_idx = row_idx - 2
            if 0 <= leg_idx < len(legend_items):
                leg_val, leg_color = legend_items[leg_idx]
                cl = ws.cell(row=row_idx, column=3, value=leg_val)
                cl.border    = THIN_BORDER
                cl.alignment = Alignment(horizontal="left", wrap_text=True)
                if leg_color:
                    cl.fill = PatternFill(fill_type="solid", fgColor=leg_color)

    note_row  = len(rows_def) + 2
    note      = ws.cell(
        row=note_row, column=1,
        value=(
            "⚠ Puede cambiar manualmente los valores en la columna '¿En cobertura?' "
            "de la hoja 'Cobertura' y este resumen se actualizará automáticamente."
        ),
    )
    note.font      = Font(italic=True, color="666666", size=9)
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=3)
    ws.row_dimensions[note_row].height = 30


# ── Utilidad de offset de columna ─────────────────────────────────────────────

def get_column_letter_offset(col_letter: str, offset: int) -> str:
    """Devuelve la letra de columna desplazada `offset` posiciones desde col_letter."""
    col_idx = 0
    for ch in col_letter.upper():
        col_idx = col_idx * 26 + (ord(ch) - ord("A") + 1)
    return get_column_letter(col_idx + offset)
