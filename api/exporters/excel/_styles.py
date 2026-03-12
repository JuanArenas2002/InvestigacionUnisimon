"""
Constantes de estilo, helpers de fuentes/rellenos y funciones
compartidas entre todos los módulos del exportador Excel.
"""
from datetime import datetime

from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Paleta de colores ─────────────────────────────────────────────────────────

COLOR_HEADER_BG   = "1F4E79"   # azul oscuro
COLOR_HEADER_FONT = "FFFFFF"   # blanco
COLOR_ACTIVE      = "C6EFCE"   # verde claro
COLOR_DISCONT     = "FFCCCC"   # rojo claro
COLOR_UNKNOWN     = "FFF2CC"   # amarillo claro
COLOR_ERROR       = "E0E0E0"   # gris claro
COLOR_ALT_ROW     = "EBF3FB"   # azul muy claro (filas alternadas)
COLOR_IN_COV      = "C6EFCE"   # verde
COLOR_OUT_COV     = "FFCCCC"   # rojo
COLOR_NO_DATA     = "FFF2CC"   # amarillo
COLOR_NOT_FOUND   = "E0E0E0"   # gris
COLOR_FALLBACK    = "FFFF99"   # amarillo vivo — resuelto por fallback sin ISSN directo

# ── Borde fino estándar ───────────────────────────────────────────────────────

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ── Caches de objetos de estilo (evita crear miles de instancias) ────────────

_fill_pool: dict[str, PatternFill] = {}
_align_pool: dict[tuple, Alignment] = {}
_font_pool: dict[tuple, Font] = {}


def _fill(color: str) -> PatternFill:
    """Retorna (y cachea) un PatternFill sólido del color dado."""
    if color not in _fill_pool:
        _fill_pool[color] = PatternFill(fill_type="solid", fgColor=color)
    return _fill_pool[color]


def _align(h: str, v: str = "center", wrap: bool = False) -> Alignment:
    """Retorna (y cachea) un Alignment con horizontal, vertical y wrap_text."""
    k = (h, v, wrap)
    if k not in _align_pool:
        _align_pool[k] = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    return _align_pool[k]


def _font(
    bold: bool = False,
    color: str = "000000",
    size: int = 10,
    italic: bool = False,
) -> Font:
    """Retorna (y cachea) un Font con los atributos indicados."""
    k = (bold, color, size, italic)
    if k not in _font_pool:
        _font_pool[k] = Font(bold=bold, color=color, size=size, italic=italic)
    return _font_pool[k]


# ── Helpers de color por valor de cobertura ───────────────────────────────────

def _normalize_header(h: str) -> str:
    """Devuelve el encabezado en minúsculas y sin espacios extremos."""
    return str(h).strip().lower()


def _in_cov_cell_color(in_cov: str) -> tuple[str, str]:
    """Retorna (bg_color, font_color) según el valor de '¿En cobertura?'."""
    v = str(in_cov).strip().lower()
    if v == "sí":
        return ("1E6B2F", "FFFFFF")
    if v.startswith("no"):
        return ("922B21", "FFFFFF")
    if v == "sin datos":
        return ("7D6608", "FFFFFF")
    return ("595959", "FFFFFF")


def _coverage_row_color(
    in_cov: str,
    found: bool,
    row_idx: int,
    found_via: str = "issn",
) -> str:
    """
    Color de fondo para la fila según el resultado de cobertura.

    - Verde    : encontrada vía ISSN y en cobertura
    - Rojo     : encontrada vía ISSN y fuera de cobertura
    - Amarillo vivo: encontrada por fallback (título/DOI/EID)
    - Amarillo tenue: sin datos de cobertura
    - Gris     : no encontrada en Scopus
    """
    if not found:
        return COLOR_NOT_FOUND
    if found_via and found_via not in ("issn", ""):
        return COLOR_FALLBACK if row_idx % 2 == 0 else "FFFFCC"
    if in_cov == "Sí":
        return COLOR_IN_COV if row_idx % 2 == 0 else "D9F0DD"
    if in_cov.startswith("No"):
        return COLOR_OUT_COV if row_idx % 2 == 0 else "FFE0E0"
    return COLOR_NO_DATA if row_idx % 2 == 0 else "FFFBE6"


# ── Palabras clave para detectar columnas de autores ─────────────────────────

_AUTHOR_COL_KEYWORDS = ("author", "affiliat", "correspondence")


def _is_author_col(col_name: str) -> bool:
    """True si el encabezado corresponde a columnas de autores/afiliaciones."""
    norm = _normalize_header(col_name)
    return any(kw in norm for kw in _AUTHOR_COL_KEYWORDS)


# ── Cabecera de hoja compartida ───────────────────────────────────────────────

def _write_sheet_header(ws, all_headers: list, title_text: str) -> None:
    """Escribe la fila de título (fila 1) y la fila de encabezados (fila 2)."""
    num_cols = len(all_headers)
    header_fill = _fill(COLOR_HEADER_BG)

    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font      = Font(bold=True, size=12, color="FFFFFF")
    title_cell.fill      = header_fill
    ws.row_dimensions[1].height = 22
    if num_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)

    for col_idx, col_name in enumerate(all_headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font      = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
    ws.row_dimensions[2].height = 28
