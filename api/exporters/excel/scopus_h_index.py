"""
Generador de Excel con h-index de autores desde Scopus.

Hojas:
- H-Index Autores: métricas + perfil por autor
- Errores: autores que no pudieron procesarse
- Resumen: estadísticas generales
"""

import io
import logging
from datetime import datetime
from typing import Dict, List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("scopus_h_index_excel_exporter")

# ── Estilos ──────────────────────────────────────────────────────────────────

HEADER_FILL  = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ERROR_FILL   = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def _header(ws, row: int, col: int, value, width: float, fill=HEADER_FILL):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill   = fill
    cell.font   = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN
    ws.column_dimensions[get_column_letter(col)].width = width
    return cell


def _cell(ws, row: int, col: int, value, align=LEFT):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border    = THIN
    cell.alignment = align
    return cell


# ── Columnas de la hoja principal ────────────────────────────────────────────

RESULT_COLS = [
    # (header,          key,             width,  center?)
    ("Scopus ID",       "author_id",     16,     True),
    ("Nombre",          "name",          32,     False),
    ("Institución",     "inst",          35,     False),
    ("ORCID",           "orcid",         22,     True),
    ("H-Index",         "h_index",       10,     True),
    ("Documentos",      "document_count",12,     True),
    ("Citas Totales",   "citation_count",14,     True),
    ("Citado Por",      "cited_by_count",12,     True),
    ("Coautores",       "coauthor_count",12,     True),
    ("Áreas",           "areas",         40,     False),
    ("Desde",           "year_from",     8,      True),
    ("Hasta",           "year_to",       8,      True),
]


def generate_h_index_excel(author_results: List[Dict]) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    successful = [r for r in author_results if r.get("status") == "success"]
    errors     = [r for r in author_results if r.get("status") != "success"]

    # ── Hoja 1: resultados ───────────────────────────────────────────────────
    ws = wb.create_sheet("H-Index Autores", 0)

    for col_i, col_def in enumerate(RESULT_COLS, 1):
        _header(ws, 1, col_i, col_def[0], col_def[2])

    for row_i, author in enumerate(successful, 2):
        fill = ALT_FILL if row_i % 2 == 0 else None
        for col_i, (_, key, _, center) in enumerate(RESULT_COLS, 1):
            val  = author.get(key, "")
            cell = _cell(ws, row_i, col_i, val, CENTER if center else LEFT)
            if fill:
                cell.fill = fill

    # ── Hoja 2: errores ──────────────────────────────────────────────────────
    if errors:
        we = wb.create_sheet("Errores", 1)
        for col_i, (header, width) in enumerate(
            [("Scopus ID", 16), ("Mensaje de Error", 60)], 1
        ):
            _header(we, 1, col_i, header, width, fill=ERROR_FILL)

        for row_i, err in enumerate(errors, 2):
            _cell(we, row_i, 1, err.get("author_id", ""))
            _cell(we, row_i, 2, err.get("error", ""))
            for col_i in range(1, 3):
                we.cell(row_i, col_i).fill = ERROR_FILL

    # ── Hoja 3: resumen ──────────────────────────────────────────────────────
    sheet_idx = 2 if errors else 1
    ws_s = wb.create_sheet("Resumen", sheet_idx)

    h_vals = [r.get("h_index") or 0 for r in successful]
    avg_h  = round(sum(h_vals) / len(h_vals), 2) if h_vals else 0
    total_docs = sum(r.get("document_count", 0) or 0 for r in successful)
    total_cit  = sum(r.get("citation_count",  0) or 0 for r in successful)

    rows = [
        ("Total Autores Procesados",  len(author_results)),
        ("Autores Exitosos",          len(successful)),
        ("Autores con Error",         len(errors)),
        ("", ""),
        ("H-Index Promedio",          avg_h),
        ("H-Index Máximo",            max(h_vals, default=0)),
        ("H-Index Mínimo",            min(h_vals, default=0)),
        ("", ""),
        ("Total Documentos",          total_docs),
        ("Total Citas",               total_cit),
        ("", ""),
        ("Fecha de Generación",       datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]

    for row_i, (label, value) in enumerate(rows, 1):
        cl = ws_s.cell(row=row_i, column=1, value=label)
        cv = ws_s.cell(row=row_i, column=2, value=value)
        if label:
            cl.font = Font(bold=True, size=11)
            if label not in ("Total Autores Procesados", "Autores Exitosos",
                             "Autores con Error", "Fecha de Generación"):
                cl.fill = HEADER_FILL
                cl.font = Font(bold=True, color="FFFFFF", size=11)
                cv.fill = SUMMARY_FILL
                cv.font = Font(bold=True, size=11)

    ws_s.column_dimensions["A"].width = 30
    ws_s.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
