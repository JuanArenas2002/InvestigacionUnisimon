"""
Generador de Excel con la producción de autores de Scopus.

Estructura:
- Hoja 1: Resumen (resumen por autor)
- Hoja 2: Todas las Publicaciones (tabla única con todas las publicaciones de todos los autores)
"""

import io
import logging
from typing import List, Dict, Optional
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("scopus_excel_exporter")

# ════════════════════════════════════════════════════════════════════════════════
# ESTILOS DE EXCEL
# ════════════════════════════════════════════════════════════════════════════════

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

ALTER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
ERROR_FILL = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

BORDER_THIN = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

PUBLICATION_COLUMNS = [
    ("Autor", "author_name", 20),
    ("Scopus ID", "scopus_id", 15),
    ("Año", "publication_year", 8),
    ("Título", "title", 50),
    ("URL Scopus", "url", 40),
    ("DOI", "doi", 25),
    ("Revista/Fuente", "source_title", 35),
    ("Tipo Fuente", "source_type", 12),
    ("Tipo Publicación", "publication_type", 15),
    ("ISSN", "issn", 15),
    ("E-ISSN", "eissn", 15),
    ("Volumen", "volume", 8),
    ("Fascículo", "issue", 8),
    ("Páginas", "pages", 10),
    ("Citas", "citation_count", 8),
    ("Open Access", "is_open_access", 12),
    ("Estado OA", "oa_status", 15),
    ("Palabras Clave", "keywords", 40),
    ("Resumen", "description", 60),
    ("Autores", "authors", 50),
]


def generate_author_production_excel(author_results: List[Dict]) -> bytes:
    """
    Genera un Excel con la producción de varios autores.
    
    Estructura:
    - Hoja 1: Resumen (tabla con info por autor)
    - Hoja 2: Todas las Publicaciones (tabla única con todas las pubs)
    
    Args:
        author_results: Lista de dicts con resultado por autor
                       (output de ScopusAuthorProductionService.process_author_ids)
    
    Returns:
        Bytes del archivo Excel
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Eliminar hoja vacía por defecto
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # HOJA 1: RESUMEN
    # ═══════════════════════════════════════════════════════════════════════════════
    
    ws_summary = wb.create_sheet("Resumen", 0)
    _create_summary_sheet(ws_summary, author_results)
    
    # ═══════════════════════════════════════════════════════════════════════════════
    # HOJA 2: TODAS LAS PUBLICACIONES
    # ═══════════════════════════════════════════════════════════════════════════════
    
    ws_publications = wb.create_sheet("Todas las Publicaciones", 1)
    _create_publications_sheet(ws_publications, author_results)
    
    # Guardar a bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    logger.info(f"Excel generado con {len(author_results)} autores")
    return output.getvalue()


def _create_summary_sheet(ws, author_results: List[Dict]) -> None:
    """Crea la hoja de resumen (solo autores con publicaciones)."""
    
    # Filtrar solo autores con status success y publicaciones > 0
    authors_with_pubs = [
        author for author in author_results
        if author.get("status") == "success" and author.get("publications_count", 0) > 0
    ]
    
    # Header
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Resumen de Producción por Autor — Scopus"
    title_cell.font = Font(bold=True, size=14)
    
    timestamp_cell = ws.cell(row=1, column=2)
    timestamp_cell.value = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    timestamp_cell.font = Font(italic=True, size=10)
    
    ws.append([])  # Fila en blanco (row 2)

    # Tabla de resumen (inicia en row 3)
    headers = ["Autor ID", "Nombre del Perfil", "Institución Actual", "Áreas", "Publicaciones", "Estado", "Error"]
    ws.append(headers)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN

    # Datos (solo autores con publicaciones)
    for row_idx, author_data in enumerate(authors_with_pubs, start=4):
        author_id = author_data.get("author_id", "")
        author_name = author_data.get("author_name", "")
        institution_current = author_data.get("institution_current", "")
        subject_areas = author_data.get("subject_areas", "")
        pub_count = author_data.get("publications_count", 0)
        status = author_data.get("status", "")
        error = author_data.get("error", "")

        row_data = [author_id, author_name, institution_current, subject_areas, pub_count, status, error or ""]
        ws.append(row_data)

        # Styling
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER_THIN

            if status == "error":
                cell.fill = ERROR_FILL
            elif row_idx % 2 == 0:
                cell.fill = ALTER_FILL

    # Ajustar ancho de columnas
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 40


def _create_publications_sheet(ws, author_results: List[Dict]) -> None:
    """Crea una hoja única con TODAS las publicaciones únicas de todos los autores."""
    
    # Recolectar todas las publicaciones y deduplicar a nivel global
    publications_dict = {}  # key: scopus_id/doi/title, value: pub
    
    for author_data in author_results:
        if author_data.get("status") == "success":
            author_name = author_data.get("author_name", "")
            author_id = author_data.get("author_id", "")
            publications = author_data.get("publications", [])
            
            for pub in publications:
                # Generar clave única para deduplicación
                scopus_id = pub.get("source_id", "")
                doi = pub.get("doi", "")
                title = pub.get("title", "")

                # Prioridad: Scopus ID > DOI > Título (primeros 50 chars)
                if scopus_id:
                    dedup_key = f"scopus_{scopus_id}"
                elif doi:
                    dedup_key = f"doi_{doi}"
                else:
                    dedup_key = f"title_{title[:50]}"

                # Si no supera en importancia, agregar a este autor
                # USO el nombre VERDADERO del perfil, no del Excel
                if dedup_key not in publications_dict:
                    pub["author_name"] = author_data.get("author_name", author_id)
                    publications_dict[dedup_key] = pub
    
    all_publications = list(publications_dict.values())
    
    # Filtrar publicaciones inválidas (sin título ni scopus_id)
    valid_publications = [
        pub for pub in all_publications
        if pub.get("title", "").strip() or pub.get("scopus_id", "").strip()
    ]
    
    logger.info(f"Publicaciones originales: {len(all_publications)}, válidas: {len(valid_publications)}")
    
    # Header
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"Todas las Publicaciones ({len(valid_publications)} total - únicas)"
    title_cell.font = Font(bold=True, size=12)
    
    ws.append([])  # Fila en blanco (row 2)
    
    # Headers de tabla (row 3)
    headers = [col[0] for col in PUBLICATION_COLUMNS]
    ws.append(headers)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN
    
    # Datos de publicaciones (a partir de row 4)
    for pub_idx, pub in enumerate(valid_publications, start=4):
        for col_idx, (_, field, _) in enumerate(PUBLICATION_COLUMNS, start=1):
            cell = ws.cell(row=pub_idx, column=col_idx)
            cell.value = pub.get(field, "")
            cell.border = BORDER_THIN
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            if pub_idx % 2 == 0:
                cell.fill = ALTER_FILL
    
    # Ajustar anchos de columnas
    for col_idx, (_, _, width) in enumerate(PUBLICATION_COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width



