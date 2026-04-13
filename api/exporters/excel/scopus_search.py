"""
Exportador de Excel para resultados de búsqueda masiva en Scopus.
Genera un archivo con dos sheets: Encontrados y No encontrados.
"""
import io
import logging
from typing import List, Dict

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("excel")

# Colores
COLOR_HEADER_BG = "4472C4"  # Azul
COLOR_HEADER_FONT = "FFFFFF"  # Blanco
COLOR_FOUND_BG = "C6EFCE"  # Verde claro
COLOR_NOT_FOUND_BG = "FFC7CE"  # Rojo claro
COLOR_WHITE = "FFFFFF"

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def _get_header_style():
    """Estilo para encabezados."""
    return {
        'fill': PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid"),
        'font': Font(bold=True, color=COLOR_HEADER_FONT),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': THIN_BORDER,
    }


def _format_cell(cell, content, bg_color=None, bold=False, wrap=True, center=False):
    """Aplica formato a una celda."""
    cell.value = content
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.font = Font(bold=bold)
    if center:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=wrap)
    else:
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=wrap)
    cell.border = THIN_BORDER


def generate_scopus_search_excel(
    found_results: List[Dict],
    not_found_results: List[Dict],
) -> bytes:
    """
    Genera un archivo Excel con dos sheets: Encontrados y No encontrados.
    
    Args:
        found_results: Lista de dicts con publicaciones encontradas
        not_found_results: Lista de dicts con publicaciones no encontradas
    
    Returns:
        Bytes del archivo Excel
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Eliminar sheet por defecto
    
    # Definir columnas
    found_cols = [
        ("# Fila", "_row_num", 7),
        ("Título Original", "title", 50),
        ("Año", "year", 8),
        ("DOI", "doi", 38),
        ("ISSN", "issn", 14),
        ("Revista", "magazine", 35),
        ("ID Scopus", "scopus_id", 20),
        ("Título en Scopus", "scopus_title", 40),
        ("Revista (Scopus)", "scopus_journal", 35),
        ("DOI (Scopus)", "scopus_doi", 35),
        ("ISSN (Scopus)", "scopus_issn", 14),
        ("Método Búsqueda", "search_method", 15),
        ("Query", "search_query", 40),
    ]
    
    not_found_cols = [
        ("# Fila", "_row_num", 7),
        ("Título", "title", 50),
        ("Año", "year", 8),
        ("DOI", "doi", 38),
        ("ISSN", "issn", 14),
        ("Revista", "magazine", 35),
        ("Método Búsqueda", "search_method", 15),
        ("Query", "search_query", 40),
    ]
    
    # Sheet 1: Encontrados
    ws_found = wb.create_sheet("Encontrados", 0)
    _write_sheet(ws_found, found_cols, found_results, COLOR_FOUND_BG)
    
    # Sheet 2: No Encontrados
    ws_not_found = wb.create_sheet("No Encontrados", 1)
    _write_sheet(ws_not_found, not_found_cols, not_found_results, COLOR_NOT_FOUND_BG)
    
    # Sheet 3: Resumen
    ws_summary = wb.create_sheet("Resumen", 0)
    _write_summary_sheet(ws_summary, len(found_results), len(not_found_results))
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _write_sheet(ws, columns, rows, status_color):
    """Escribe un sheet con encabezados y datos."""
    # Encabezados
    header_style = _get_header_style()
    for col_idx, (col_name, _, col_width) in enumerate(columns, 1):
        cell = ws.cell(1, col_idx)
        _format_cell(cell, col_name, bg_color=COLOR_HEADER_BG, bold=True, center=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    # Datos
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, (_, field_key, _) in enumerate(columns, 1):
            cell = ws.cell(row_idx, col_idx)
            value = row_data.get(field_key, "")
            
            # Formatear según tipo
            if field_key == "_row_num":
                _format_cell(cell, value, bg_color=COLOR_WHITE, center=True)
            elif field_key in ("year", "search_method"):
                _format_cell(cell, value, bg_color=COLOR_WHITE, center=True)
            elif isinstance(value, list):
                _format_cell(cell, ", ".join(str(v) for v in value) if value else "", bg_color=COLOR_WHITE)
            else:
                _format_cell(cell, value or "", bg_color=COLOR_WHITE)
    
    # Ajustar altura de encabezado
    ws.row_dimensions[1].height = 30


def _write_summary_sheet(ws, found_count, not_found_count):
    """Escribe el sheet de resumen."""
    header_style = _get_header_style()
    
    # Título
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = "Búsqueda de Productos en Scopus - Resumen"
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Datos
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    
    data = [
        ("Encontrados en Scopus", found_count),
        ("No encontrados", not_found_count),
        ("Total procesados", found_count + not_found_count),
        ("Tasa de éxito (%)", round((found_count / (found_count + not_found_count) * 100), 2) if (found_count + not_found_count) > 0 else 0),
    ]
    
    for row_idx, (label, value) in enumerate(data, 3):
        # Label
        cell_label = ws.cell(row_idx, 1)
        cell_label.value = label
        cell_label.font = Font(bold=True)
        cell_label.alignment = Alignment(horizontal='left', vertical='center')
        cell_label.border = THIN_BORDER
        
        # Value
        cell_value = ws.cell(row_idx, 2)
        cell_value.value = value
        cell_value.alignment = Alignment(horizontal='center', vertical='center')
        cell_value.border = THIN_BORDER
        
        # Color según tipo
        if "Encontrados" in label:
            cell_value.fill = PatternFill(start_color=COLOR_FOUND_BG, end_color=COLOR_FOUND_BG, fill_type="solid")
        elif "No encontrados" in label:
            cell_value.fill = PatternFill(start_color=COLOR_NOT_FOUND_BG, end_color=COLOR_NOT_FOUND_BG, fill_type="solid")
        else:
            cell_value.fill = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")
        
        ws.row_dimensions[row_idx].height = 20
