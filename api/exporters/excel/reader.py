"""
Funciones para leer archivos Excel subidos por el usuario:
  - ISSNs desde la columna A  (para cobertura masiva de revistas)
  - Publicaciones desde un export de Scopus (para verificación de cobertura)
"""
import io
import logging
import re

import openpyxl

from ._styles import _normalize_header

logger = logging.getLogger("excel")

# ── Nombres de columnas que genera el propio exportador ──────────────────────
# Se usan para detectar si el Excel de entrada es un resultado previo del
# endpoint /scopus/check-publications-coverage y omitir esas columnas al releer.

_COVERAGE_COL_NAMES: frozenset = frozenset([
    "Revista en Scopus",
    "Título oficial (Scopus)",
    "Editorial (Scopus)",
    "Estado revista",
    "Periodos de cobertura",
    "¿En cobertura?",
    "ISSN resuelto (Scopus)",
    "E-ISSN resuelto (Scopus)",
    "Fuente",
    "En Scopus",
    "Revista (Scopus)",
    "Editorial",
    "Periodos cobertura",
    "Áreas temáticas",
    "Encontrado vía",
    "ISSN resuelto",
    "E-ISSN resuelto",
    "#",
])

# ── Mapeo de nombres de columna Scopus → clave interna ───────────────────────
# Cada lista incluye el nombre del export Scopus original y variantes usadas
# en el propio reporte generado (para soportar re-procesamiento).

_SCOPUS_COL_MAP = {
    "title":          ["title",        "título del artículo"],
    "year":           ["year",         "año"],
    "source_title":   ["source title", "revista (scopus)"],
    "issn":           ["issn",         "issn resuelto"],
    "eissn":          ["eissn", "e-issn", "electronic issn", "e-issn resuelto"],
    "isbn":           ["isbn"],
    "doi":            ["doi"],
    "document_type":  ["document type", "tipo de publicación"],
    "authors":        ["authors"],
    "eid":            ["eid"],
    "link":           ["link"],
    "language":       ["language of original document", "language"],
    "open_access":    ["open access"],
    "cited_by":       ["cited by"],
    "publisher":      ["publisher"],
}


def _deduplicate_headers(headers: list) -> list:
    """Añade sufijo _2, _3… a nombres de columna duplicados."""
    seen: dict = {}
    result = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            result.append(f"{h}_{seen[h]}")
        else:
            seen[h] = 1
            result.append(h)
    return result


def _looks_like_header_row(row_values: tuple) -> bool:
    """
    Devuelve True si la fila parece contener encabezados de columna reales.
    Comprueba que al menos una celda coincida con palabras clave conocidas.
    """
    _KNOWN_HEADERS = {
        "title", "year", "issn", "isbn", "doi", "eid",
        "source title", "document type", "authors", "publisher",
        "cited by", "open access", "language",
        "revista en scopus", "estado revista", "¿en cobertura?",
        "título oficial (scopus)", "editorial (scopus)",
        "fuente", "en scopus", "título del artículo", "año",
        "tipo de publicación", "revista (scopus)", "periodos cobertura",
        "encontrado vía", "áreas temáticas", "issn resuelto", "e-issn resuelto",
    }
    non_empty = [str(c).strip().lower() for c in row_values if c is not None and str(c).strip()]
    return any(v in _KNOWN_HEADERS for v in non_empty)


# ── read_issns_from_excel ─────────────────────────────────────────────────────

def read_issns_from_excel(file_bytes: bytes) -> list:
    """
    Lee un archivo .xlsx y extrae ISSNs/E-ISSNs de la columna A.

    Modo automático:
    - Lee ISSN/E-ISSN de columna A (pueden estar separados por ; o ,)
    - Detecta si el contenido son ISSNs o nombres de revistas
    - Omite la primera fila si parece un encabezado
    - Acepta ISSNs con o sin guion: '2595-3982' ó '25953982'
    - E-ISSNs: '2146-4553'
    - Nombres de revistas: "Nature", "Journal of Biochemistry", etc.
    - Elimina duplicados preservando el orden
    - Ignora celdas vacías

    Returns:
        Lista de [ISSNs | E-ISSNs | Nombres] como strings. El extractor maneja todos los tipos.

    Raises:
        ValueError: Si el archivo no es válido o la columna A está vacía.
    """
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
    except Exception as e:
        raise ValueError(f"No se pudo leer el archivo Excel: {e}")

    ws = wb.active
    ISSN_RE = re.compile(r"^\d{4}-?\d{3}[\dXx]$", re.IGNORECASE)

    seen: set = set()
    items: list = []

    for row_idx, row in enumerate(
        ws.iter_rows(min_col=1, max_col=1, values_only=True), start=1
    ):
        raw = row[0]
        if raw is None:
            continue
        value = str(raw).strip()
        if not value:
            continue

        # Skip header row (solo si es claramente un encabezado)
        if row_idx == 1 and value.lower() in ("issn", "e-issn", "eissn", "source title", "journal", "year", "id"):
            continue

        # Dividir múltiples valores separados por ; o , (ej: "ISSN; E-ISSN" o "ISSN, E-ISSN")
        parts = re.split(r'[;,]', value)
        
        for part in parts:
            part = part.strip()
            if not part:
                continue
            
            normalized = part.replace("-", "")
            
            # Es un ISSN/E-ISSN válido (XXXX-XXXD o XXXXXXXD)
            if ISSN_RE.match(part) or re.match(r"^[\dXx]{7,8}$", normalized, re.I):
                if part not in seen:
                    seen.add(part)
                    items.append(part)
            # Es un nombre de revista (texto con al menos 3 caracteres)
            elif len(part) >= 3 and not part.isdigit():
                if part not in seen:
                    seen.add(part)
                    items.append(part)

    wb.close()

    if not items:
        raise ValueError(
            "No se encontraron datos válidos en la columna A del archivo. "
            "Aceptamos: ISSNs (2595-3982 o 25953982), E-ISSNs (2146-4553), nombres de revistas, "
            "o múltiples separados por punto y coma (ISSN; E-ISSN)."
        )

    return items


# ── read_publications_from_excel ──────────────────────────────────────────────

def read_publications_from_excel(file_bytes: bytes) -> tuple:
    """
    Lee un Excel de exportación de Scopus (o el Excel resultado del propio
    endpoint) y retorna los datos de publicaciones.

    Detecta automáticamente si la fila 1 es un título del reporte (celda
    fusionada) y en ese caso usa la fila 2 como encabezados reales.

    Returns:
        Tuple (headers: list[str], rows: list[dict]):
          - headers: nombres originales de columna en orden
            (sin las columnas de cobertura si el archivo es un resultado previo)
          - rows: cada fila como dict con TODOS los valores originales,
                  MÁS claves internas normalizadas (__title, __year, …)

    Raises:
        ValueError: Si el archivo no es válido o no tiene encabezados.
    """
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
    except Exception as e:
        raise ValueError(f"No se pudo leer el archivo Excel: {e}")

    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        raise ValueError("El archivo Excel está vacío.")

    header_row_idx = 0
    if not _looks_like_header_row(all_rows[0]):
        if len(all_rows) > 1 and _looks_like_header_row(all_rows[1]):
            header_row_idx = 1
            logger.info(
                "[read_publications] Fila 1 detectada como título del reporte; "
                "usando fila 2 como encabezados."
            )
        else:
            logger.warning(
                "[read_publications] No se reconoció una fila de encabezados estándar; "
                "se usará la primera fila como encabezados."
            )

    raw_headers: list = [
        str(c).strip() if c is not None else f"Col_{i}"
        for i, c in enumerate(all_rows[header_row_idx])
    ]
    raw_headers = _deduplicate_headers(raw_headers)

    logger.info(
        f"[read_publications] header_row_idx={header_row_idx}, "
        f"total filas de datos={len(all_rows) - header_row_idx - 1}, "
        f"headers detectados ({len(raw_headers)}): "
        f"{raw_headers[:8]}{'...' if len(raw_headers) > 8 else ''}"
    )

    rows: list = []
    norm_headers = {_normalize_header(h): h for h in raw_headers}

    for row in all_rows[header_row_idx + 1:]:
        if all(c is None for c in row):
            continue

        row_dict: dict = {}
        for col_name, val in zip(raw_headers, row):
            row_dict[col_name] = val if val is not None else ""

        for internal_key, candidates in _SCOPUS_COL_MAP.items():
            for cand in candidates:
                orig_col = norm_headers.get(cand)
                if orig_col:
                    row_dict[f"__{internal_key}"] = row_dict.get(orig_col, "")
                    break
            else:
                row_dict[f"__{internal_key}"] = ""

        rows.append(row_dict)

    if not raw_headers:
        raise ValueError("El archivo Excel no tiene encabezados reconocibles.")
    if not rows:
        raise ValueError("El archivo Excel no contiene filas de datos.")

    # Descartar columnas de cobertura previas (si el Excel es un resultado anterior)
    headers = [h for h in raw_headers if h not in _COVERAGE_COL_NAMES]
    if len(headers) < len(raw_headers):
        dropped = [h for h in raw_headers if h in _COVERAGE_COL_NAMES]
        logger.info(
            f"[read_publications] Excel resultado re-procesado: "
            f"se ignoraron {len(dropped)} columnas previas de cobertura: {dropped}"
        )

    return headers, rows
