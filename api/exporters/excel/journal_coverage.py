"""
Genera el Excel de cobertura de revistas en Scopus (Serial Title API).

Función pública:
  generate_journal_coverage_excel(results) → bytes
"""
import io
import logging
from datetime import datetime
from typing import List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ._styles import (
    COLOR_HEADER_BG,
    COLOR_HEADER_FONT,
    COLOR_ACTIVE,
    COLOR_DISCONT,
    COLOR_ALT_ROW,
    COLOR_ERROR,
    THIN_BORDER,
)

logger = logging.getLogger("excel")

# ── Definición de columnas ────────────────────────────────────────────────────

COLUMNS = [
    ("ISSN",             "issn",            15),
    ("Título",           "title",           45),
    ("Editorial",        "publisher",       30),
    ("Estado",           "status",          14),
    ("¿Descontinuada?",  "is_discontinued", 16),
    ("Desde (año)",      "coverage_from",   13),
    ("Hasta (año)",      "coverage_to",     13),
    ("Áreas temáticas",  "subject_areas",   40),
    ("Source ID Scopus", "source_id",       18),
    ("Error",            "error",           35),
]


# ── Función principal ─────────────────────────────────────────────────────────

def generate_journal_coverage_excel(results: List[dict]) -> bytes:
    """
    Genera un Excel en memoria a partir de los resultados de cobertura.

    Args:
        results: Lista de dicts retornados por SerialTitleExtractor.get_bulk_coverage()

    Returns:
        bytes del archivo .xlsx listo para enviar como respuesta HTTP.
    """
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Cobertura de Revistas"
    _write_title_row(ws, len(COLUMNS))
    _write_header_row(ws)
    _write_data_rows(ws, results)
    _auto_adjust_columns(ws)

    ws_summary = wb.create_sheet("Resumen")
    _write_summary_sheet(ws_summary, results)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


# ── Helpers privados ──────────────────────────────────────────────────────────

def _write_title_row(ws, num_cols: int) -> None:
    """Fila 1: título del reporte."""
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = (
        f"Reporte de Cobertura de Revistas en Scopus  —  "
        f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    title_cell.font      = Font(bold=True, size=13, color=COLOR_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)


def _write_header_row(ws) -> None:
    """Fila 2: encabezados de columna con estilo."""
    header_fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)
    for col_idx, (header, _, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value     = header
        cell.font      = Font(bold=True, color=COLOR_HEADER_FONT, size=11)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER
    ws.row_dimensions[2].height = 30


def _write_data_rows(ws, results: List[dict]) -> None:
    """Filas de datos a partir de la fila 3."""
    for row_idx, item in enumerate(results, start=3):
        fill_color = _row_fill_color(item, row_idx)
        row_fill   = PatternFill(fill_type="solid", fgColor=fill_color)

        for col_idx, (_, key, _) in enumerate(COLUMNS, start=1):
            value = item.get(key)
            if key == "subject_areas" and isinstance(value, list):
                value = " | ".join(value) if value else None
            elif key == "is_discontinued":
                value = "Sí" if value else "No"

            cell           = ws.cell(row=row_idx, column=col_idx)
            cell.value     = value
            cell.fill      = row_fill
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        ws.row_dimensions[row_idx].height = 20


def _row_fill_color(item: dict, row_idx: int) -> str:
    """Retorna el color de fondo según el estado de la revista."""
    if item.get("error"):
        return COLOR_ERROR
    status = (item.get("status") or "").lower()
    if "discontinued" in status or "inactive" in status:
        return COLOR_DISCONT
    if status == "inactiva":
        return "FCE5CD"
    if "active" in status:
        return COLOR_ACTIVE if row_idx % 2 == 0 else "D9F0DD"
    return COLOR_ALT_ROW if row_idx % 2 == 0 else "FFFFFF"


def _auto_adjust_columns(ws) -> None:
    """Aplica los anchos de columna definidos en COLUMNS."""
    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_summary_sheet(ws, results: List[dict]) -> None:
    """Hoja de resumen con estadísticas del lote."""
    total     = len(results)
    found     = sum(1 for r in results if not r.get("error") and r.get("title"))
    not_found = sum(1 for r in results if r.get("error") == "Revista no encontrada en Scopus.")
    errors    = sum(1 for r in results if r.get("error") and r.get("error") != "Revista no encontrada en Scopus.")
    active    = sum(1 for r in results if (r.get("status") or "").lower() == "active")
    discont   = sum(1 for r in results if (r.get("status") or "").lower() in ("discontinued", "inactive", "inactiva"))
    unknown   = sum(1 for r in results if (r.get("status") or "").lower() == "unknown" and not r.get("error"))

    header_fill = PatternFill(fill_type="solid", fgColor=COLOR_HEADER_BG)

    rows_def = [
        ("Métrica",                          "Valor"),
        ("Total ISSNs consultados",          total),
        ("Revistas encontradas",             found),
        ("No encontradas (404)",             not_found),
        ("Errores de API",                   errors),
        ("",                                 ""),
        ("Activas",                          active),
        ("Descontinuadas",                   discont),
        ("Estado desconocido",               unknown),
        ("",                                 ""),
        ("Fecha de generación",              datetime.now().strftime("%d/%m/%Y %H:%M")),
    ]

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    for row_idx, (label, value) in enumerate(rows_def, start=1):
        cell_a = ws.cell(row=row_idx, column=1, value=label)
        cell_b = ws.cell(row=row_idx, column=2, value=value)
        if row_idx == 1:
            for cell in (cell_a, cell_b):
                cell.font      = Font(bold=True, color=COLOR_HEADER_FONT)
                cell.fill      = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border    = THIN_BORDER
        elif label:
            cell_a.font = Font(bold=True)
            cell_a.border = THIN_BORDER
            cell_b.border = THIN_BORDER
            cell_b.alignment = Alignment(horizontal="center")
            if row_idx % 2 == 0:
                even_fill = PatternFill(fill_type="solid", fgColor=COLOR_ALT_ROW)
                cell_a.fill = even_fill
                cell_b.fill = even_fill
