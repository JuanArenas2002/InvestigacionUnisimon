import openpyxl
import os

# Crear archivo Excel con IDs de Scopus conocidos y válidos
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 'author_id'

# IDs reales de autores con perfiles públicos en Scopus
valid_ids = [
    '57193767797',  # Usuario anterior que funcionó
    '7404530122',   # Probado antes
    '35093378600',  # Probado antes
]

for idx, author_id in enumerate(valid_ids, 2):
    ws[f'A{idx}'] = author_id

os.makedirs('c:/temp', exist_ok=True)
wb.save('c:/temp/test_valid_authors.xlsx')
print('✅ Excel con IDs válidos creado: c:/temp/test_valid_authors.xlsx')
print(f'📝 IDs incluidos: {valid_ids}')
