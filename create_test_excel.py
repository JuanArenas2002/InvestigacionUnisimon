"""
Script para crear un archivo Excel de prueba con publicaciones.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill

# Crear workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Publicaciones"

# Encabezados
headers = ["Título", "Año", "DOI", "ISSN", "Revista", "Autores", "Institución"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(1, col_idx)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

# Datos de ejemplo (publicaciones reales de ejemplo)
data = [
    [
        "Deep learning for image recognition",
        2023,
        "10.1038/s41586-023-06585-5",
        "0028-0836",
        "Nature",
        "Smith, J.; Brown, M.",
        "MIT"
    ],
    [
        "Machine learning in healthcare",
        2022,
        "10.1016/j.jmi.2022.100041",
        "2589-750X",
        "Journal of Medical Internet Research",
        "Johnson, A.; Lee, S.",
        "Stanford University"
    ],
    [
        "Climate change impacts on biodiversity",
        2023,
        "10.1126/science.abo7906",
        "0036-8075",
        "Science",
        "Green, P.; Wilson, T.",
        "Oxford University"
    ],
    [
        "Artificial intelligence in drug discovery",
        2024,
        "",  # Sin DOI
        "0926-6690",
        "Journal of Computer-Aided Molecular Design",
        "Anderson, K.; Taylor, L.",
        "Cambridge University"
    ],
    [
        "Quantum computing applications",
        2023,
        "10.1038/s41534-023-00680-1",
        "",  # Sin ISSN
        "npj Quantum Information",
        "Chen, X.; Kumar, R.",
        "Toronto University"
    ],
]

# Escribir datos
for row_idx, row_data in enumerate(data, 2):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row_idx, col_idx).value = value

# Ajustar anchos de columna
ws.column_dimensions['A'].width = 50
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 35
ws.column_dimensions['F'].width = 30
ws.column_dimensions['G'].width = 25

# Guardar
wb.save("test_scopus_search.xlsx")
print("✓ Archivo de prueba creado: test_scopus_search.xlsx")
print(f"  - Contiene {len(data)} publicaciones de ejemplo")
print("  - Columnas: Título, Año, DOI, ISSN, Revista, Autores, Institución")
