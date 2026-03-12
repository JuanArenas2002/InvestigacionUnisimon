"""
Enriquecedor de publicaciones vía PyAlex / OpenAlex API.

Responsabilidades:
  - Leer un Excel con publicaciones (título, año, DOI).
  - Enriquecer cada fila consultando OpenAlex en tres etapas:
      1. Lote DOI  — hasta MAX_BATCH DOIs por petición (filter_or).
      2. Fallback título — para DOIs que el lote no resolvió.
      3. Búsqueda por título — para registros sin DOI.
  - Guardar el resultado enriquecido en un nuevo Excel.
"""

import io
import logging
import socket
import time
from pathlib import Path
from typing import List

import pyalex
from pyalex import Works

from config import openalex_config, institution
from extractors.base import normalize_doi, normalize_author_name
from ._rate_limit import OpenAlexRateLimitError, extract_retry_after

logger = logging.getLogger(__name__)


class  OpenAlexEnricher:
    """
    Enriquece un listado de publicaciones consultando OpenAlex a través de PyAlex.

    Acepta un Excel con (al menos) las columnas:
        Título  |  Año  |  doi

    Estrategia de búsqueda por orden de prioridad:
      1. Lote DOI  — hasta MAX_BATCH DOIs por petición (filtro OR).
                     Rápido y exacto. Un solo HTTP request por lote.
      2. Sin DOI   — búsqueda full-text por título + filtro por año.
                     Se elige el candidato con mayor similitud usando
                     RapidFuzz (token_sort_ratio, umbral MIN_SCORE %).

    Uso típico::

        enricher = OpenAlexEnricher()

        # Desde un archivo Excel
        rows = enricher.enrich_from_excel("publicaciones.xlsx")

        # Desde bytes (útil dentro de FastAPI)
        rows = enricher.enrich_from_excel_bytes(file_bytes)

        # Guardar resultado
        enricher.save_to_excel(rows, "resultado_openalex.xlsx")

    Campos OpenAlex añadidos a cada fila (prefijo ``oa_``):
        oa_encontrado, oa_metodo, oa_work_id, oa_titulo, oa_año, oa_doi,
        oa_tipo, oa_revista, oa_issn, oa_editorial, oa_open_access,
        oa_status_oa, oa_citas, oa_idioma, oa_url, oa_autores
    """

    MAX_BATCH      = 25    # DOIs por petición (25 para evitar URLs demasiado largas)
    MIN_SCORE      = 80.0  # umbral score compuesto (título 85% + año 15%) — búsqueda general
    MIN_SCORE_ISSN = 68.0  # umbral permisivo cuando ISSN ya confirma la revista
    MIN_SCORE_SOURCE = 70.0  # umbral para coincidencia por nombre de revista

    # Campos que se piden a OpenAlex (reduce tamaño de respuesta)
    _SELECT = [
        "id", "doi", "title", "publication_year", "type",
        "primary_location", "open_access", "authorships",
        "cited_by_count", "language",
    ]

    def __init__(self, email: str = None):
        try:
            import pyalex as _pyalex
            from pyalex import Works as _Works
            self._pyalex = _pyalex
            self._Works  = _Works
        except ImportError:
            raise ImportError(
                "PyAlex no está instalado.  Ejecuta:  pip install pyalex"
            )

        self.email   = email or institution.contact_email
        self.api_key = openalex_config.api_key

        pyalex.config.email                = self.email
        pyalex.config.api_key              = self.api_key or None
        pyalex.config.max_retries          = 2
        pyalex.config.retry_backoff_factor = 0.5
        # 429 se maneja manualmente — NO dejar que urllib3 espere el Retry-After
        # (puede ser miles de segundos → la app parecería congelada)
        pyalex.config.retry_http_codes     = [500, 502, 503, 504]

        # Timeout global de socket para que las peticiones nunca cuelguen
        socket.setdefaulttimeout(30)

        key_status = f"key={'***' + self.api_key[-4:] if self.api_key else 'no configurada'}"
        logger.info(f"[OpenAlexEnricher] Iniciado — polite pool: {self.email} | {key_status}")

    # ── API pública ──────────────────────────────────────────────────────────

    def enrich_from_excel(self, file_path) -> list[dict]:
        """
        Lee un Excel (.xlsx) por ruta y devuelve las filas enriquecidas.

        Args:
            file_path: str o Path al archivo .xlsx.

        Returns:
            Lista de dicts con los campos originales + campos ``oa_*``.
        """
        rows = self._read_excel_path(Path(file_path))
        return self.enrich(rows)

    def enrich_from_excel_bytes(self, file_bytes: bytes) -> list[dict]:
        """
        Igual que ``enrich_from_excel`` pero acepta los bytes del archivo
        directamente (útil en endpoints FastAPI con UploadFile).
        """
        rows = self._read_excel_bytes(file_bytes)
        return self.enrich(rows)

    def enrich(self, publications: list[dict]) -> list[dict]:
        """
        Enriquece una lista de dicts que contienen al menos:
            - clave ``titulo`` / ``Título`` / ``title``
            - clave ``año``    / ``Año``    / ``year``
            - clave ``doi``    / ``DOI``

        Returns:
            La misma lista con campos ``oa_*`` añadidos a cada dict.
        """
        n = len(publications)
        logger.info(f"[OpenAlexEnricher] Enriqueciendo {n} publicaciones…")

        oa_map: list[dict | None] = [None] * n
        method: list[str | None]  = [None] * n

        # Cache intra-ejecución: evita rellamar a la API si el mismo
        # título+año aparece más de una vez en el input (duplicados).
        # Clave: (title_norm, year)  Valor: (oa_dict | None, method_str | None)
        self._search_cache: dict[tuple, tuple] = {}

        with_doi    = [(i, p) for i, p in enumerate(publications) if self._doi(p)]
        without_doi = [(i, p) for i, p in enumerate(publications) if not self._doi(p)]

        logger.info(
            f"[OpenAlexEnricher] Con DOI: {len(with_doi)} | "
            f"Sin DOI: {len(without_doi)}"
        )

        # 1. Búsqueda por lotes de DOI (precisa y rápida)
        try:
            self._enrich_by_doi_batch(oa_map, method, with_doi)
        except OpenAlexRateLimitError as e:
            logger.error(f"[OpenAlexEnricher] RATE LIMIT ALCANZADO — {e}")
            raise

        # 2. Fallback por título: registros CON doi que el lote no resolvió
        doi_not_found = [(i, p) for i, p in with_doi if oa_map[i] is None]
        if doi_not_found:
            logger.info(
                f"[OpenAlexEnricher] Fallback título para {len(doi_not_found)} "
                f"con DOI no resuelto…"
            )
            try:
                self._enrich_by_title(oa_map, method, doi_not_found)
            except OpenAlexRateLimitError as e:
                logger.error(f"[OpenAlexEnricher] RATE LIMIT en fallback título — {e}")
                raise

        # 3. Búsqueda por título para los que no tienen DOI
        try:
            self._enrich_by_title(oa_map, method, without_doi)
        except OpenAlexRateLimitError as e:
            logger.error(f"[OpenAlexEnricher] RATE LIMIT ALCANZADO durante búsqueda por título — {e}")
            raise

        # 4. Fallback por ISSN + título: para los que siguen sin resolver y tienen ISSN
        still_not_found = [
            (i, p) for i, p in enumerate(publications)
            if oa_map[i] is None and self._issns(p)
        ]
        if still_not_found:
            logger.info(
                f"[OpenAlexEnricher] Fallback ISSN+título para "
                f"{len(still_not_found)} publicaciones sin resolver…"
            )
            try:
                self._enrich_by_issn(oa_map, method, still_not_found)
            except OpenAlexRateLimitError as e:
                logger.error(f"[OpenAlexEnricher] RATE LIMIT en fallback ISSN — {e}")
                raise

        # 5. Fallback por nombre de revista + título: sin ISSN pero con columna 'revista'
        still_not_found2 = [
            (i, p) for i, p in enumerate(publications)
            if oa_map[i] is None and self._revista(p) and not self._issns(p)
        ]
        if still_not_found2:
            logger.info(
                f"[OpenAlexEnricher] Fallback nombre-revista para "
                f"{len(still_not_found2)} publicaciones sin resolver…"
            )
            try:
                self._enrich_by_source_name(oa_map, method, still_not_found2)
            except OpenAlexRateLimitError as e:
                logger.error(f"[OpenAlexEnricher] RATE LIMIT en fallback nombre-revista — {e}")
                raise

        # 6. Último recurso: búsqueda solo por título sin filtro de año
        #    (para los que siguen sin resolverse y tienen título)
        last_resort = [
            (i, p) for i, p in enumerate(publications)
            if oa_map[i] is None and self._title(p)
        ]
        if last_resort:
            logger.info(
                f"[OpenAlexEnricher] Último recurso (solo título) para "
                f"{len(last_resort)} publicaciones…"
            )
            try:
                self._enrich_by_title_only(oa_map, method, last_resort)
            except OpenAlexRateLimitError as e:
                logger.error(f"[OpenAlexEnricher] RATE LIMIT en último recurso — {e}")
                raise

        found = sum(1 for x in oa_map if x is not None)
        logger.info(
            f"[OpenAlexEnricher] Encontrados: {found}/{n} "
            f"({found / n * 100:.1f}%)"
        )

        return [
            self._flatten(pub, oa_map[i], method[i])
            for i, pub in enumerate(publications)
        ]

    def save_to_excel(self, enriched_rows: list[dict], output_path) -> Path:
        """
        Guarda los resultados enriquecidos en un Excel (.xlsx).

        Las columnas originales (Título, Año, doi) van primero; los campos
        OpenAlex (oa_*) van después con encabezado azul diferenciado.

        Args:
            enriched_rows: lista devuelta por ``enrich()`` / ``enrich_from_excel()``.
            output_path:   str o Path donde guardar el .xlsx.

        Returns:
            Path del archivo guardado.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl requerido: pip install openpyxl")

        if not enriched_rows:
            raise ValueError("No hay filas para guardar.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enriquecimiento OpenAlex"

        headers = list(enriched_rows[0].keys())

        # ── Fila 1: título del reporte ───────────────────────────────────────
        from datetime import datetime as _dt
        _hdr_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        _oa_fill  = PatternFill(fill_type="solid", fgColor="16537e")
        _hdr_font = Font(bold=True, color="FFFFFF", size=10)

        title_cell = ws.cell(
            row=1, column=1,
            value=(
                f"Enriquecimiento OpenAlex  —  "
                f"{len(enriched_rows)} publicaciones  —  "
                f"Generado: {_dt.now().strftime('%d/%m/%Y %H:%M')}"
            ),
        )
        title_cell.font  = Font(bold=True, size=12, color="FFFFFF")
        title_cell.fill  = _hdr_fill
        ws.row_dimensions[1].height = 22
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

        # ── Fila 2: encabezados ─────────────────────────────────────────────
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.fill      = _oa_fill if h.startswith("oa_") else _hdr_fill
            cell.font      = _hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 28

        # ── Filas de datos ───────────────────────────────────────────────────
        _found_fill  = PatternFill(fill_type="solid", fgColor="EBF3FB")
        _notfnd_fill = PatternFill(fill_type="solid", fgColor="F5F5F5")
        _yes_fill    = PatternFill(fill_type="solid", fgColor="C6EFCE")
        _no_fill     = PatternFill(fill_type="solid", fgColor="FFCCCC")

        for ri, row in enumerate(enriched_rows, start=3):
            found    = row.get("oa_encontrado", False)
            base_fll = _found_fill if found else _notfnd_fill
            for ci, h in enumerate(headers, start=1):
                val  = row.get(h, "")
                if isinstance(val, bool):
                    val = "Sí" if val else "No"
                cell = ws.cell(row=ri, column=ci, value=val if val is not None else "")
                cell.alignment = Alignment(
                    horizontal="left", vertical="center",
                    wrap_text=(h in ("oa_autores", "oa_titulo", "titulo")),
                )
                if h == "oa_encontrado":
                    cell.fill = _yes_fill if found else _no_fill
                    cell.font = Font(bold=True)
                elif h == "oa_open_access":
                    cell.fill = _yes_fill if val == "Sí" else base_fll
                else:
                    cell.fill = base_fll
            ws.row_dimensions[ri].height = 16

        # ── Anchos de columna ────────────────────────────────────────────────
        _widths = {
            "titulo":        50, "oa_titulo":  50, "oa_autores":   46,
            "doi":           38, "oa_doi":     38, "oa_url":       38,
            "oa_work_id":    34, "oa_revista": 34, "oa_editorial": 28,
            "oa_encontrado": 14, "oa_metodo":  14, "oa_tipo":      16,
            "oa_year":        8, "año":         8, "oa_status_oa": 14,
        }
        for ci, h in enumerate(headers, start=1):
            w = _widths.get(h, 18 if h.startswith("oa_") else 14)
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.freeze_panes = "A3"

        output_path = Path(output_path)
        wb.save(output_path)
        logger.info(f"[OpenAlexEnricher] Excel guardado: {output_path}")
        return output_path

    # ── Internos: búsqueda ───────────────────────────────────────────────────

    def _enrich_by_doi_batch(
        self,
        oa_map:  list,
        method:  list,
        indexed: list[tuple[int, dict]],
    ) -> None:
        """Resuelve DOIs en lotes — una sola petición por lote."""
        import requests as _req

        for batch_start in range(0, len(indexed), self.MAX_BATCH):
            batch = indexed[batch_start : batch_start + self.MAX_BATCH]

            doi_to_idx: dict[str, int] = {}
            doi_full_list: list[str]   = []
            for idx, pub in batch:
                d = self._doi(pub)
                if not d:
                    continue
                full = d if d.startswith("https://") else f"https://doi.org/{d}"
                doi_to_idx[d]   = idx
                doi_full_list.append(full)

            if not doi_full_list:
                continue

            n_lote     = len(doi_full_list)
            lote_label = f"{batch_start + 1}–{batch_start + len(batch)}"
            logger.info(f"[OpenAlexEnricher] Lote DOI {lote_label}: {n_lote} DOIs")

            try:
                query_obj = (
                    self._Works()
                    .select(self._SELECT)
                    .filter_or(doi=doi_full_list)
                )
                url_preview = query_obj.url[:120]
                logger.info(
                    f"[OpenAlexEnricher] → GET {url_preview}"
                    f"{'...' if len(query_obj.url) > 120 else ''}"
                )
                works = query_obj.get(per_page=min(n_lote, 200))
                logger.info(
                    f"[OpenAlexEnricher] ← Lote {lote_label}: "
                    f"{len(works)} works recibidos"
                )
                found_in_batch = 0
                for work in works:
                    w_doi = normalize_doi(str(work.get("doi") or ""))
                    if w_doi in doi_to_idx:
                        i = doi_to_idx[w_doi]
                        oa_map[i] = dict(work)
                        method[i] = "doi"
                        found_in_batch += 1
                logger.info(
                    f"[OpenAlexEnricher] ✓ Lote {lote_label}: "
                    f"{found_in_batch}/{n_lote} coincidencias"
                )
                time.sleep(0.1)

            except OpenAlexRateLimitError:
                raise
            except (socket.timeout, _req.exceptions.Timeout) as exc:
                logger.error(
                    f"[OpenAlexEnricher] TIMEOUT en lote {lote_label} — "
                    f"la API no respondió en 30s ({exc})"
                )
            except _req.exceptions.ConnectionError as exc:
                logger.error(
                    f"[OpenAlexEnricher] ERROR DE CONEXIÓN en lote {lote_label}: {exc}"
                )
            except _req.exceptions.HTTPError as exc:
                retry_after = extract_retry_after(exc)
                if retry_after is not None:
                    raise OpenAlexRateLimitError(retry_after)
                logger.warning(
                    f"[OpenAlexEnricher] HTTP error en lote {lote_label}: {exc}"
                )
            except Exception as exc:
                retry_after = extract_retry_after(exc)
                if retry_after is not None:
                    raise OpenAlexRateLimitError(retry_after)
                logger.warning(
                    f"[OpenAlexEnricher] Lote DOI {lote_label} falló "
                    f"({type(exc).__name__}: {exc}); reintentando individualmente…"
                )
                for full_doi, idx in {
                    (d if d.startswith("https://") else f"https://doi.org/{d}"): i
                    for i, pub in batch
                    for d in [self._doi(pub)] if d
                }.items():
                    try:
                        logger.debug(f"[OpenAlexEnricher]   → DOI individual: {full_doi}")
                        work = self._Works().select(self._SELECT)[full_doi]
                        oa_map[idx] = dict(work)
                        method[idx] = "doi"
                        time.sleep(0.12)
                    except Exception as e2:
                        retry_after2 = extract_retry_after(e2)
                        if retry_after2 is not None:
                            raise OpenAlexRateLimitError(retry_after2)
                        logger.debug(
                            f"[OpenAlexEnricher] DOI individual no encontrado "
                            f"'{full_doi}': {e2}"
                        )

    def _enrich_by_title(
        self,
        oa_map:  list,
        method:  list,
        indexed: list[tuple[int, dict]],
    ) -> None:
        """Busca por título + año para publicaciones sin DOI."""
        total = len(indexed)
        for seq, (idx, pub) in enumerate(indexed, start=1):
            title = self._title(pub)
            year  = self._year(pub)
            if not title:
                logger.debug(f"[OpenAlexEnricher] Título vacío en fila {idx}, saltando")
                continue
            if seq % 25 == 1:
                logger.info(
                    f"[OpenAlexEnricher] Por título: {seq}/{total} "
                    f"(encontrados: {sum(1 for m in method if m and 'titulo' in m)})"
                )
            try:
                title_norm  = self._normalize_title(title)
                title_clean = self._sanitize_title(title)  # minúsculas + sin ¿¡ iniciales

                # ── Cache intra-ejecución ─────────────────────────────────────────
                _cache_key = (title_norm, year)
                if hasattr(self, "_search_cache") and _cache_key in self._search_cache:
                    _cached_oa, _cached_met = self._search_cache[_cache_key]
                    if _cached_oa is not None:
                        oa_map[idx] = _cached_oa
                        method[idx] = _cached_met
                        logger.debug(f"[OpenAlexEnricher]   ✓ Cache hit '{title[:55]}'")
                    continue

                all_seen: list = []   # acumula candidatos de todos los intentos

                # Intento 1: título limpio (minúsculas, sin ¿¡) + año exacto
                # Usa search_filter(title=) que solo busca en el campo title.
                # title_clean conserva tildes para mayor precisión.
                query = (
                    self._Works()
                    .select(self._SELECT)
                    .search_filter(title=title_clean)
                )
                if year:
                    query = query.filter(publication_year=year)
                candidates = query.get(per_page=15)
                all_seen.extend(candidates)
                best = self._best_match(title, candidates, year)

                # Intento 2: título normalizado (sin tildes) + año
                # Cubre casos donde el título en OA usa codificación diferente
                if best is None and title_norm != title_clean:
                    query2 = (
                        self._Works()
                        .select(self._SELECT)
                        .search_filter(title=title_norm)
                    )
                    if year:
                        query2 = query2.filter(publication_year=year)
                    candidates2 = query2.get(per_page=15)
                    all_seen.extend(candidates2)
                    best = self._best_match(title, candidates2, year)
                    if best:
                        time.sleep(0.1)

                # ── Salida temprana: si el pool R1/R2 ya contiene un candidato
                # ── con título ≥88%, el año es el único obstáculo.
                # ── R3+R4 devolverían los mismos candidatos vía distinto índice.
                if best is None and all_seen:
                    _early = self._best_match_loose(title, all_seen, min_title_score=88.0)
                    if _early:
                        _has_doi_e = bool(self._doi(pub))
                        oa_map[idx] = _early
                        method[idx] = "titulo_fallback_doi_verificar" if _has_doi_e else "titulo_verificar"
                        logger.debug(
                            f"[OpenAlexEnricher]   ~ Early exit '{title[:55]}' "
                            f"(título ≥88% en pool R1/R2, R3+R4 omitidos)"
                        )
                        time.sleep(0.1)
                        continue

                # Intento 3: búsqueda web (.search) + año
                # Equiv. al buscador web de OpenAlex; usa title_clean (minúsculas, con tildes).
                if best is None:
                    q3 = self._Works().select(self._SELECT).search(title_clean)
                    if year:
                        q3 = q3.filter(publication_year=year)
                    candidates3 = q3.get(per_page=15)
                    all_seen.extend(candidates3)
                    best = self._best_match(title, candidates3, year)
                    if best:
                        time.sleep(0.1)

                # Intento 4: búsqueda web SIN filtro de año
                # Solo si R3 ya devolvió candidatos (el artículo existe en OA
                # pero el año differía). Si R3 retornó vacío, R4 tampoco
                # encontrará nada para este título.
                if best is None and year and candidates3:
                    candidates4 = (
                        self._Works()
                        .select(self._SELECT)
                        .search(title_clean)
                    ).get(per_page=15)
                    all_seen.extend(candidates4)
                    best = self._best_match(title, candidates4, year)
                    if best:
                        time.sleep(0.1)

                has_doi = bool(self._doi(pub))
                if best:
                    oa_map[idx] = best
                    method[idx] = "titulo_fallback_doi" if has_doi else "titulo"
                    logger.debug(
                        f"[OpenAlexEnricher]   ✓ '{title[:60]}' → encontrado "
                        f"({'fallback' if has_doi else 'sin doi'})"
                    )
                else:
                    # Verificación: título muy similar pero algo difiere (p. ej. año)
                    best_v = self._best_match_loose(title, all_seen)
                    if best_v:
                        oa_map[idx] = best_v
                        method[idx] = "titulo_fallback_doi_verificar" if has_doi else "titulo_verificar"
                        logger.debug(
                            f"[OpenAlexEnricher]   ~ '{title[:60]}' → VERIFICAR "
                            f"(título similar, datos difieren)"
                        )

                # Guardar en cache (incluso si no se encontró — evita rellamar)
                if hasattr(self, "_search_cache"):
                    self._search_cache[_cache_key] = (oa_map[idx], method[idx])

                time.sleep(0.15)

            except OpenAlexRateLimitError:
                raise
            except (socket.timeout, Exception) as exc:
                retry_after = extract_retry_after(exc)
                if retry_after is not None:
                    raise OpenAlexRateLimitError(retry_after)
                logger.debug(
                    f"[OpenAlexEnricher] Búsqueda título falló "
                    f"'{title[:50]}': {type(exc).__name__}: {exc}"
                )

    def _enrich_by_issn(
        self,
        oa_map:  list,
        method:  list,
        indexed: list[tuple[int, dict]],
    ) -> None:
        """
        Busca publicaciones filtrando por ISSN de la revista + texto del título.
        Útil cuando el DOI no existe o no coincide pero el ISSN es conocido.

        Estrategia:
          1. primary_location.source.issn + título + año
          2. primary_location.source.issn + título (sin año)
          3. locations.source.issn + título (fuente en localización secundaria)
          4. Itera sobre todos los ISSNs del registro (input puede tener varios)
        """
        total = len(indexed)
        for seq, (idx, pub) in enumerate(indexed, start=1):
            issns = self._issns(pub)   # lista de ISSNs normalizados
            title = self._title(pub)
            year  = self._year(pub)
            if not issns or not title:
                continue
            if seq % 25 == 1:
                logger.info(
                    f"[OpenAlexEnricher] ISSN fallback: {seq}/{total}"
                )
            title_norm  = self._normalize_title(title)
            title_clean = self._sanitize_title(title)  # minúsculas + sin ¿¡ iniciales
            best: dict | None = None
            matched_issn = ""

            for issn_raw in issns:
                if best:
                    break
                issn_clean = self._normalize_issn(issn_raw)
                if not issn_clean:
                    continue
                try:
                    # Intento A: primary_location + año
                    q = (
                        self._Works()
                        .select(self._SELECT)
                        .filter(**{"primary_location.source.issn": issn_clean})
                        .search_filter(title=title_clean)
                    )
                    if year:
                        q = q.filter(publication_year=year)
                    candidates_a = q.get(per_page=10)
                    best = self._best_match(title, candidates_a, year, min_score=self.MIN_SCORE_ISSN)

                    # Intento B: primary_location sin año — title_clean con tildes
                    candidates_b: list = []
                    if best is None and year:
                        q_ny = (
                            self._Works()
                            .select(self._SELECT)
                            .filter(**{"primary_location.source.issn": issn_clean})
                            .search_filter(title=title_clean)
                        )
                        candidates_b = q_ny.get(per_page=10)
                        best = self._best_match(title, candidates_b, year, min_score=self.MIN_SCORE_ISSN)
                        if best:
                            time.sleep(0.1)

                    # ── Salida temprana A+B: si el pool ya tiene candidato loose
                    # ── no hacer C/C_ny (locations.issn rara vez añade algo mejor)
                    if best is None and (candidates_a or candidates_b):
                        _pool_ab = list(candidates_a) + list(candidates_b)
                        _early_v = self._best_match_loose(title, _pool_ab)
                        if _early_v:
                            oa_map[idx] = _early_v
                            method[idx] = "issn_verificar"
                            logger.debug(
                                f"[OpenAlexEnricher] ~ ISSN early exit '{issn_clean}' "
                                f"'{title[:50]}' (pool A+B, C/C_ny omitidos)"
                            )
                            break  # no iterar más ISSNs ni hacer C/C_ny

                    # Intento C: locations.source.issn (cubre preprints / versiones secundarias)
                    candidates_c: list = []
                    if best is None:
                        q_loc = (
                            self._Works()
                            .select(self._SELECT)
                            .filter(**{"locations.source.issn": issn_clean})
                            .search_filter(title=title_clean)
                        )
                        if year:
                            q_loc = q_loc.filter(publication_year=year)
                        candidates_c = q_loc.get(per_page=10)
                        best = self._best_match(title, candidates_c, year, min_score=self.MIN_SCORE_ISSN)
                        # C_ny (sin año) se omite: el early-exit A+B ya cubre
                        # coincidencias loose; añadir una 4a llamada por ISSN
                        # rara vez mejora el resultado y eleva el costo.
                        if best:
                            time.sleep(0.1)

                    # Sin match estricto → verificación con pool A+B+C
                    if best is None:
                        all_issn_seen = list(candidates_a) + list(candidates_b) + list(candidates_c)
                        best_v = self._best_match_loose(title, all_issn_seen)
                        if best_v:
                            oa_map[idx] = best_v
                            method[idx] = "issn_verificar"
                            logger.debug(
                                f"[OpenAlexEnricher] ~ ISSN VERIFICAR '{issn_clean}' "
                                f"'{title[:50]}'"
                            )
                            break  # no iterar más ISSNs

                    if best:
                        matched_issn = issn_clean

                    time.sleep(0.12)

                except OpenAlexRateLimitError:
                    raise
                except Exception as exc:
                    retry_after = extract_retry_after(exc)
                    if retry_after is not None:
                        raise OpenAlexRateLimitError(retry_after)
                    logger.debug(
                        f"[OpenAlexEnricher] ISSN fallback falló "
                        f"issn={issn_clean} '{title[:50]}': {exc}"
                    )

            if best:
                oa_map[idx] = best
                method[idx] = "issn"
                logger.debug(
                    f"[OpenAlexEnricher] ✓ ISSN match '{matched_issn}' "
                    f"'{title[:55]}'"
                )

            time.sleep(0.05)

    def _best_match(
        self,
        query_title: str,
        candidates: list,
        year: int | None = None,
        *,
        min_score: float | None = None,
    ) -> dict | None:
        """
        Elige el candidato con mayor score compuesto título + año.

        Score compuesto = título × 0.85 + año × 0.15
          - título : token_sort_ratio máximo entre versión raw y normalizada
          - año    : 100 si mismo año | 60 si ±1 | 25 si ±2 | 10 si ±3 | 0 si más

        Umbrales:
          - min_score (kwarg)  → usa ese valor (p. ej. MIN_SCORE_ISSN para ISSN)
          - sin kwarg          → usa MIN_SCORE (80%)
        """
        if not candidates:
            return None
        try:
            from rapidfuzz import fuzz
        except ImportError:
            return dict(candidates[0]) if candidates else None

        threshold  = min_score if min_score is not None else self.MIN_SCORE
        best_score = 0.0
        best_work  = None
        qt_lower   = query_title.lower()
        qt_norm    = self._normalize_title(query_title)

        for work in candidates:
            cand_raw  = str(work.get("title") or "").lower()
            cand_norm = self._normalize_title(str(work.get("title") or ""))

            title_score = max(
                fuzz.token_sort_ratio(qt_lower, cand_raw),
                fuzz.token_sort_ratio(qt_norm,  cand_norm),
            )

            # Año: tolerancia ampliada para cubrir online-first, ahead-of-print,
            # diferencias entre versión preprint y publicación definitiva
            work_year = work.get("publication_year")
            if year and work_year:
                diff = abs(int(work_year) - int(year))
                if diff == 0:
                    year_score = 100
                elif diff == 1:
                    year_score = 60
                elif diff == 2:
                    year_score = 25
                elif diff == 3:
                    year_score = 10
                else:
                    year_score = 0
            else:
                year_score = 0

            composite = title_score * 0.85 + year_score * 0.15

            if composite > best_score:
                best_score = composite
                best_work  = work

        if best_score >= threshold:
            best_title_score = max(
                fuzz.token_sort_ratio(qt_lower, str(best_work.get("title") or "").lower()),
                fuzz.token_sort_ratio(qt_norm, self._normalize_title(str(best_work.get("title") or ""))),
            )
            best_year = best_work.get("publication_year")
            logger.debug(
                f"[OpenAlexEnricher] Match OK (umbral {threshold:.0f}%) — "
                f"título: {best_title_score:.0f}% | "
                f"año: {best_year} {'✓' if year and best_year and int(best_year)==int(year) else '~'} | "
                f"compuesto: {best_score:.1f}%"
            )
            return dict(best_work)

        logger.debug(
            f"[OpenAlexEnricher] Sin match — "
            f"mejor compuesto: {best_score:.1f}% < umbral {threshold:.0f}%"
        )
        return None

    def _enrich_by_source_name(
        self,
        oa_map:  list,
        method:  list,
        indexed: list[tuple[int, dict]],
    ) -> None:
        """
        Busca por nombre de revista (columna 'revista') + título.
        Se activa cuando hay revista pero no hay ISSN en el input.

        Estrategia:
          1. Filtro por primary_location.source.display_name.search + título + año
          2. Ídem sin año
        """
        total = len(indexed)
        for seq, (idx, pub) in enumerate(indexed, start=1):
            revista = self._revista(pub)
            title   = self._title(pub)
            year    = self._year(pub)
            if not revista or not title:
                continue
            if seq % 25 == 1:
                logger.info(
                    f"[OpenAlexEnricher] Fallback revista: {seq}/{total}"
                )
            title_clean = self._sanitize_title(title)  # minúsculas + sin ¿¡
            try:
                q = (
                    self._Works()
                    .select(self._SELECT)
                    .filter(**{"primary_location.source.display_name.search": revista})
                    .search_filter(title=title_clean)
                )
                if year:
                    q = q.filter(publication_year=year)
                candidates = q.get(per_page=10)
                best = self._best_match(
                    title, candidates, year,
                    min_score=self.MIN_SCORE_SOURCE,
                )

                if best is None and year:
                    q_ny = (
                        self._Works()
                        .select(self._SELECT)
                        .filter(**{"primary_location.source.display_name.search": revista})
                        .search_filter(title=title_clean)
                    )
                    candidates_ny = q_ny.get(per_page=10)
                    best = self._best_match(
                        title, candidates_ny, year,
                        min_score=self.MIN_SCORE_SOURCE,
                    )
                    if best:
                        time.sleep(0.1)

                if best:
                    oa_map[idx] = best
                    method[idx] = "revista"
                    logger.debug(
                        f"[OpenAlexEnricher] ✓ Revista match '{revista[:40]}' "
                        f"'{title[:50]}'"
                    )
                else:
                    all_c = list(candidates) + (list(candidates_ny) if best is None and year else [])
                    best_v = self._best_match_loose(title, all_c)
                    if best_v:
                        oa_map[idx] = best_v
                        method[idx] = "revista_verificar"
                        logger.debug(
                            f"[OpenAlexEnricher] ~ Revista VERIFICAR '{revista[:35]}' '{title[:45]}'"
                        )

                time.sleep(0.15)

            except OpenAlexRateLimitError:
                raise
            except Exception as exc:
                retry_after = extract_retry_after(exc)
                if retry_after is not None:
                    raise OpenAlexRateLimitError(retry_after)
                logger.debug(
                    f"[OpenAlexEnricher] Fallback revista falló "
                    f"revista='{revista[:30]}' '{title[:40]}': {exc}"
                )

    def _enrich_by_title_only(
        self,
        oa_map:  list,
        method:  list,
        indexed: list[tuple[int, dict]],
    ) -> None:
        """
        Último recurso: búsqueda únicamente por título sin ningún filtro adicional.
        Usa un umbral alto (88%) para evitar falsos positivos.
        """
        total = len(indexed)
        for seq, (idx, pub) in enumerate(indexed, start=1):
            title = self._title(pub)
            year  = self._year(pub)
            if not title:
                continue
            if seq % 25 == 1:
                logger.info(
                    f"[OpenAlexEnricher] Último recurso (solo título): {seq}/{total}"
                )
            title_norm  = self._normalize_title(title)
            title_clean = self._sanitize_title(title)  # minúsculas + sin ¿¡
            title_short = self._truncate_title_for_search(title)
            try:
                # Solo título limpio (minúsculas, con tildes), sin filtros
                candidates = (
                    self._Works()
                    .select(self._SELECT)
                    .search_filter(title=title_clean)
                ).get(per_page=15)
                best = self._best_match(title, candidates, year, min_score=88.0)

                # Con título corto si el largo no funcionó
                # Omitir si c1 ya devolvió ≥8 candidatos: el artículo no tiene
                # mejor posicionamiento con título truncado.
                candidates2: list = []
                if best is None and title_short and title_short != title_norm and len(candidates) < 8:
                    candidates2 = (
                        self._Works()
                        .select(self._SELECT)
                        .search_filter(title=title_short)
                    ).get(per_page=15)
                    best = self._best_match(title, candidates2, year, min_score=88.0)

                # Búsqueda general (title+abstract) — equiv. al buscador web de OpenAlex
                # Usa title_clean (minúsculas + tildes)
                candidates3: list = []
                if best is None:
                    candidates3 = (
                        self._Works()
                        .select(self._SELECT)
                        .search(title_clean)
                    ).get(per_page=15)
                    best = self._best_match(title, candidates3, year, min_score=88.0)

                if best:
                    oa_map[idx] = best
                    method[idx] = "titulo_solo"
                    logger.debug(
                        f"[OpenAlexEnricher] ✓ Último recurso '{title[:60]}'"
                    )
                else:
                    # Verificación: título muy similar aunque no alcanza umbral estricto
                    all_c = list(candidates) + list(candidates2) + list(candidates3)
                    best_v = self._best_match_loose(title, all_c)
                    if best_v:
                        oa_map[idx] = best_v
                        method[idx] = "titulo_solo_verificar"
                        logger.debug(
                            f"[OpenAlexEnricher] ~ Último recurso VERIFICAR '{title[:55]}'"
                        )

                time.sleep(0.2)

            except OpenAlexRateLimitError:
                raise
            except Exception as exc:
                retry_after = extract_retry_after(exc)
                if retry_after is not None:
                    raise OpenAlexRateLimitError(retry_after)
                logger.debug(
                    f"[OpenAlexEnricher] Último recurso falló '{title[:50]}': {exc}"
                )

    # Stopwords que no aportan discriminación en búsqueda de títulos
    _TITLE_STOPWORDS = frozenset({
        "de", "del", "la", "el", "los", "las", "un", "una", "unos", "unas",
        "en", "y", "o", "a", "para", "por", "con", "sin", "sobre", "entre",
        "the", "a", "an", "of", "in", "and", "or", "for", "to", "is", "are",
        "its", "with", "from", "at", "by", "on", "as",
    })

    @classmethod
    def _normalize_title(cls, title: str) -> str:
        """
        Normalización profunda para mejorar el matching de títulos:
          1. Lowercase.
          2. Elimina tildes y diacríticos (NFKD decomposition).
          3. Elimina puntuación (mantiene alfanuméricos y espacios).
          4. Colapsa espacios múltiples.
          5. Elimina stopwords triviales.
        """
        import unicodedata, re
        # Paso 1-2: lowercase + quitar diacríticos
        nfkd = unicodedata.normalize("NFKD", title.lower())
        no_accents = "".join(c for c in nfkd if not unicodedata.combining(c))
        # Paso 3: reemplazar puntuación por espacio
        no_punct = re.sub(r"[^\w\s]", " ", no_accents)
        # Paso 4: colapsar espacios
        clean = re.sub(r"\s+", " ", no_punct).strip()
        # Paso 5: eliminar stopwords
        tokens = [w for w in clean.split() if w not in cls._TITLE_STOPWORDS]
        return " ".join(tokens)

    @staticmethod
    def _sanitize_title(title: str) -> str:
        """
        Prepara el título para enviarlo a la API de OpenAlex:
          - Convierte a minúsculas (OpenAlex es case-insensitive pero ALL-CAPS
            puede generar ruido en la búsqueda de relevancia).
          - Elimina los signos de apertura españoles ¿ e ¡ que anteceden al
            título — OpenAlex no los indexa al comienzo del título.
          - Conserva tildes y acentos para maximizar la precisión del índice.
        Ejemplo:
          '¿LA RELACIÓN UNIVERSIDAD EMPRESA...' → 'la relación universidad empresa...'
        """
        return title.strip().lstrip("¿¡").strip().lower()

    def _best_match_loose(
        self,
        query_title: str,
        candidates: list,
        min_title_score: float = 78.0,
    ) -> dict | None:
        """
        Busca el candidato con mayor similitud de título IGNORANDO el año.
        Solo se acepta si el token_sort_ratio del título supera min_title_score.

        Se usa para la hoja 'Verificación manual': el artículo existe en OpenAlex
        con título muy similar pero algún metadato difiere (año, subtítulo, etc.).
        """
        if not candidates:
            return None
        try:
            from rapidfuzz import fuzz
        except ImportError:
            return dict(candidates[0]) if candidates else None

        qt_lower = query_title.lower()
        qt_norm  = self._normalize_title(query_title)
        best_score = 0.0
        best_work  = None

        for work in candidates:
            cand_raw  = str(work.get("title") or "").lower()
            cand_norm = self._normalize_title(str(work.get("title") or ""))
            score = max(
                fuzz.token_sort_ratio(qt_lower, cand_raw),
                fuzz.token_sort_ratio(qt_norm,  cand_norm),
            )
            if score > best_score:
                best_score = score
                best_work  = work

        if best_score >= min_title_score:
            logger.debug(
                f"[OpenAlexEnricher] Verificar match "
                f"(título: {best_score:.0f}%) '{query_title[:55]}'"
            )
            return dict(best_work)
        return None

    @classmethod
    def _truncate_title_for_search(cls, title: str, max_words: int = 10) -> str:
        """
        Devuelve las primeras `max_words` palabras de contenido del título
        (tras normalización y eliminación de stopwords).
        Útil para títulos muy largos o con diferencias al final.
        """
        normalized = cls._normalize_title(title)
        words = normalized.split()
        if len(words) <= max_words:
            return normalized
        return " ".join(words[:max_words])

    @staticmethod
    def _normalize_issn(raw: str) -> str:
        """Normaliza un ISSN al formato XXXX-XXXX. Devuelve '' si no es válido."""
        import re
        clean = re.sub(r"[^0-9Xx]", "", raw).upper()
        if len(clean) == 8:
            return clean[:4] + "-" + clean[4:]
        # ya viene con guión
        m = re.match(r"^(\d{4}-[\dX]{4})$", raw.strip().upper())
        return m.group(1) if m else ""

    # ── Internos: transformación ─────────────────────────────────────────────

    def _flatten(self, orig: dict, oa: dict | None, met: str | None) -> dict:
        """Combina el dict original del Excel con los campos OpenAlex (prefijo oa_)."""
        out = dict(orig)

        _empty_oa = {
            "oa_encontrado":  False,
            "oa_confianza":   None,
            "oa_metodo":      None,
            "oa_work_id":     None,
            "oa_titulo":      None,
            "oa_año":         None,
            "oa_doi":         None,
            "oa_tipo":        None,
            "oa_revista":     None,
            "oa_issn":        None,
            "oa_issn_todos":  None,
            "oa_editorial":   None,
            "oa_open_access": None,
            "oa_status_oa":   None,
            "oa_citas":       None,
            "oa_idioma":      None,
            "oa_url":         None,
            "oa_autores":                None,
            "oa_autor_institucional":     None,
            "oa_autores_institucionales": None,
        }
        if not oa:
            out.update(_empty_oa)
            return out

        primary_loc = oa.get("primary_location") or {}
        source      = primary_loc.get("source") or {}
        open_access = oa.get("open_access") or {}

        authors_txt = "; ".join(
            str((a.get("author") or {}).get("display_name") or "")
            for a in (oa.get("authorships") or [])
        ).strip("; ")

        # Autores con afiliación institucional (via ROR)
        _ror = institution.ror_id
        inst_names: list[str] = []
        for _a in (oa.get("authorships") or []):
            for _inst in (_a.get("institutions") or []):
                if _inst.get("ror") == _ror:
                    _name = (_a.get("author") or {}).get("display_name") or ""
                    if _name:
                        inst_names.append(_name)
                    break
        total_authors = len(oa.get("authorships") or [])
        n_inst = len(inst_names)
        if n_inst > 0:
            # Al menos un autor con afiliación directa via ROR → institucional confirmado
            has_inst: bool | str = True
        elif total_authors > 0:
            # Hay autores en OA pero ninguno tiene nuestro ROR → revisar (OA puede tener
            # datos de afiliación incompletos o desactualizados)
            has_inst = "verificar"
        else:
            # Sin datos de autoría en OA en absoluto
            has_inst = False

        es_verificar = bool(met and met.endswith("_verificar"))
        out.update({
            "oa_encontrado":  True,
            "oa_confianza":   "verificar" if es_verificar else "confirmado",
            "oa_metodo":      met,
            "oa_work_id":     oa.get("id"),
            "oa_titulo":      oa.get("title"),
            "oa_año":         oa.get("publication_year"),
            "oa_doi":         normalize_doi(str(oa.get("doi") or "")),
            "oa_tipo":        oa.get("type"),
            "oa_revista":     source.get("display_name"),
            "oa_issn":        source.get("issn_l"),
            "oa_issn_todos":  "; ".join(source.get("issn") or []) or None,
            "oa_editorial":   source.get("host_organization_name"),
            "oa_open_access": open_access.get("is_oa"),
            "oa_status_oa":   open_access.get("oa_status"),
            "oa_citas":       oa.get("cited_by_count", 0),
            "oa_idioma":      oa.get("language"),
            "oa_url": (
                primary_loc.get("landing_page_url")
                or oa.get("doi")
                or oa.get("id")
            ),
            "oa_autores": authors_txt or None,
            "oa_autor_institucional":     has_inst,
            "oa_autores_institucionales": "; ".join(inst_names) or None,
        })
        return out

    # ── Internos: lectura de Excel ───────────────────────────────────────────

    _COL_ALIASES: dict[str, list[str]] = {
        "titulo":          ["título", "titulo", "title", "Título", "Title", "TÍTULO", "TITLE",
                             "nombre del producto", "nombre producto"],
        "año":             ["año", "Año", "year", "Year", "anio", "Anio", "AÑO", "YEAR",
                             "año de publicación", "año publicacion"],
        "doi":             ["doi", "DOI", "Doi"],
        "issn":            ["issn", "ISSN", "Issn", "issn-l", "issn_l", "ISSN-L",
                             "e-issn", "eissn"],
        "revista":         ["revista", "Revista", "journal", "Journal", "JOURNAL",
                             "fuente", "source", "Source title", "Source Title",
                             "nombre revista"],
        "pmid":            ["pmid", "PMID", "Pmid", "pubmed id", "PubMed ID"],
        "tipologia":       ["tipología minciencias", "tipologia minciencias",
                             "Tipología Minciencias", "tipologia", "Tipología"],
        "tipo_documento":  ["tipo documento", "Tipo Documento", "tipo_documento",
                             "document type", "Document Type", "tipo"],
        "topico":          ["tópico primario", "topico primario", "Tópico Primario",
                             "primary topic", "Primary Topic", "topico"],
    }

    @classmethod
    def _map_headers(cls, raw_headers: list[str]) -> dict[str, str]:
        """Devuelve {header_original → clave_interna}."""
        col_map: dict[str, str] = {}
        for orig in raw_headers:
            norm = orig.strip().lower()
            for internal, aliases in cls._COL_ALIASES.items():
                if orig in aliases or norm in [a.lower() for a in aliases]:
                    col_map[orig] = internal
                    break
        missing = [k for k in ("titulo", "año", "doi") if k not in col_map.values()]
        if missing:
            logger.warning(
                f"[OpenAlexEnricher] Columnas no detectadas: {missing}. "
                f"Encabezados encontrados: {raw_headers}"
            )
        return col_map

    @classmethod
    def _parse_ws(cls, ws) -> list[dict]:
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise ValueError("El archivo Excel está vacío.")

        raw_headers = [
            str(c).strip() if c is not None else f"Col_{i}"
            for i, c in enumerate(all_rows[0])
        ]
        col_map = cls._map_headers(raw_headers)

        rows: list[dict] = []
        for raw_row in all_rows[1:]:
            if all(c is None for c in raw_row):
                continue
            row_dict: dict = {}
            for orig_h, cell_val in zip(raw_headers, raw_row):
                key = col_map.get(orig_h, orig_h)
                row_dict[key] = cell_val if cell_val is not None else ""
            rows.append(row_dict)

        logger.info(
            f"[OpenAlexEnricher] Excel leído: {len(rows)} filas, "
            f"columnas detectadas: {list(col_map.values())}"
        )
        return rows

    def _read_excel_path(self, path: Path) -> list[dict]:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl requerido: pip install openpyxl")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = self._parse_ws(ws)
        wb.close()
        return rows

    def _read_excel_bytes(self, file_bytes: bytes) -> list[dict]:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl requerido: pip install openpyxl")
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = self._parse_ws(ws)
        wb.close()
        return rows

    # ── Internos: helpers de extracción ─────────────────────────────────────

    def _doi(self, pub: dict) -> str:
        raw = str(
            pub.get("doi") or pub.get("DOI") or pub.get("Doi") or ""
        ).strip()
        return normalize_doi(raw) if raw else ""

    def _title(self, pub: dict) -> str:
        return str(
            pub.get("titulo") or pub.get("Título") or pub.get("título") or
            pub.get("title")  or pub.get("Title")  or ""
        ).strip()

    def _year(self, pub: dict) -> int | None:
        raw = (
            pub.get("año") or pub.get("Año") or
            pub.get("year") or pub.get("Year")
        )
        try:
            return int(raw) if raw else None
        except (ValueError, TypeError):
            return None

    def _issn(self, pub: dict) -> str:
        """Extrae el primer ISSN válido del dict (compatibilidad con código antiguo)."""
        lst = self._issns(pub)
        return lst[0] if lst else ""

    def _issns(self, pub: dict) -> list[str]:
        """
        Extrae todos los ISSNs presentes en el registro.
        El campo puede contener varios separados por ';', ',' o espacio.
        """
        raw = str(pub.get("issn") or "").strip()
        if not raw or raw in ("None", "nan", ""):
            return []
        import re
        parts = re.split(r"[;,\s]+", raw)
        result = []
        for p in parts:
            n = self._normalize_issn(p.strip())
            if n and n not in result:
                result.append(n)
        return result

    def _revista(self, pub: dict) -> str:
        """Extrae el nombre de la revista del registro."""
        raw = str(
            pub.get("revista") or pub.get("Revista") or
            pub.get("journal") or pub.get("Journal") or ""
        ).strip()
        return raw if raw not in ("", "None", "nan") else ""








