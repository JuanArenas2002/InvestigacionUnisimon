import io
import logging
from pathlib import Path


COL_ALIASES: dict[str, list[str]] = {
    "titulo": [
        "título", "titulo", "title", "Título", "Title", "TÍTULO", "TITLE",
        "nombre del producto", "nombre producto",
    ],
    "año": [
        "año", "Año", "year", "Year", "anio", "Anio", "AÑO", "YEAR",
        "año de publicación", "año publicacion",
    ],
    "doi": ["doi", "DOI", "Doi"],
    "issn": [
        "issn", "ISSN", "Issn", "issn-l", "issn_l", "ISSN-L",
        "e-issn", "eissn",
    ],
    "revista": [
        "revista", "Revista", "journal", "Journal", "JOURNAL",
        "fuente", "source", "Source title", "Source Title", "nombre revista",
    ],
    "pmid": ["pmid", "PMID", "Pmid", "pubmed id", "PubMed ID"],
    "tipologia": [
        "tipología minciencias", "tipologia minciencias",
        "Tipología Minciencias", "tipologia", "Tipología",
    ],
    "tipo_documento": [
        "tipo documento", "Tipo Documento", "tipo_documento",
        "document type", "Document Type", "tipo",
    ],
    "topico": [
        "tópico primario", "topico primario", "Tópico Primario",
        "primary topic", "Primary Topic", "topico",
    ],
}


def map_headers(raw_headers: list[str], logger: logging.Logger) -> dict[str, str]:
    col_map: dict[str, str] = {}
    aliases_lower = {
        internal: {alias.lower() for alias in aliases}
        for internal, aliases in COL_ALIASES.items()
    }
    for original in raw_headers:
        normalized = original.strip().lower()
        for internal, aliases in COL_ALIASES.items():
            if original in aliases or normalized in aliases_lower[internal]:
                col_map[original] = internal
                break
    missing = [key for key in ("titulo", "año", "doi") if key not in col_map.values()]
    if missing:
        logger.warning(
            f"[OpenAlexEnricher] Columnas no detectadas: {missing}. "
            f"Encabezados encontrados: {raw_headers}"
        )
    return col_map


def parse_worksheet(ws, logger: logging.Logger) -> list[dict]:
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise ValueError("El archivo Excel está vacío.")

    raw_headers = [
        str(cell).strip() if cell is not None else f"Col_{index}"
        for index, cell in enumerate(all_rows[0])
    ]
    col_map = map_headers(raw_headers, logger)

    rows: list[dict] = []
    for raw_row in all_rows[1:]:
        if all(cell is None for cell in raw_row):
            continue
        row_dict: dict = {}
        for original, cell_value in zip(raw_headers, raw_row):
            key = col_map.get(original, original)
            row_dict[key] = cell_value if cell_value is not None else ""
        rows.append(row_dict)

    logger.info(
        f"[OpenAlexEnricher] Excel leído: {len(rows)} filas, "
        f"columnas detectadas: {list(col_map.values())}"
    )
    return rows


def read_excel_path(path: Path, logger: logging.Logger) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl requerido: pip install openpyxl")
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = parse_worksheet(worksheet, logger)
    workbook.close()
    return rows


def read_excel_bytes(file_bytes: bytes, logger: logging.Logger) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl requerido: pip install openpyxl")
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    rows = parse_worksheet(worksheet, logger)
    workbook.close()
    return rows