"""
Extractor de OpenAlex API.

Refactorizado para usar la interfaz BaseExtractor y producir StandardRecord.
Usa PyAlex como cliente HTTP — maneja paginación con cursor, retry,
backoff exponencial y polite pool automáticamente.
"""

import io
import json
import logging
import time
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

import pyalex
from pyalex import Works

from config import openalex_config, institution, SourceName
from extractors.base import (
    BaseExtractor,
    StandardRecord,
    normalize_doi,
    normalize_year,
    normalize_author_name,
)

logger = logging.getLogger(__name__)


class OpenAlexAPIError(Exception):
    """Excepción para errores de la API de OpenAlex"""
    pass


class OpenAlexExtractor(BaseExtractor):
    """
    Extractor de publicaciones desde OpenAlex API (vía PyAlex).

    PyAlex se encarga de:
      - cursor-based pagination (paginate())
      - retry con backoff exponencial
      - polite pool (mailto)
      - rate limiting automático
    """

    source_name = SourceName.OPENALEX

    def __init__(
        self,
        ror_id: str = None,
        email: str = None,
        max_retries: int = None,
    ):
        self.ror_id = self._validate_ror_id(ror_id or institution.ror_id)
        self.email  = email or institution.contact_email
        self.config = openalex_config

        # Configurar PyAlex globalmente (polite pool + retry)
        pyalex.config.email                = self.email
        pyalex.config.max_retries          = max_retries or self.config.max_retries
        pyalex.config.retry_backoff_factor = 0.5
        pyalex.config.retry_http_codes     = [429, 500, 502, 503, 504]

        logger.info(f"OpenAlexExtractor inicializado para ROR: {self.ror_id}")

    # ---------------------------------------------------------
    # INTERFAZ BaseExtractor
    # ---------------------------------------------------------

    def extract(
        self,
        year_from: Optional[int] = None,
        year_to:   Optional[int] = None,
        max_results: Optional[int] = None,
        publication_types: Optional[List[str]] = None,
        **kwargs,
    ) -> List[StandardRecord]:
        """
        Extrae publicaciones de OpenAlex filtrando por ROR de la institución.
        PyAlex maneja automáticamente la paginación con cursor.
        """
        self._validate_year_range(year_from, year_to)

        # ── Construir query ───────────────────────────────────────────────────
        query = Works().filter(
            authorships={"institutions": {"ror": self.ror_id}}
        )

        # Filtros de fecha
        if year_from and year_to:
            query = query.filter(publication_year=f"{year_from}-{year_to}")
        elif year_from:
            query = query.filter(from_publication_date=f"{year_from}-01-01")
        elif year_to:
            query = query.filter(to_publication_date=f"{year_to}-12-31")

        # Filtro de tipos de publicación
        if publication_types:
            query = query.filter(type="|".join(publication_types))

        logger.info(
            f"Extrayendo de OpenAlex: "
            f"{year_from or 'inicio'} – {year_to or 'presente'}"
        )

        records: List[StandardRecord] = []
        total_fetched = 0

        try:
            # paginate() usa cursor automáticamente y devuelve un iterador
            for work in query.paginate(
                per_page=self.config.max_per_page,
                n_max=max_results,
            ):
                try:
                    record = self._parse_record(work)
                    records.append(record)
                    total_fetched += 1
                except Exception as e:
                    logger.warning(f"Error parseando work: {e}")
                    continue

                if total_fetched % 200 == 0:
                    logger.info(f"  Extraídos: {total_fetched}")

        except Exception as e:
            raise OpenAlexAPIError(f"Error comunicándose con OpenAlex: {e}")

        logger.info(f"Extracción completa: {total_fetched} registros")
        return self._post_process(records)

    def search_by_doi(self, doi: str) -> Optional[StandardRecord]:
        """
        Busca un work de OpenAlex por DOI y lo devuelve como StandardRecord.
        Retorna None si no existe (404) o si hay error.
        """
        doi_clean = normalize_doi(str(doi or "").strip())
        if not doi_clean:
            return None
        doi_url = (
            doi_clean
            if doi_clean.startswith("https://")
            else f"https://doi.org/{doi_clean}"
        )
        try:
            work = Works()[doi_url]
            return self._parse_record(work)
        except Exception as e:
            logger.debug(f"search_by_doi: DOI {doi_clean!r} → {e}")
            return None

    def _parse_record(self, work: dict) -> StandardRecord:
        """Convierte un work de OpenAlex a StandardRecord"""
        ids_data         = work.get("ids") or {}
        primary_location = work.get("primary_location") or {}
        source_data      = primary_location.get("source") or {}
        open_access      = work.get("open_access") or {}

        all_authors, institutional_authors = self._extract_authors(work)

        url = (
            primary_location.get("landing_page_url")
            or work.get("doi")
            or work.get("id")
        )

        return StandardRecord(
            source_name=self.source_name,
            source_id=work.get("id"),
            doi=work.get("doi"),
            pmid=ids_data.get("pmid"),
            pmcid=ids_data.get("pmcid"),
            title=work.get("title"),
            publication_year=work.get("publication_year"),
            publication_date=work.get("publication_date"),
            publication_type=work.get("type"),
            language=work.get("language"),
            source_journal=source_data.get("display_name"),
            issn=source_data.get("issn_l"),
            is_open_access=open_access.get("is_oa", False),
            oa_status=open_access.get("oa_status"),
            authors=all_authors,
            institutional_authors=institutional_authors,
            citation_count=work.get("cited_by_count", 0),
            url=url,
            raw_data=work,
        )

    # ---------------------------------------------------------
    # LÓGICA INTERNA
    # ---------------------------------------------------------

    @staticmethod
    def _validate_ror_id(ror_id: str) -> str:
        if not ror_id:
            raise ValueError("ROR ID no puede estar vacío")
        if not ror_id.startswith("https://ror.org/"):
            if ror_id.startswith("ror.org/"):
                ror_id = f"https://{ror_id}"
            elif "/" not in ror_id:
                ror_id = f"https://ror.org/{ror_id}"
            else:
                raise ValueError(f"Formato de ROR ID inválido: {ror_id}")
        return ror_id

    @staticmethod
    def _validate_year_range(year_from, year_to):
        current_year = datetime.now().year
        if year_from and (year_from < 1900 or year_from > current_year + 1):
            raise ValueError(f"Año inicial inválido: {year_from}")
        if year_to and (year_to < 1900 or year_to > current_year + 1):
            raise ValueError(f"Año final inválido: {year_to}")
        if year_from and year_to and year_from > year_to:
            raise ValueError(f"year_from ({year_from}) > year_to ({year_to})")

    def _extract_authors(self, work: dict):
        """Retorna (all_authors, institutional_authors) como listas de dicts"""
        all_authors  = []
        institutional = []

        for authorship in work.get("authorships", []):
            author_data = authorship.get("author") or {}
            raw_name    = author_data.get("display_name") or ""
            clean_name  = normalize_author_name(raw_name)

            info = {
                "name":           clean_name,
                "orcid":          author_data.get("orcid"),
                "openalex_id":    author_data.get("id"),
                "is_institutional": False,
            }
            all_authors.append(info)

            for inst in authorship.get("institutions", []):
                if inst.get("ror") == self.ror_id:
                    info["is_institutional"] = True
                    institutional.append(info)
                    break

        return all_authors, institutional

    # ---------------------------------------------------------
    # UTILIDADES
    # ---------------------------------------------------------

    def save_to_json(
        self,
        records: List[StandardRecord],
        filename:   str,
        output_dir: str = "OpenAlexJson",
    ) -> Path:
        """Guarda registros en JSON"""
        path = Path(output_dir)
        path.mkdir(parents=True, exist_ok=True)
        filepath = path / filename

        data = [r.to_dict() for r in records]
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        logger.info(f"Guardado: {filepath}")
        return filepath


# ─────────────────────────────────────────────────────────────────────────────
# ENRIQUECEDOR DE PUBLICACIONES VÍA PYALEX
# ─────────────────────────────────────────────────────────────────────────────

class OpenAlexEnricher:
    """
    Enriquece un listado de publicaciones consultando OpenAlex a través de PyAlex.

    Acepta un Excel con (al menos) las columnas:
        Título  |  Año  |  doi

    Estrategia de búsqueda por orden de prioridad:
      1. Lote DOI  — hasta MAX_BATCH DOIs por petición (filtro OR).
                     Rápido y exacto. Un solo HTTP request por lote.
      2. Sin DOI   — búsqueda full-text por título + filtro por año.
                     Se elige el candidato con mayor similitud usando
                     RapidFuzz (token_sort_ratio, umbral MIN_SCORE %).

    Uso típico::

        enricher = OpenAlexEnricher()

        # Desde un archivo Excel
        rows = enricher.enrich_from_excel("publicaciones.xlsx")

        # Desde bytes (útil dentro de FastAPI)
        rows = enricher.enrich_from_excel_bytes(file_bytes)

        # Guardar resultado
        enricher.save_to_excel(rows, "resultado_openalex.xlsx")

    Campos OpenAlex añadidos a cada fila (prefijo ``oa_``):
        oa_encontrado, oa_metodo, oa_work_id, oa_titulo, oa_año, oa_doi,
        oa_tipo, oa_revista, oa_issn, oa_editorial, oa_open_access,
        oa_status_oa, oa_citas, oa_idioma, oa_url, oa_autores
    """

    MAX_BATCH = 50    # DOIs por petición (API soporta >100, usamos 50 por seguridad)
    MIN_SCORE = 80.0  # similitud mínima % para aceptar match por título

    # Campos que se piden a OpenAlex (reduce tamaño de respuesta)
    _SELECT = [
        "id", "doi", "title", "publication_year", "type",
        "primary_location", "open_access", "authorships",
        "cited_by_count", "language",
    ]

    def __init__(self, email: str = None):
        try:
            import pyalex
            from pyalex import Works
            self._pyalex = pyalex
            self._Works  = Works
        except ImportError:
            raise ImportError(
                "PyAlex no está instalado.  Ejecuta:  pip install pyalex"
            )

        self.email = email or institution.contact_email
        pyalex.config.email           = self.email
        pyalex.config.max_retries     = 3
        pyalex.config.retry_backoff_factor = 0.5

        logger.info(f"[OpenAlexEnricher] Iniciado — polite pool: {self.email}")

    # ── API pública ───────────────────────────────────────────────────────────

    def enrich_from_excel(self, file_path) -> list[dict]:
        """
        Lee un Excel (.xlsx) por ruta y devuelve las filas enriquecidas.

        Args:
            file_path: str o Path al archivo .xlsx.

        Returns:
            Lista de dicts con los campos originales + campos ``oa_*``.
        """
        rows = self._read_excel_path(Path(file_path))
        return self.enrich(rows)

    def enrich_from_excel_bytes(self, file_bytes: bytes) -> list[dict]:
        """
        Igual que ``enrich_from_excel`` pero acepta los bytes del archivo
        directamente (útil en endpoints FastAPI con UploadFile).
        """
        rows = self._read_excel_bytes(file_bytes)
        return self.enrich(rows)

    def enrich(self, publications: list[dict]) -> list[dict]:
        """
        Enriquece una lista de dicts que contienen al menos:
            - clave ``titulo`` / ``Título`` / ``title``
            - clave ``año``    / ``Año``    / ``year``
            - clave ``doi``    / ``DOI``

        Returns:
            La misma lista con campos ``oa_*`` añadidos a cada dict.
        """
        n = len(publications)
        logger.info(f"[OpenAlexEnricher] Enriqueciendo {n} publicaciones…")

        oa_map: list[dict | None] = [None] * n
        method: list[str | None]  = [None] * n

        with_doi    = [(i, p) for i, p in enumerate(publications) if self._doi(p)]
        without_doi = [(i, p) for i, p in enumerate(publications) if not self._doi(p)]

        logger.info(
            f"[OpenAlexEnricher] Con DOI: {len(with_doi)} | "
            f"Sin DOI: {len(without_doi)}"
        )

        # 1. Búsqueda por lotes de DOI (precisa y rápida)
        self._enrich_by_doi_batch(oa_map, method, with_doi)

        # 2. Búsqueda por título para los que no tienen DOI
        self._enrich_by_title(oa_map, method, without_doi)

        found = sum(1 for x in oa_map if x is not None)
        logger.info(
            f"[OpenAlexEnricher] Encontrados: {found}/{n} "
            f"({found / n * 100:.1f}%)"
        )

        return [
            self._flatten(pub, oa_map[i], method[i])
            for i, pub in enumerate(publications)
        ]

    def save_to_excel(self, enriched_rows: list[dict], output_path) -> Path:
        """
        Guarda los resultados enriquecidos en un Excel (.xlsx).

        Las columnas originales (Título, Año, doi) van primero; los campos
        OpenAlex (oa_*) van después con encabezado azul diferenciado.

        Args:
            enriched_rows: lista devuelta por ``enrich()`` / ``enrich_from_excel()``.
            output_path:   str o Path donde guardar el .xlsx.

        Returns:
            Path del archivo guardado.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl requerido: pip install openpyxl")

        if not enriched_rows:
            raise ValueError("No hay filas para guardar.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Enriquecimiento OpenAlex"

        headers = list(enriched_rows[0].keys())

        # ── Fila 1: título del reporte ────────────────────────────────────────
        from datetime import datetime as _dt
        _hdr_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
        _oa_fill  = PatternFill(fill_type="solid", fgColor="16537e")
        _hdr_font = Font(bold=True, color="FFFFFF", size=10)

        title_cell = ws.cell(
            row=1, column=1,
            value=(
                f"Enriquecimiento OpenAlex  —  "
                f"{len(enriched_rows)} publicaciones  —  "
                f"Generado: {_dt.now().strftime('%d/%m/%Y %H:%M')}"
            ),
        )
        title_cell.font  = Font(bold=True, size=12, color="FFFFFF")
        title_cell.fill  = _hdr_fill
        ws.row_dimensions[1].height = 22
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

        # ── Fila 2: encabezados ──────────────────────────────────────────────
        for ci, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=ci, value=h)
            cell.fill      = _oa_fill if h.startswith("oa_") else _hdr_fill
            cell.font      = _hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 28

        # ── Filas de datos ────────────────────────────────────────────────────
        _found_fill  = PatternFill(fill_type="solid", fgColor="EBF3FB")   # azul muy claro
        _notfnd_fill = PatternFill(fill_type="solid", fgColor="F5F5F5")   # gris
        _yes_fill    = PatternFill(fill_type="solid", fgColor="C6EFCE")   # verde claro
        _no_fill     = PatternFill(fill_type="solid", fgColor="FFCCCC")   # rojo claro

        for ri, row in enumerate(enriched_rows, start=3):
            found    = row.get("oa_encontrado", False)
            base_fll = _found_fill if found else _notfnd_fill
            for ci, h in enumerate(headers, start=1):
                val  = row.get(h, "")
                if isinstance(val, bool):
                    val = "Sí" if val else "No"
                cell = ws.cell(row=ri, column=ci, value=val if val is not None else "")
                cell.alignment = Alignment(
                    horizontal="left", vertical="center",
                    wrap_text=(h in ("oa_autores", "oa_titulo", "titulo")),
                )
                if h == "oa_encontrado":
                    cell.fill = _yes_fill if found else _no_fill
                    cell.font = Font(bold=True)
                elif h == "oa_open_access":
                    cell.fill = _yes_fill if val == "Sí" else base_fll
                else:
                    cell.fill = base_fll
            ws.row_dimensions[ri].height = 16

        # ── Anchos de columna ─────────────────────────────────────────────────
        _widths = {
            "titulo":        50, "oa_titulo":  50, "oa_autores":   46,
            "doi":           38, "oa_doi":     38, "oa_url":       38,
            "oa_work_id":    34, "oa_revista": 34, "oa_editorial": 28,
            "oa_encontrado": 14, "oa_metodo":  14, "oa_tipo":      16,
            "oa_year":        8, "año":         8, "oa_status_oa": 14,
        }
        for ci, h in enumerate(headers, start=1):
            w = _widths.get(h, 18 if h.startswith("oa_") else 14)
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.freeze_panes = "A3"

        output_path = Path(output_path)
        wb.save(output_path)
        logger.info(f"[OpenAlexEnricher] Excel guardado: {output_path}")
        return output_path

    # ── Internos: búsqueda ────────────────────────────────────────────────────

    def _enrich_by_doi_batch(
        self,
        oa_map:  list,
        method:  list,
        indexed: list[tuple[int, dict]],
    ) -> None:
        """Resuelve DOIs en lotes — una sola petición por lote."""
        for batch_start in range(0, len(indexed), self.MAX_BATCH):
            batch = indexed[batch_start : batch_start + self.MAX_BATCH]

            # doi_normalizado → índice original
            doi_to_idx: dict[str, int] = {}
            doi_full_list: list[str]   = []
            for idx, pub in batch:
                d = self._doi(pub)
                if not d:
                    continue
                full = d if d.startswith("https://") else f"https://doi.org/{d}"
                doi_to_idx[d]   = idx
                doi_full_list.append(full)

            if not doi_full_list:
                continue

            n_lote = len(doi_full_list)
            logger.info(
                f"[OpenAlexEnricher] Lote DOI "
                f"{batch_start + 1}–{batch_start + len(batch)}: "
                f"{n_lote} DOIs"
            )

            try:
                works = (
                    self._Works()
                    .select(self._SELECT)
                    .filter(doi=doi_full_list)
                    .get(per_page=min(n_lote, 200))
                )
                for work in works:
                    w_doi = normalize_doi(str(work.get("doi") or ""))
                    if w_doi in doi_to_idx:
                        i = doi_to_idx[w_doi]
                        oa_map[i] = dict(work)
                        method[i] = "doi"
                time.sleep(0.1)

            except Exception as exc:
                logger.warning(
                    f"[OpenAlexEnricher] Lote DOI falló ({exc}); "
                    "reintentando individualmente…"
                )
                for full_doi, idx in {
                    (d if d.startswith("https://") else f"https://doi.org/{d}"): i
                    for i, pub in batch
                    for d in [self._doi(pub)] if d
                }.items():
                    try:
                        work = self._Works().select(self._SELECT)[full_doi]
                        oa_map[idx] = dict(work)
                        method[idx] = "doi"
                        time.sleep(0.12)
                    except Exception as e2:
                        logger.debug(
                            f"[OpenAlexEnricher] DOI individual no encontrado "
                            f"'{full_doi}': {e2}"
                        )

    def _enrich_by_title(
        self,
        oa_map:  list,
        method:  list,
        indexed: list[tuple[int, dict]],
    ) -> None:
        """Busca por título + año para publicaciones sin DOI."""
        for idx, pub in indexed:
            title = self._title(pub)
            year  = self._year(pub)
            if not title:
                continue
            try:
                query = (
                    self._Works()
                    .select(self._SELECT)
                    .search_filter(title=title)
                )
                if year:
                    query = query.filter(publication_year=year)

                # top 5 por relevancia — suficiente para el fuzzy match
                candidates = query.get(per_page=5)
                best = self._best_match(title, candidates)
                if best:
                    oa_map[idx] = best
                    method[idx] = "titulo"

                time.sleep(0.15)

            except Exception as exc:
                logger.debug(
                    f"[OpenAlexEnricher] Búsqueda título falló "
                    f"'{title[:50]}': {exc}"
                )

    def _best_match(self, query_title: str, candidates: list) -> dict | None:
        """Elige el candidato con mayor similitud (token_sort_ratio ≥ MIN_SCORE %)."""
        if not candidates:
            return None
        try:
            from rapidfuzz import fuzz
        except ImportError:
            # Sin rapidfuzz: tomar el primer resultado
            return dict(candidates[0]) if candidates else None

        best_score = 0.0
        best_work  = None
        qt_lower   = query_title.lower()

        for work in candidates:
            cand = str(work.get("title") or "").lower()
            score = fuzz.token_sort_ratio(qt_lower, cand)
            if score > best_score:
                best_score = score
                best_work  = work

        if best_score >= self.MIN_SCORE:
            return dict(best_work)

        logger.debug(
            f"[OpenAlexEnricher] Mejor match ({best_score:.0f}%) "
            f"< umbral {self.MIN_SCORE:.0f}% → descartado"
        )
        return None

    # ── Internos: transformación ──────────────────────────────────────────────

    def _flatten(self, orig: dict, oa: dict | None, met: str | None) -> dict:
        """Combina el dict original del Excel con los campos OpenAlex (prefijo oa_)."""
        out = dict(orig)

        _empty_oa = {
            "oa_encontrado":  False,
            "oa_metodo":      None,
            "oa_work_id":     None,
            "oa_titulo":      None,
            "oa_año":         None,
            "oa_doi":         None,
            "oa_tipo":        None,
            "oa_revista":     None,
            "oa_issn":        None,
            "oa_issn_todos":  None,
            "oa_editorial":   None,
            "oa_open_access": None,
            "oa_status_oa":   None,
            "oa_citas":       None,
            "oa_idioma":      None,
            "oa_url":         None,
            "oa_autores":     None,
        }
        if not oa:
            out.update(_empty_oa)
            return out

        primary_loc = oa.get("primary_location") or {}
        source      = primary_loc.get("source") or {}
        open_access = oa.get("open_access") or {}

        # Texto de autores: "Apellido1, N1; Apellido2, N2; …"
        authors_txt = "; ".join(
            str((a.get("author") or {}).get("display_name") or "")
            for a in (oa.get("authorships") or [])
        ).strip("; ")

        out.update({
            "oa_encontrado":  True,
            "oa_metodo":      met,
            "oa_work_id":     oa.get("id"),
            "oa_titulo":      oa.get("title"),
            "oa_año":         oa.get("publication_year"),
            "oa_doi":         normalize_doi(str(oa.get("doi") or "")),
            "oa_tipo":        oa.get("type"),
            "oa_revista":     source.get("display_name"),
            "oa_issn":        source.get("issn_l"),
            "oa_issn_todos":  "; ".join(source.get("issn") or []) or None,
            "oa_editorial":   source.get("host_organization_name"),
            "oa_open_access": open_access.get("is_oa"),
            "oa_status_oa":   open_access.get("oa_status"),
            "oa_citas":       oa.get("cited_by_count", 0),
            "oa_idioma":      oa.get("language"),
            "oa_url": (
                primary_loc.get("landing_page_url")
                or oa.get("doi")
                or oa.get("id")
            ),
            "oa_autores": authors_txt or None,
        })
        return out

    # ── Internos: lectura de Excel ────────────────────────────────────────────

    # Variantes de nombre de columna aceptadas → clave interna
    _COL_ALIASES: dict[str, list[str]] = {
        "titulo": ["título", "titulo", "title", "Título", "Title", "TÍTULO", "TITLE"],
        "año":    ["año", "Año", "year", "Year", "anio", "Anio", "AÑO", "YEAR"],
        "doi":    ["doi", "DOI", "Doi"],
    }

    @classmethod
    def _map_headers(cls, raw_headers: list[str]) -> dict[str, str]:
        """Devuelve {header_original → clave_interna}."""
        col_map: dict[str, str] = {}
        for orig in raw_headers:
            norm = orig.strip().lower()
            for internal, aliases in cls._COL_ALIASES.items():
                if orig in aliases or norm in [a.lower() for a in aliases]:
                    col_map[orig] = internal
                    break
        missing = [k for k in ("titulo", "año", "doi") if k not in col_map.values()]
        if missing:
            logger.warning(
                f"[OpenAlexEnricher] Columnas no detectadas: {missing}. "
                f"Encabezados encontrados: {raw_headers}"
            )
        return col_map

    @classmethod
    def _parse_ws(cls, ws) -> list[dict]:
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            raise ValueError("El archivo Excel está vacío.")

        raw_headers = [
            str(c).strip() if c is not None else f"Col_{i}"
            for i, c in enumerate(all_rows[0])
        ]
        col_map = cls._map_headers(raw_headers)

        rows: list[dict] = []
        for raw_row in all_rows[1:]:
            if all(c is None for c in raw_row):
                continue
            row_dict: dict = {}
            for orig_h, cell_val in zip(raw_headers, raw_row):
                # Guardar bajo la clave interna si esta col fue mapeada,
                # de lo contrario bajo el nombre original
                key = col_map.get(orig_h, orig_h)
                row_dict[key] = cell_val if cell_val is not None else ""
            rows.append(row_dict)

        logger.info(
            f"[OpenAlexEnricher] Excel leído: {len(rows)} filas, "
            f"columnas detectadas: {list(col_map.values())}"
        )
        return rows

    def _read_excel_path(self, path: Path) -> list[dict]:
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl requerido: pip install openpyxl")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = self._parse_ws(ws)
        wb.close()
        return rows

    def _read_excel_bytes(self, file_bytes: bytes) -> list[dict]:
        try:
            import io, openpyxl
        except ImportError:
            raise ImportError("openpyxl requerido: pip install openpyxl")
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = self._parse_ws(ws)
        wb.close()
        return rows

    # ── Internos: helpers de extracción ──────────────────────────────────────

    def _doi(self, pub: dict) -> str:
        raw = str(
            pub.get("doi") or pub.get("DOI") or pub.get("Doi") or ""
        ).strip()
        return normalize_doi(raw) if raw else ""

    def _title(self, pub: dict) -> str:
        return str(
            pub.get("titulo") or pub.get("Título") or pub.get("título") or
            pub.get("title")  or pub.get("Title")  or ""
        ).strip()

    def _year(self, pub: dict) -> int | None:
        raw = (
            pub.get("año") or pub.get("Año") or
            pub.get("year") or pub.get("Year")
        )
        try:
            return int(raw) if raw else None
        except (ValueError, TypeError):
            return None
