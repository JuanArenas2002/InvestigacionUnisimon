import requests
import os

# Archivo de prueba
file_path = 'c:/temp/test_authors.xlsx'
api_url = 'http://localhost:8000/api/scopus/author-h-index'

# Hacer petición POST
print("🔄 Enviando archivo al endpoint...")
with open(file_path, 'rb') as f:
    files = {'file': f}
    response = requests.post(api_url, files=files, timeout=60)

# Guardar resultado
if response.status_code == 200:
    with open('c:/temp/h_index_result.xlsx', 'wb') as f:
        f.write(response.content)
    print(f'✅ Respuesta exitosa - {len(response.content)} bytes')
    print(f'✅ Archivo guardado: c:/temp/h_index_result.xlsx')
    
    # Leer el archivo para verificar
    import openpyxl
    wb = openpyxl.load_workbook('c:/temp/h_index_result.xlsx')
    print(f'📊 Hojas disponibles: {wb.sheetnames}')
    
    # Ver primera hoja
    ws = wb.active
    print(f'\n📋 Primeras filas de {ws.title}:')
    for row in range(1, min(6, ws.max_row + 1)):
        values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        print(f"  Fila {row}: {values}")
    
else:
    print(f'❌ Error {response.status_code}')
    print(f'Respuesta: {response.text[:500]}')
